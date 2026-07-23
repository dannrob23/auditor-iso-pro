"""
Módulo de Autodiagnóstico IA-SEC
Evalúa la madurez del modelo de IA mediante un cuestionario estructurado
y genera automáticamente una propuesta de implementación basada en ISO 19011.
"""
import json
import os
import tempfile
from datetime import datetime
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# DIMENSIONES Y PREGUNTAS DEL AUTODIAGNÓSTICO (basado en ISO 19011 + ISO 23894)
# ══════════════════════════════════════════════════════════════════════════════

DIMENSIONES_AUTO = [
    {
        "nombre": "1. Gobernanza de IA",
        "descripcion": "Políticas, liderazgo y estructura organizacional para la IA (ISO 27001 Cláusula 5 / ISO 42001 Cláusula 5 / NIST GOVERN)",
        "preguntas": [
            {
                "id": "GOV-1",
                "texto": "¿Existe una política de IA aprobada por la alta dirección que defina principios éticos, niveles de autonomía y supervisión humana?",
                "opciones": [
                    ("No existe", 0),
                    ("En borrador", 1),
                    ("Existe pero no se aplica", 2),
                    ("Implementada parcialmente", 3),
                    ("Completamente implementada", 4),
                    ("Optimizada con mejora continua", 5),
                ]
            },
            {
                "id": "GOV-2",
                "texto": "¿Hay un responsable formal (CISO de IA / Oficial de IA) con autoridad para detener despliegues de modelos críticos?",
                "opciones": [
                    ("No hay responsable asignado", 0),
                    ("Rol definido pero sin autoridad formal", 1),
                    ("Rol asignado, autoridad limitada", 2),
                    ("Rol con autoridad para detener despliegues", 3),
                    ("Rol con autoridad + comité de ética de IA", 4),
                    ("Estructura de gobierno de IA madura con reporting directo a junta", 5),
                ]
            },
            {
                "id": "GOV-3",
                "texto": "¿Los roles y responsabilidades para sistemas de IA están formalmente definidos e integrados con el SGSI?",
                "opciones": [
                    ("No definidos", 0),
                    ("Definidos informalmente", 1),
                    ("Definidos pero no integrados con SGSI", 2),
                    ("Definidos e integrados parcialmente", 3),
                    ("Completamente integrados con SGSI", 4),
                    ("Integración total con auditoría independiente", 5),
                ]
            }
        ]
    },
    {
        "nombre": "2. Gestión de Riesgos de IA",
        "descripcion": "Identificación, evaluación y tratamiento de riesgos específicos de IA (ISO 27001 Cláusula 6.1 / ISO 42001 Cláusula 6.1 / NIST MAP-MEASURE / ISO 23894)",
        "preguntas": [
            {
                "id": "RIES-1",
                "texto": "¿El proceso de gestión de riesgos de la organización incluye riesgos específicos de IA (opacidad algorítmica, deriva del modelo, ataques adversariales)?",
                "opciones": [
                    ("No se consideran riesgos de IA", 0),
                    ("Identificados pero no evaluados", 1),
                    ("Evaluados informalmente", 2),
                    ("Evaluados con metodología definida", 3),
                    ("Integrados en el proceso de riesgos corporativo", 4),
                    ("Gestionados con monitoreo continuo y métricas cuantitativas", 5),
                ]
            },
            {
                "id": "RIES-2",
                "texto": "¿Se realiza evaluación de impacto de IA (AIF) antes del despliegue de nuevos modelos según ISO 42001 A.2?",
                "opciones": [
                    ("No se realiza", 0),
                    ("Ocasionalmente, sin metodología", 1),
                    ("Para modelos críticos, metodología básica", 2),
                    ("Para todos los modelos, metodología definida", 3),
                    ("Para todos los modelos, con revisión independiente", 4),
                    ("AIF obligatorio + revisión externa periódica", 5),
                ]
            },
            {
                "id": "RIES-3",
                "texto": "¿Se monitorea la deriva del modelo (Model Drift) y la degradación del rendimiento en producción?",
                "opciones": [
                    ("No se monitorea", 0),
                    ("Monitoreo manual esporádico", 1),
                    ("Monitoreo automatizado básico", 2),
                    ("Monitoreo automatizado con alertas", 3),
                    ("Monitoreo con dashboards y KPIs definidos", 4),
                    ("Monitoreo predictivo con acción correctiva automática", 5),
                ]
            },
            {
                "id": "RIES-4",
                "texto": "¿Se realizan pruebas de robustez adversarial a los modelos antes del despliegue?",
                "opciones": [
                    ("No se realizan", 0),
                    ("Pruebas básicas no sistemáticas", 1),
                    ("Pruebas para modelos críticos", 2),
                    ("Pruebas sistemáticas con metodología definida", 3),
                    ("Pruebas automatizadas en pipeline CI/CD", 4),
                    ("Red teaming de IA con periodicidad definida", 5),
                ]
            }
        ]
    },
    {
        "nombre": "3. Ciclo de Vida del Modelo",
        "descripcion": "Desarrollo seguro, validación y despliegue (ISO 27001 A.8.25-8.29 / ISO 42001 A.7 / NIST MAP 3.1-3.5)",
        "preguntas": [
            {
                "id": "CICLO-1",
                "texto": "¿El desarrollo de modelos sigue un proceso MLSecOps con controles de seguridad en cada etapa del pipeline?",
                "opciones": [
                    ("No hay proceso definido", 0),
                    ("Desarrollo informal sin controles de seguridad", 1),
                    ("MLOps básico, sin seguridad integrada", 2),
                    ("MLSecOps con controles en etapas clave", 3),
                    ("MLSecOps automatizado en CI/CD", 4),
                    ("MLSecOps maduro con security gates y validación continua", 5),
                ]
            },
            {
                "id": "CICLO-2",
                "texto": "¿Los datos de entrenamiento se validan para detectar sesgos, envenenamiento y calidad antes de su uso?",
                "opciones": [
                    ("No se validan", 0),
                    ("Validación manual básica", 1),
                    ("Validación automatizada de calidad", 2),
                    ("Validación de calidad + detección de sesgos", 3),
                    ("Validación completa con reportes de equidad", 4),
                    ("Validación continua con federated learning y privacidad diferencial", 5),
                ]
            },
            {
                "id": "CICLO-3",
                "texto": "¿Los modelos se registran con model cards, versionado y firmas criptográficas?",
                "opciones": [
                    ("No se registran", 0),
                    ("Registro manual básico", 1),
                    ("Model registry con versionado", 2),
                    ("Model registry + model cards", 3),
                    ("Model registry + model cards + firmas digitales", 4),
                    ("Model registry inmutable + SBOM de IA + firmas verificables", 5),
                ]
            }
        ]
    },
    {
        "nombre": "4. Seguridad Técnica del Modelo",
        "descripcion": "Protección de modelos, datos y APIs (ISO 27001 A.8.8, A.8.15, A.5.15 / ISO 42001 A.4 / NIST MANAGE 2.2)",
        "preguntas": [
            {
                "id": "TEC-1",
                "texto": "¿Las APIs de inferencia del modelo están protegidas con autenticación, rate limiting y validación de entrada?",
                "opciones": [
                    ("No hay APIs expuestas o no están protegidas", 0),
                    ("Autenticación básica sin rate limiting", 1),
                    ("Autenticación + rate limiting", 2),
                    ("Autenticación + rate limiting + validación de entrada", 3),
                    ("API Gateway con WAF y detección de anomalías", 4),
                    ("Protección completa con Zero Trust y monitoreo adversarial", 5),
                ]
            },
            {
                "id": "TEC-2",
                "texto": "¿Los logs de inferencia del modelo son inmutables y permiten trazabilidad forense completa?",
                "opciones": [
                    ("No se registran logs de inferencia", 0),
                    ("Logs básicos sin protección", 1),
                    ("Logs protegidos contra modificación", 2),
                    ("Logs inmutables con trazabilidad", 3),
                    ("Logs inmutables + correlación con eventos de seguridad", 4),
                    ("Sistema de logging blockchain/WORM con capacidad forense completa", 5),
                ]
            },
            {
                "id": "TEC-3",
                "texto": "¿El modelo y sus datos de entrenamiento están cifrados en reposo y en tránsito?",
                "opciones": [
                    ("No están cifrados", 0),
                    ("Cifrado en tránsito solamente", 1),
                    ("Cifrado en reposo y tránsito básico", 2),
                    ("Cifrado completo con gestión de claves (HSM)", 3),
                    ("Cifrado completo + firma de modelos", 4),
                    ("Cifrado completo + HSM + preparación post-cuántica", 5),
                ]
            },
            {
                "id": "TEC-4",
                "texto": "¿El control de acceso a los pipelines ML sigue el principio de Zero Trust?",
                "opciones": [
                    ("Acceso libre a entornos ML", 0),
                    ("Control de acceso básico por usuario", 1),
                    ("Segregación de entornos dev/staging/prod", 2),
                    ("Zero Trust con MFA para operaciones críticas", 3),
                    ("Zero Trust + aprobación multi-persona para despliegues", 4),
                    ("Zero Trust completo + just-in-time access + auditoría continua", 5),
                ]
            }
        ]
    },
    {
        "nombre": "5. Supervisión Humana y Transparencia",
        "descripcion": "Human-in-the-loop, explicabilidad y comunicación (ISO 27001 Cláusula 7.4, A.6.3 / ISO 42001 A.3 / NIST GOVERN 3.1)",
        "preguntas": [
            {
                "id": "HUM-1",
                "texto": "¿Las decisiones críticas automatizadas tienen un mecanismo formal de supervisión humana (Human-in-the-Loop)?",
                "opciones": [
                    ("Sin supervisión humana", 0),
                    ("Supervisión humana reactiva post-decisión", 1),
                    ("Human-on-the-loop (monitoreo sin intervención directa)", 2),
                    ("Human-in-the-loop para decisiones irreversibles", 3),
                    ("HITL con tiempos de respuesta validados", 4),
                    ("HITL + override documentado + métricas de efectividad", 5),
                ]
            },
            {
                "id": "HUM-2",
                "texto": "¿Los operadores y usuarios finales reciben capacitación sobre las capacidades y limitaciones del sistema de IA?",
                "opciones": [
                    ("No hay capacitación", 0),
                    ("Capacitación básica en el uso del sistema", 1),
                    ("Capacitación sobre capacidades y limitaciones", 2),
                    ("Capacitación + concientización sobre riesgos de IA", 3),
                    ("Programa de formación continua certificado", 4),
                    ("Formación + simulacros de fallo + certificación periódica", 5),
                ]
            },
            {
                "id": "HUM-3",
                "texto": "¿El modelo puede explicar sus decisiones de forma comprensible para los operadores (explicabilidad)?",
                "opciones": [
                    ("Sin capacidad de explicación", 0),
                    ("Explicabilidad básica (feature importance)", 1),
                    ("Explicabilidad con SHAP/LIME para auditoría", 2),
                    ("Explicabilidad en tiempo real para operadores", 3),
                    ("Explicabilidad en tiempo real + post-hoc para reguladores", 4),
                    ("Explicabilidad completa + reportes automáticos de auditoría", 5),
                ]
            },
            {
                "id": "HUM-4",
                "texto": "¿Se informa a los usuarios cuando interactúan con un sistema de IA (transparencia)?",
                "opciones": [
                    ("No se informa", 0),
                    ("Se informa en términos y condiciones", 1),
                    ("Se informa en el punto de interacción", 2),
                    ("Se informa + se explican las capacidades del sistema", 3),
                    ("Se informa + se ofrece canal de apelación", 4),
                    ("Disclosure completo + model cards públicos + canal de apelación", 5),
                ]
            }
        ]
    },
    {
        "nombre": "6. Monitoreo y Mejora Continua",
        "descripcion": "Métricas, auditoría y ciclo de mejora (ISO 27001 Cláusula 9-10 / ISO 42001 Cláusula 9-10 / NIST MEASURE / ISO 19011 Seguimiento)",
        "preguntas": [
            {
                "id": "MON-1",
                "texto": "¿Existe un dashboard de monitoreo continuo con KPIs de rendimiento, deriva, equidad y seguridad del modelo?",
                "opciones": [
                    ("No existe monitoreo", 0),
                    ("Monitoreo manual periódico", 1),
                    ("Dashboard básico con métricas técnicas", 2),
                    ("Dashboard con KPIs de rendimiento y deriva", 3),
                    ("Dashboard integral: rendimiento, deriva, equidad, seguridad", 4),
                    ("Dashboard con alertas predictivas y acción correctiva automatizada", 5),
                ]
            },
            {
                "id": "MON-2",
                "texto": "¿Se realizan auditorías periódicas de los sistemas de IA siguiendo la metodología ISO 19011?",
                "opciones": [
                    ("No se auditan", 0),
                    ("Auditorías reactivas post-incidente", 1),
                    ("Auditorías anuales básicas", 2),
                    ("Auditorías semestrales con checklist definido", 3),
                    ("Auditoría continua con herramientas automatizadas", 4),
                    ("Programa de auditoría maduro: continua + externa + certificable", 5),
                ]
            },
            {
                "id": "MON-3",
                "texto": "¿Existe un plan de continuidad operativa para fallos catastróficos de IA (modo degradado, RTO < 15 min)?",
                "opciones": [
                    ("No existe plan de continuidad", 0),
                    ("Plan básico sin pruebas", 1),
                    ("Plan documentado con RTO definido", 2),
                    ("Plan probado con simulacros anuales", 3),
                    ("Plan probado con simulacros semestrales + modo degradado validado", 4),
                    ("Plan maduro: hot standby, failover automático, simulacros trimestrales", 5),
                ]
            }
        ]
    },
    {
        "nombre": "7. Cadena de Suministro y Terceros",
        "descripcion": "Proveedores de IA, modelos externos y dependencias (ISO 27001 A.5.19-5.23 / ISO 42001 A.8 / NIST GOVERN 6.1-6.2)",
        "preguntas": [
            {
                "id": "CAD-1",
                "texto": "¿Los proveedores de modelos de IA (APIs, modelos pre-entrenados, datasets) son evaluados en seguridad?",
                "opciones": [
                    ("No se evalúan proveedores", 0),
                    ("Evaluación básica sin criterios definidos", 1),
                    ("Evaluación con cuestionario de seguridad", 2),
                    ("Evaluación con auditoría de proveedores", 3),
                    ("AI-SBOM obligatorio para todos los proveedores", 4),
                    ("AI-SBOM + auditoría in situ + cláusulas contractuales de seguridad", 5),
                ]
            },
            {
                "id": "CAD-2",
                "texto": "¿Se tiene un inventario completo de dependencias de IA (librerías, frameworks, modelos base)?",
                "opciones": [
                    ("No hay inventario", 0),
                    ("Inventario manual parcial", 1),
                    ("Inventario automatizado básico", 2),
                    ("SBOM de IA con versionado completo", 3),
                    ("SBOM + escaneo de vulnerabilidades automatizado", 4),
                    ("SBOM + escaneo + parcheo automatizado + vigilancia de amenazas", 5),
                ]
            }
        ]
    }
]


# ══════════════════════════════════════════════════════════════════════════════
# CÁLCULO DE MADUREZ
# ══════════════════════════════════════════════════════════════════════════════

def calcular_madurez(respuestas: dict) -> dict:
    """
    Calcula el nivel de madurez por dimensión y el puntaje general.
    Respuestas: {pregunta_id: valor_seleccionado (0-5)}
    Retorna: {dimensiones, puntajes, nivel_global, descripcion}
    """
    resultados = {}
    puntaje_total = 0
    maximo_total = 0

    for dim in DIMENSIONES_AUTO:
        puntaje_dim = 0
        maximo_dim = 0
        preguntas_dim = []

        for p in dim["preguntas"]:
            valor = respuestas.get(p["id"], 0)
            max_opcion = max(v for _, v in p["opciones"])
            puntaje_dim += valor
            maximo_dim += max_opcion
            preguntas_dim.append({
                "id": p["id"],
                "texto": p["texto"],
                "valor": valor,
                "maximo": max_opcion,
                "porcentaje": round(valor / max_opcion * 100, 1)
            })

        porcentaje_dim = round(puntaje_dim / maximo_dim * 100, 1) if maximo_dim > 0 else 0
        nivel_dim = _nivel_madurez(porcentaje_dim)

        resultados[dim["nombre"]] = {
            "puntaje": puntaje_dim,
            "maximo": maximo_dim,
            "porcentaje": porcentaje_dim,
            "nivel": nivel_dim,
            "preguntas": preguntas_dim
        }

        puntaje_total += puntaje_dim
        maximo_total += maximo_dim

    porcentaje_global = round(puntaje_total / maximo_total * 100, 1) if maximo_total > 0 else 0
    nivel_global = _nivel_madurez(porcentaje_global)

    return {
        "dimensiones": resultados,
        "puntaje_total": puntaje_total,
        "maximo_total": maximo_total,
        "porcentaje_global": porcentaje_global,
        "nivel_global": nivel_global,
        "descripcion_global": _descripcion_global(nivel_global, porcentaje_global)
    }


def _nivel_madurez(porcentaje: float) -> str:
    if porcentaje >= 90:
        return "Optimizado"
    elif porcentaje >= 70:
        return "Gestionado"
    elif porcentaje >= 50:
        return "Definido"
    elif porcentaje >= 30:
        return "Repetible"
    elif porcentaje >= 10:
        return "Inicial"
    else:
        return "Inexistente"


def _descripcion_global(nivel: str, pct: float) -> str:
    desc = {
        "Inexistente": "La organización no ha comenzado a gestionar los riesgos de IA. No hay procesos, roles ni controles definidos. Exposición crítica a incidentes.",
        "Inicial": "La organización reconoce la necesidad de gestionar IA pero lo hace de forma reactiva y dependiente de personas clave. Sin procesos estandarizados.",
        "Repetible": "Existen procesos básicos que se aplican de forma inconsistente. Alta dependencia de personas. La documentación es parcial.",
        "Definido": "Los procesos están documentados, estandarizados y se aplican consistentemente. Los controles están implementados pero no se miden sistemáticamente.",
        "Gestionado": "Los controles se monitorean con métricas cuantitativas. Hay mejora continua basada en datos. La organización está preparada para certificación.",
        "Optimizado": "La gestión de IA está completamente integrada en la cultura organizacional. Automatización, mejora continua y auditoría independiente. Referente del sector."
    }
    return desc.get(nivel, "")


# ══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE PROPUESTA BASADA EN RESULTADOS
# ══════════════════════════════════════════════════════════════════════════════

def generar_prompt_propuesta(resultados: dict, sector: str, contexto: str) -> str:
    """
    Genera el prompt para el LLM basado en los resultados del autodiagnóstico.
    La propuesta sigue la metodología ISO 19011 (5 fases) e incluye ISO 23894.
    """
    dim_desc = []
    for dim_nombre, dim_data in resultados["dimensiones"].items():
        areas_bajas = [p["texto"][:80] for p in dim_data["preguntas"] if p["porcentaje"] < 50]
        areas_medias = [p["texto"][:80] for p in dim_data["preguntas"] if 50 <= p["porcentaje"] < 80]
        dim_desc.append(f"""
### {dim_nombre} — {dim_data['nivel']} ({dim_data['porcentaje']}%)
- Puntaje: {dim_data['puntaje']}/{dim_data['maximo']}
- Áreas críticas (menos de 50%):
{chr(10).join('  - ❌ ' + a for a in areas_bajas) if areas_bajas else '  - Ninguna'}
- Áreas en desarrollo (50-80%):
{chr(10).join('  - ⚠️ ' + a for a in areas_medias) if areas_medias else '  - Ninguna'}
""")

    dim_texto = "\n".join(dim_desc)

    prompt = f"""Eres un consultor senior en ciberseguridad e IA para infraestructura crítica.
Basado en los siguientes resultados de autodiagnóstico de madurez, genera una **Propuesta de Implementación** detallada.

## Resultados del Autodiagnóstico IA-SEC

**Nivel de Madurez Global: {resultados['nivel_global']} ({resultados['porcentaje_global']}%)**
{resultados['descripcion_global']}

**Sector:** {sector}
**Contexto organizacional:** {contexto}

## Desglose por Dimensión:
{dim_texto}

## INSTRUCCIONES — Estructura obligatoria:

Genera la propuesta siguiendo la metodología ISO 19011:2018 con las siguientes secciones:

1. **Resumen Ejecutivo** — Nivel de madurez global, hallazgos principales, urgencia.
2. **Metodología Aplicada** — ISO 19011 (5 fases de auditoría) + ISO 23894 (gestión de riesgos de IA).
3. **Catálogo de Riesgos Identificados** (ISO 23894:2023): Opacidad algorítmica, Deriva del modelo, Vulnerabilidad adversarial, Sesgo algorítmico, Privacidad, Impacto social, Seguridad física. Indicar cuáles aplican según los resultados.
4. **Hoja de Ruta de Implementación** por fases (Fase 1: 0-3m, Fase 2: 3-6m, Fase 3: 6-12m).
5. **Priorización de Controles** — Basada en las dimensiones con menor puntaje.
6. **Arquitectura de Seguridad de IA recomendada**.
7. **KPIs y Métricas de Monitoreo** (ISO 23894 Cláusula 6.2).
8. **Quick Wins** — Acciones inmediatas (primeras 2 semanas).
9. **Plan de Seguimiento** — Próximas auditorías, frecuencia de revisión.

Formato: Markdown estructurado con emojis. Enfocado en acción y resultados medibles."""
    return prompt


# ══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE ENTREGABLES PROFESIONALES
# ══════════════════════════════════════════════════════════════════════════════

def _generar_excel_autodiagnostico(resultados, propuesta, sector="", empresa=""):
    wb = Workbook()
    thin_border = Border(left=Side(style="thin", color="D0D5DD"), right=Side(style="thin", color="D0D5DD"),
                         top=Side(style="thin", color="D0D5DD"), bottom=Side(style="thin", color="D0D5DD"))
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    accent_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="1E293B")
    meta_label_font = Font(name="Calibri", size=10, bold=True, color="1E40AF")
    meta_val_font = Font(name="Calibri", size=10, color="334155")
    score_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    title_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # ── Sheet 1: Portada ──
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_properties.tabColor = "1E40AF"
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 45

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "CID AUDITOR — AUTODIAGNÓSTICO DE MADUREZ IA-SEC"
    c.font = title_font; c.fill = header_fill; c.alignment = title_align
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A2:G2")
    c2 = ws["A2"]
    c2.value = "ISO 27001 · ISO 42001 · NIST AI RMF · ISO 19011 · ISO 23894"
    c2.font = Font(name="Calibri", size=10, color="94A3B8")
    c2.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    pct = resultados["porcentaje_global"]
    nivel = resultados["nivel_global"]
    score_color = "16A34A" if pct >= 70 else "CA8A04" if pct >= 40 else "DC2626"
    score_bg = "DCFCE7" if pct >= 70 else "FEF9C3" if pct >= 40 else "FEE2E2"

    ws.merge_cells("A4:B4")
    ws["A4"].value = "MADUREZ GLOBAL"
    ws["A4"].font = meta_label_font; ws["A4"].alignment = left_align

    ws.merge_cells("C4:G4")
    c = ws["C4"]
    c.value = f"{pct}% — {nivel}"
    c.font = Font(name="Calibri", size=14, bold=True, color=score_color)
    c.fill = PatternFill(start_color=score_bg, end_color=score_bg, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 35

    meta_items = [
        ("Sector", sector),
        ("Empresa", empresa),
        ("Fecha", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Dimensiones evaluadas", str(len(resultados["dimensiones"]))),
    ]
    for i, (label, val) in enumerate(meta_items, 6):
        ws.cell(row=i, column=1, value=label).font = meta_label_font
        ws.cell(row=i, column=1).border = thin_border
        ws.merge_cells(f"A{i}:B{i}")
        c = ws.cell(row=i, column=3, value=val)
        c.font = meta_val_font; c.border = thin_border
        ws.merge_cells(f"C{i}:G{i}")

    ws.merge_cells("A11:G11")
    ws["A11"].value = "RESUMEN POR DIMENSIÓN"
    ws["A11"].font = Font(name="Calibri", size=12, bold=True, color="1E40AF")
    ws["A11"].alignment = left_align

    dim_headers = ["Dimensión", "Puntaje", "Máximo", "%", "Nivel"]
    for j, h in enumerate(dim_headers, 1):
        c = ws.cell(row=12, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

    for row_idx, (dim_nombre, dim_res) in enumerate(resultados["dimensiones"].items(), 13):
        dp = dim_res["porcentaje"]
        d_fill = green_fill if dp >= 70 else yellow_fill if dp >= 40 else red_fill
        ws.cell(row=row_idx, column=1, value=dim_nombre).font = data_font
        ws.cell(row=row_idx, column=1).fill = d_fill; ws.cell(row=row_idx, column=1).border = thin_border
        ws.cell(row=row_idx, column=2, value=dim_res["puntaje"]).font = data_font
        ws.cell(row=row_idx, column=2).fill = d_fill; ws.cell(row=row_idx, column=2).border = thin_border
        ws.cell(row=row_idx, column=2).alignment = center_align
        ws.cell(row=row_idx, column=3, value=dim_res["maximo"]).font = data_font
        ws.cell(row=row_idx, column=3).fill = d_fill; ws.cell(row=row_idx, column=3).border = thin_border
        ws.cell(row=row_idx, column=3).alignment = center_align
        ws.cell(row=row_idx, column=4, value=dp / 100).font = data_font
        ws.cell(row=row_idx, column=4).fill = d_fill; ws.cell(row=row_idx, column=4).border = thin_border
        ws.cell(row=row_idx, column=4).number_format = "0%"
        ws.cell(row=row_idx, column=4).alignment = center_align
        ws.cell(row=row_idx, column=5, value=dim_res["nivel"]).font = Font(name="Calibri", size=10, bold=True, color="1E293B")
        ws.cell(row=row_idx, column=5).fill = d_fill; ws.cell(row=row_idx, column=5).border = thin_border
        ws.cell(row=row_idx, column=5).alignment = center_align

    # ── Sheet 2: Dimensiones ──
    ws2 = wb.create_sheet("Dimensiones")
    ws2.sheet_properties.tabColor = "2563EB"
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 18

    for j, h in enumerate(["Dimensión", "Puntaje", "Máximo", "%", "Nivel de Madurez"], 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
    ws2.row_dimensions[1].height = 30

    for row_idx, (dim_nombre, dim_res) in enumerate(resultados["dimensiones"].items(), 2):
        dp = dim_res["porcentaje"]
        d_fill = green_fill if dp >= 70 else yellow_fill if dp >= 40 else red_fill
        ws2.cell(row=row_idx, column=1, value=dim_nombre).font = data_font
        ws2.cell(row=row_idx, column=1).fill = d_fill; ws2.cell(row=row_idx, column=1).border = thin_border
        ws2.cell(row=row_idx, column=2, value=dim_res["puntaje"]).font = data_font
        ws2.cell(row=row_idx, column=2).fill = d_fill; ws2.cell(row=row_idx, column=2).border = thin_border
        ws2.cell(row=row_idx, column=2).alignment = center_align
        ws2.cell(row=row_idx, column=3, value=dim_res["maximo"]).font = data_font
        ws2.cell(row=row_idx, column=3).fill = d_fill; ws2.cell(row=row_idx, column=3).border = thin_border
        ws2.cell(row=row_idx, column=3).alignment = center_align
        ws2.cell(row=row_idx, column=4, value=dp / 100).font = data_font
        ws2.cell(row=row_idx, column=4).fill = d_fill; ws2.cell(row=row_idx, column=4).border = thin_border
        ws2.cell(row=row_idx, column=4).number_format = "0%"
        ws2.cell(row=row_idx, column=4).alignment = center_align
        ws2.cell(row=row_idx, column=5, value=dim_res["nivel"]).font = Font(name="Calibri", size=10, bold=True, color="1E293B")
        ws2.cell(row=row_idx, column=5).fill = d_fill; ws2.cell(row=row_idx, column=5).border = thin_border
        ws2.cell(row=row_idx, column=5).alignment = center_align

    # ── Sheet 3: Detalle Preguntas ──
    ws3 = wb.create_sheet("Detalle Preguntas")
    ws3.sheet_properties.tabColor = "3B82F6"
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 55
    ws3.column_dimensions["D"].width = 19
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 14

    for j, h in enumerate(["ID", "Dimensión", "Pregunta", "Valor", "Máximo", "%"], 1):
        c = ws3.cell(row=1, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
    ws3.row_dimensions[1].height = 30

    row_idx = 2
    for dim_nombre, dim_res in resultados["dimensiones"].items():
        for p in dim_res["preguntas"]:
            pp = p["porcentaje"]
            p_fill = green_fill if pp >= 70 else yellow_fill if pp >= 40 else red_fill
            ws3.cell(row=row_idx, column=1, value=p["id"]).font = data_font
            ws3.cell(row=row_idx, column=1).fill = p_fill; ws3.cell(row=row_idx, column=1).border = thin_border
            ws3.cell(row=row_idx, column=1).alignment = center_align
            ws3.cell(row=row_idx, column=2, value=dim_nombre).font = data_font
            ws3.cell(row=row_idx, column=2).fill = p_fill; ws3.cell(row=row_idx, column=2).border = thin_border
            ws3.cell(row=row_idx, column=3, value=p["texto"]).font = data_font
            ws3.cell(row=row_idx, column=3).fill = p_fill; ws3.cell(row=row_idx, column=3).border = thin_border
            ws3.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="center")
            ws3.cell(row=row_idx, column=4, value=p["valor"]).font = data_font
            ws3.cell(row=row_idx, column=4).fill = p_fill; ws3.cell(row=row_idx, column=4).border = thin_border
            ws3.cell(row=row_idx, column=4).alignment = center_align
            ws3.cell(row=row_idx, column=5, value=p["maximo"]).font = data_font
            ws3.cell(row=row_idx, column=5).fill = p_fill; ws3.cell(row=row_idx, column=5).border = thin_border
            ws3.cell(row=row_idx, column=5).alignment = center_align
            ws3.cell(row=row_idx, column=6, value=pp / 100).font = data_font
            ws3.cell(row=row_idx, column=6).fill = p_fill; ws3.cell(row=row_idx, column=6).border = thin_border
            ws3.cell(row=row_idx, column=6).number_format = "0%"
            ws3.cell(row=row_idx, column=6).alignment = center_align
            row_idx += 1

    # ── Sheet 4: Propuesta ──
    ws4 = wb.create_sheet("Propuesta")
    ws4.sheet_properties.tabColor = "60A5FA"
    ws4.column_dimensions["A"].width = 120

    ws4.merge_cells("A1:A1")
    ws4["A1"].value = "PROPUESTA DE IMPLEMENTACIÓN"
    ws4["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws4["A1"].fill = header_fill; ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 35

    lines = propuesta.split("\n")
    for i, line in enumerate(lines):
        r = 3 + i
        ws4.cell(row=r, column=1, value=line).font = Font(name="Calibri", size=10, color="1E293B")
        ws4.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    with open(tmp.name, "rb") as f:
        bytes_data = f.read()
    os.unlink(tmp.name)
    return bytes_data


def _generar_pdf_autodiagnostico(resultados, propuesta, sector="", empresa=""):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from datetime import datetime
        import io

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        PURPLE = HexColor("#1E40AF")
        DARK = HexColor("#1E293B")
        GRAY = HexColor("#64748B")
        GREEN = HexColor("#16A34A")
        YELLOW = HexColor("#CA8A04")
        RED = HexColor("#DC2626")
        CODE_BG = HexColor("#F1F5F9")

        story = []
        s_title = ParagraphStyle("Title", parent=styles["Title"], fontSize=20, textColor=PURPLE,
                                 spaceAfter=6, alignment=TA_CENTER, fontName="Helvetica-Bold")
        s_sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=GRAY,
                               alignment=TA_CENTER, spaceAfter=4)
        s_h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, textColor=PURPLE,
                              spaceBefore=16, spaceAfter=8, fontName="Helvetica-Bold")
        s_h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11, textColor=DARK,
                              spaceBefore=12, spaceAfter=6, fontName="Helvetica-Bold")
        s_body = ParagraphStyle("Body", parent=styles["Normal"], fontSize=9, textColor=DARK,
                                leading=14, spaceAfter=4)
        s_meta = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=9, textColor=GRAY, leading=13)

        story.append(Paragraph("CID Auditor", s_title))
        story.append(Paragraph("Autodiagnóstico de Madurez IA-SEC", s_sub))
        story.append(Paragraph("ISO 27001 · ISO 42001 · NIST AI RMF · ISO 19011 · ISO 23894", s_sub))
        story.append(HRFlowable(width="100%", thickness=2, color=PURPLE, spaceAfter=12))
        story.append(Spacer(1, 0.3*cm))

        pct = resultados["porcentaje_global"]
        nivel = resultados["nivel_global"]
        score_color = GREEN if pct >= 70 else YELLOW if pct >= 40 else RED
        story.append(Paragraph(f'Madurez Global: <font color="{score_color.hexval()}">{pct}% — {nivel}</font>',
                               ParagraphStyle("Score", parent=s_h1, fontSize=18, alignment=TA_CENTER, spaceBefore=8, spaceAfter=8)))
        story.append(Spacer(1, 0.3*cm))

        meta_data = [
            ["Sector:", sector or "No especificado"],
            ["Empresa:", empresa or "No especificada"],
            ["Fecha:", datetime.now().strftime("%d/%m/%Y %H:%M")],
            ["Dimensiones:", str(len(resultados["dimensiones"]))],
        ]
        meta_table = Table(meta_data, colWidths=[4*cm, 12*cm])
        meta_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TEXTCOLOR", (0, 0), (0, -1), PURPLE),
            ("TEXTCOLOR", (1, 0), (1, -1), DARK),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.5*cm))

        story.append(Paragraph("Resultados por Dimensión", s_h1))
        dim_header = [Paragraph("<b>Dimensión</b>", s_meta),
                      Paragraph("<b>Puntaje</b>", s_meta),
                      Paragraph("<b>%</b>", s_meta),
                      Paragraph("<b>Nivel</b>", s_meta)]
        dim_rows = [dim_header]
        for dim_nombre, dim_res in resultados["dimensiones"].items():
            dp = dim_res["porcentaje"]
            sc = GREEN if dp >= 70 else YELLOW if dp >= 40 else RED
            dim_rows.append([
                Paragraph(dim_nombre, s_body),
                Paragraph(f"{dim_res['puntaje']}/{dim_res['maximo']}", s_body),
                Paragraph(f'<font color="{sc.hexval()}"><b>{dp}%</b></font>', s_body),
                Paragraph(dim_res["nivel"], s_body),
            ])

        dim_table = Table(dim_rows, colWidths=[7*cm, 3*cm, 2.5*cm, 3.5*cm])
        dim_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), PURPLE),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
            ("ALIGN", (1, 0), (-2, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(dim_table)
        story.append(Spacer(1, 0.5*cm))

        story.append(PageBreak())
        story.append(Paragraph("Propuesta de Implementación", s_h1))
        story.append(HRFlowable(width="100%", thickness=1, color=PURPLE, spaceAfter=8))

        import re as _re

        def _inline_to_html(text):
            text = _re.sub(r'`([^`]+)`', r'<font face="Courier" size="8">\1</font>', text)
            text = _re.sub(r'\*\*([^*]+)\*\*', r'<b>\1</b>', text)
            text = _re.sub(r'(?<!\*)\*([^*]+)\*(?!\*)', r'<i>\1</i>', text)
            text = _re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'<a href="\2" color="#1E40AF">\1</a>', text)
            return text

        s_code = ParagraphStyle("Code", fontName="Courier", fontSize=8,
                                textColor=DARK, leading=12, spaceAfter=6,
                                backColor=CODE_BG, leftIndent=15, rightIndent=15,
                                borderPadding=6)
        s_bullet = ParagraphStyle("Bullet", fontName="Helvetica", fontSize=9.5,
                                  textColor=DARK, leading=15, spaceAfter=3,
                                  leftIndent=20, bulletIndent=10)
        s_bq = ParagraphStyle("BQ", fontName="Helvetica-Oblique", fontSize=9.5,
                              textColor=GRAY, leading=14, spaceAfter=6,
                              leftIndent=20, rightIndent=15)
        s_code_block = ParagraphStyle("CodeBlock", fontName="Courier", fontSize=8,
                                      textColor=DARK, leading=11, spaceAfter=8,
                                      backColor=CODE_BG, leftIndent=15, rightIndent=15,
                                      borderPadding=8)

        in_code = False
        code_buffer = []
        for line in propuesta.split("\n"):
            if line.startswith("```"):
                if in_code:
                    story.append(Paragraph("<br/>".join(code_buffer), s_code_block))
                    code_buffer = []
                    in_code = False
                else:
                    in_code = True
                continue
            if in_code:
                code_buffer.append(line.replace(" ", "&nbsp;") if not line.strip() else line)
                continue

            stripped = line.strip()
            if not stripped:
                story.append(Spacer(1, 0.15*cm))
                continue

            if stripped == "---":
                story.append(HRFlowable(width="100%", thickness=1, color=GRAY, spaceAfter=8, spaceBefore=4))
            elif stripped.startswith("# ") and len(stripped) > 2:
                story.append(Paragraph(stripped[2:], s_h1))
            elif stripped.startswith("## ") and len(stripped) > 3:
                story.append(Paragraph(stripped[3:], s_h2))
            elif stripped.startswith("### ") and len(stripped) > 4:
                story.append(Paragraph(stripped[4:], ParagraphStyle("H3", fontName="Helvetica-Bold",
                              fontSize=10, textColor=DARK, spaceBefore=8, spaceAfter=4)))
            elif stripped.startswith("- ") or stripped.startswith("* "):
                text = _inline_to_html(stripped[2:])
                story.append(Paragraph(f"<bullet>&bull;</bullet> {text}", s_bullet))
            elif stripped[0].isdigit() and ". " in stripped[:4]:
                text = _inline_to_html(stripped.split(". ", 1)[1])
                story.append(Paragraph(f"<bullet>{stripped.split('. ')[0]}.</bullet> {text}", s_bullet))
            elif stripped.startswith("> "):
                story.append(Paragraph(_inline_to_html(stripped[2:]), s_bq))
            elif stripped.startswith("|"):
                cells = [c.strip() for c in stripped.split("|") if c.strip() and c.strip() != "---"]
                if cells:
                    has_sep = any("---" in c for c in stripped.split("|") if c.strip() and c.strip() != "---")
                    if not has_sep:
                        story.append(Paragraph(" | ".join(cells), s_body))
            else:
                story.append(Paragraph(_inline_to_html(stripped), s_body))

        doc.build(story)
        return buffer.getvalue()
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# INTERFAZ STREAMLIT
# ══════════════════════════════════════════════════════════════════════════════

def mostrar_autodiagnostico():
    """Pestaña completa de autodiagnóstico de madurez de IA."""
    st.markdown("#### 🔬 Autodiagnóstico de Madurez IA-SEC")
    st.markdown("""
    **Evalúa la madurez de tu organización en la gestión de sistemas de IA**  
    Basado en ISO 27001 · ISO 42001 · NIST AI RMF · ISO 19011 · ISO 23894

    Responde el cuestionario por dimensión. Al finalizar, obtendrás:
    - Nivel de madurez global y por dimensión (escala MSPI 6 niveles)
    - Propuesta de implementación generada por IA siguiendo ISO 19011
    - Catálogo de riesgos según ISO 23894
    """)

    if "auto_respuestas" not in st.session_state:
        st.session_state.auto_respuestas = {}
    if "auto_resultados" not in st.session_state:
        st.session_state.auto_resultados = None
    if "auto_propuesta" not in st.session_state:
        st.session_state.auto_propuesta = None

    tab_preguntas, tab_resultados, tab_propuesta = st.tabs([
        "📝 Cuestionario", "📊 Resultados de Madurez", "📄 Propuesta de Implementación"
    ])

    with tab_preguntas:
        st.markdown("### Cuestionario de Autodiagnóstico")
        st.markdown("Responde cada pregunta seleccionando la opción que mejor describa el estado actual de tu organización.")

        sector_auto = st.text_input(
            "🏭 Sector / Industria",
            placeholder="Ej: Salud, Finanzas, Energía, Telecomunicaciones...",
            key="auto_sector"
        )
        contexto_auto = st.text_area(
            "📋 Contexto organizacional",
            placeholder="Describe brevemente la organización: tamaño, tipo de sistemas de IA que utiliza, madurez actual...",
            height=80,
            key="auto_contexto"
        )

        respuestas = {}
        for dim in DIMENSIONES_AUTO:
            with st.expander(f"**{dim['nombre']}** — {dim['descripcion']}", expanded=True):
                for p in dim["preguntas"]:
                    opciones = [o[0] for o in p["opciones"]]
                    valores = {o[0]: o[1] for o in p["opciones"]}
                    idx_default = 0
                    saved = st.session_state.auto_respuestas.get(p["id"], None)
                    if saved is not None:
                        for i, o in enumerate(p["opciones"]):
                            if o[1] == saved:
                                idx_default = i
                                break
                    seleccion = st.radio(
                        p["texto"],
                        opciones,
                        index=idx_default,
                        key=f"auto_{p['id']}"
                    )
                    respuestas[p["id"]] = valores[seleccion]

        st.session_state.auto_respuestas = respuestas

        col_calc, col_clear = st.columns([1, 1])
        with col_calc:
            if st.button("🔬 Calcular Madurez", type="primary", use_container_width=True):
                resultados = calcular_madurez(respuestas)
                st.session_state.auto_resultados = resultados
                st.success(f"✅ Madurez calculada: **{resultados['nivel_global']}** ({resultados['porcentaje_global']}%)")
                st.balloons()

        with col_clear:
            if st.button("🔄 Reiniciar cuestionario", use_container_width=True):
                st.session_state.auto_respuestas = {}
                st.session_state.auto_resultados = None
                st.session_state.auto_propuesta = None
                st.rerun()

    with tab_resultados:
        if st.session_state.auto_resultados is None:
            st.info("Completa el cuestionario en la pestaña **📝 Cuestionario** y presiona 'Calcular Madurez'.")
        else:
            res = st.session_state.auto_resultados
            st.markdown(f"### 📊 Nivel de Madurez Global: **{res['nivel_global']}**")
            st.markdown(f"**{res['porcentaje_global']}%** — {res['descripcion_global']}")

            import pandas as pd
            dim_data = []
            for dim_nombre, dim_res in res["dimensiones"].items():
                dim_data.append({
                    "Dimensión": dim_nombre,
                    "Madurez": dim_res["nivel"],
                    "Puntaje": f"{dim_res['puntaje']}/{dim_res['maximo']}",
                    "%": dim_res["porcentaje"]
                })
            df_dim = pd.DataFrame(dim_data)
            st.dataframe(df_dim, use_container_width=True, hide_index=True,
                         column_config={
                             "Dimensión": st.column_config.TextColumn("Dimensión", width="large"),
                             "Madurez": st.column_config.TextColumn("Nivel", width="medium"),
                             "Puntaje": st.column_config.TextColumn("Puntaje", width="small"),
                             "%": st.column_config.ProgressColumn("%", format="%.1f%%", width="medium")
                         })

            import plotly.graph_objects as go
            fig = go.Figure()
            fig.add_trace(go.Scatterpolar(
                r=[res["dimensiones"][d]["porcentaje"] for d in res["dimensiones"]],
                theta=list(res["dimensiones"].keys()),
                fill='toself',
                name='Madurez IA-SEC',
                line_color='#7c3aed'
            ))
            fig.update_layout(
                polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                showlegend=False,
                paper_bgcolor="rgba(0,0,0,0)",
                font_color="white",
                height=450
            )
            st.plotly_chart(fig, use_container_width=True)

            st.markdown("### Desglose por Dimensión")
            for dim_nombre, dim_res in res["dimensiones"].items():
                with st.expander(f"{dim_nombre} — {dim_res['nivel']} ({dim_res['porcentaje']}%)"):
                    for p in dim_res["preguntas"]:
                        color = "🟢" if p["porcentaje"] >= 80 else "🟡" if p["porcentaje"] >= 50 else "🔴"
                        st.markdown(f"{color} **{p['texto'][:100]}...** — {p['valor']}/{p['maximo']} ({p['porcentaje']}%)")

            col_prop, _ = st.columns([1, 2])
            with col_prop:
                if st.button("🚀 Generar Propuesta desde Resultados", type="primary", use_container_width=True):
                    s_auto = st.session_state.get("auto_sector", "").strip()
                    c_auto = st.session_state.get("auto_contexto", "").strip()
                    if not s_auto:
                        st.warning("Ingresa el sector en el cuestionario antes de generar la propuesta.")
                    elif not c_auto:
                        st.warning("Ingresa el contexto organizacional en el cuestionario antes de generar la propuesta.")
                    else:
                        prompt = generar_prompt_propuesta(res, s_auto, c_auto)
                        st.session_state.auto_prompt = prompt
                        st.session_state.auto_propuesta = "PENDIENTE"
                        st.info("Ve a la pestaña **📄 Propuesta de Implementación** y usa el botón para generar con el asistente IA.")

    with tab_propuesta:
        if st.session_state.auto_propuesta is None:
            st.info("Primero completa el cuestionario y calcula la madurez, luego genera la propuesta desde la pestaña de resultados.")
        elif st.session_state.auto_propuesta == "PENDIENTE":
            prompt = st.session_state.get("auto_prompt", "")
            st.markdown("### 📄 Propuesta de Implementación")
            st.markdown("**Metodología:** ISO 19011:2018 (5 fases) + ISO 23894:2023 (gestión de riesgos de IA)")

            col_ia, col_md, col_pdf = st.columns([1, 1, 1])
            with col_ia:
                if st.button("🤖 Generar con Asistente IA", type="primary", use_container_width=True):
                    st.session_state["generar_auto_propuesta"] = True
                    st.rerun()
        else:
            st.markdown("### 📄 Propuesta de Implementación")
            st.markdown(st.session_state.auto_propuesta)
            col_md, col_xls, col_pdf = st.columns(3)
            with col_md:
                st.download_button(
                    "📄 .md", data=st.session_state.auto_propuesta,
                    file_name="autodiagnostico_propuesta.md",
                    mime="text/markdown", key="dl_auto_prop_md",
                    use_container_width=True)
            with col_xls:
                try:
                    res = st.session_state.auto_resultados
                    excel_bytes = _generar_excel_autodiagnostico(
                        res, st.session_state.auto_propuesta,
                        st.session_state.get("auto_sector", ""),
                        st.session_state.get("auto_empresa", ""),
                    )
                    st.download_button("📊 Excel Profesional", data=excel_bytes,
                        file_name=f"autodiagnostico_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_auto_prop_xls", type="primary", use_container_width=True)
                except Exception as e:
                    st.caption(f"Excel no disponible: {e}")

            with col_pdf:
                try:
                    res = st.session_state.auto_resultados
                    pdf_bytes = _generar_pdf_autodiagnostico(
                        res, st.session_state.auto_propuesta,
                        st.session_state.get("auto_sector", ""),
                        st.session_state.get("auto_empresa", ""),
                    )
                    st.download_button(
                        "📄 PDF Profesional", data=pdf_bytes,
                        file_name=f"autodiagnostico_{datetime.now().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf", key="dl_auto_prop_pdf",
                        use_container_width=True)
                except Exception as e:
                    st.caption(f"PDF no disponible: {e}")
