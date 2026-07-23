from core.llm_factory import ejecutar_gap_analysis, ejecutar_chat, extraer_texto_pdf
from database import guardar_reporte, guardar_vulnerabilidad
from audit_logger import log_event
from pdf_export import generar_pdf


def _generar_excel_gap_analysis(resultado: str, stats: dict, nombre_doc: str, usuario: str, proveedor: str, modelo: str) -> bytes:
    """
    Genera un Excel profesional con los resultados del Gap Analysis enriquecido con RAG.
    Incluye: Portada, Resumen Ejecutivo, Resultados parseados, Controles ISO, y Fuentes RAG.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import tempfile, os, re
    from datetime import datetime

    wb = Workbook()
    thin_border = Border(left=Side(style="thin", color="334155"), right=Side(style="thin", color="334155"),
                         top=Side(style="thin", color="334155"), bottom=Side(style="thin", color="334155"))
    dark_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    accent_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
    header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
    title_font = Font(name="Century Gothic", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Century Gothic", size=9, bold=True, color="FFFFFF")
    data_font = Font(name="Century Gothic", size=8, color="E0E0E0")

    # ── HOJA 1: PORTADA ──
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_properties.tabColor = "1a1a2e"
    ws.merge_cells("A1:F1")
    ws["A1"] = "INFORME DE GAP ANALYSIS"
    ws["A1"].font = title_font
    ws["A1"].fill = dark_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45

    meta = [
        ("Documento", nombre_doc),
        ("Fecha", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Auditor", usuario),
        ("Proveedor IA", proveedor),
        ("Modelo", modelo),
        ("Cumplimiento", f"{stats.get('cumplimiento_pct', 0)}%"),
        ("Conformes", str(stats.get('conformes', 0))),
        ("Parciales", str(stats.get('parciales', 0))),
        ("No Conformes", str(stats.get('no_conformes', 0))),
    ]
    for i, (label, val) in enumerate(meta, 4):
        ws.cell(row=i, column=1, value=label).font = Font(name="Century Gothic", size=9, bold=True, color="58C8F2")
        ws.cell(row=i, column=1).fill = header_fill
        ws.cell(row=i, column=1).border = thin_border
        ws.merge_cells(f"A{i}:B{i}")
        c = ws.cell(row=i, column=3, value=val)
        c.font = Font(name="Century Gothic", size=9, color="E0E0E0")
        c.fill = accent_fill
        c.border = thin_border
        ws.merge_cells(f"C{i}:F{i}")
    for col_l in "ABCDEF": ws.column_dimensions[col_l].width = 20

    # ── HOJA 2: RESULTADOS ──
    ws2 = wb.create_sheet("Resultados Gap Analysis")
    ws2.sheet_properties.tabColor = "a78bfa"
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "RESULTADOS DEL GAP ANALYSIS"
    ws2["A1"].font = Font(name="Century Gothic", size=12, bold=True, color="FFFFFF")
    ws2["A1"].fill = dark_fill
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 30

    # Parsear tabla markdown a filas de Excel
    lines = resultado.split("\n")
    row_idx = 3
    cols_escritas = 0
    for line in lines:
        if line.startswith("|"):
            cells = [c.strip() for c in line.split("|") if c.strip()]
            if not cells:
                continue
            if line.startswith("| -") or line.startswith("|---"):
                continue
            is_header = any(kw in cells[0].lower() for kw in ["control", "requisito", "estado", "norma"])
            for col, val in enumerate(cells, 1):
                cell = ws2.cell(row=row_idx, column=col, value=val)
                if is_header:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
            ws2.row_dimensions[row_idx].height = 28
            row_idx += 1
            cols_escritas = max(cols_escritas, len(cells))

    if stats.get('cumplimiento_pct', 0) > 0:
        row_idx += 1
        ws2.merge_cells(f"A{row_idx}:D{row_idx}")
        ws2[f"A{row_idx}"] = f"📊 RESUMEN: {stats.get('cumplimiento_pct', 0)}% de cumplimiento | "
        ws2[f"A{row_idx}"] += f"Conformes: {stats.get('conformes', 0)} | Parciales: {stats.get('parciales', 0)} | No Conformes: {stats.get('no_conformes', 0)}"
        ws2[f"A{row_idx}"].font = Font(name="Century Gothic", size=10, bold=True, color="58C8F2")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 45
    ws2.column_dimensions["D"].width = 45

    # ── HOJA 3: TEXTO COMPLETO ──
    ws3 = wb.create_sheet("Informe Completo")
    ws3.sheet_properties.tabColor = "58C8F2"
    ws3.merge_cells("A1:B1")
    ws3["A1"] = "INFORME COMPLETO DEL GAP ANALYSIS"
    ws3["A1"].font = Font(name="Century Gothic", size=12, bold=True, color="FFFFFF")
    ws3["A1"].fill = dark_fill
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 30

    for i, line in enumerate(lines):
        r = 3 + i
        ws3.merge_cells(f"A{r}:B{r}")
        c = ws3.cell(row=r, column=1, value=line)
        c.font = Font(name="Century Gothic", size=8, color="E0E0E0")
        c.alignment = Alignment(wrap_text=True)
        ws3.row_dimensions[r].height = 18
    ws3.column_dimensions["A"].width = 80
    ws3.column_dimensions["B"].width = 40

    # ── HOJA 4: FUENTES RAG ──
    ws4 = wb.create_sheet("Fuentes RAG")
    ws4.sheet_properties.tabColor = "10b981"
    ws4.merge_cells("A1:D1")
    ws4["A1"] = "FUENTES NORMATIVAS CONSULTADAS (RAG)"
    ws4["A1"].font = Font(name="Century Gothic", size=12, bold=True, color="FFFFFF")
    ws4["A1"].fill = dark_fill
    ws4["A1"].alignment = Alignment(horizontal="center")
    ws4.row_dimensions[1].height = 30

    rag_headers = ["Documento", "Sección", "Fragmento", "Relevancia"]
    for col, h in enumerate(rag_headers, 1):
        c = ws4.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    # Intentar obtener fuentes del RAG real
    try:
        from core.rag_engine import get_rag
        engine = get_rag()
        if engine.is_active:
            # Buscar contexto con el resultado para identificar qué secciones se usaron
            fuentes = engine.search(resultado[:500], top_k=5, threshold=0.2)
            for i, src in enumerate(fuentes, 4):
                ws4.cell(row=i, column=1, value=src["source"]).font = data_font
                ws4.cell(row=i, column=1).border = thin_border
                ws4.cell(row=i, column=2, value=src.get("section", "General")).font = data_font
                ws4.cell(row=i, column=2).border = thin_border
                ws4.cell(row=i, column=3, value=src["text"][:300]).font = data_font
                ws4.cell(row=i, column=3).border = thin_border
                ws4.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
                c = ws4.cell(row=i, column=4, value=src.get("score", 0))
                c.font = Font(name="Century Gothic", size=8, color="AAAAAA")
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
    except Exception:
        ws4.merge_cells("A4:D4")
        ws4["A4"] = "Las fuentes RAG se cargarán al consultar durante el análisis."
        ws4["A4"].font = Font(italic=True, color="888888")

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 70
    ws4.column_dimensions["D"].width = 12

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    with open(tmp.name, "rb") as f:
        data = f.read()
    os.unlink(tmp.name)
    return data


def analyze_pdf_policy(archivo, proveedor, modelo, temperatura, usar_rag, usuario):
    texto = extraer_texto_pdf(archivo)
    palabras = len(texto.split())
    if palabras == 0:
        return None, 0, "No se pudo extraer texto del PDF."

    log_event("GAP_ANALYSIS_START", user=usuario, detail=f"Archivo: {archivo.name} | Modelo: {proveedor}/{modelo}")

    resultado = ejecutar_gap_analysis(texto, proveedor, modelo, temperatura, usar_rag)
    stats = guardar_reporte(usuario, "PDF", archivo.name, proveedor, modelo, resultado)

    log_event("GAP_ANALYSIS_DONE", user=usuario, result="success",
              detail=f"Archivo: {archivo.name} | Cumplimiento: {stats['cumplimiento_pct']}%")

    pdf_bytes = generar_pdf(resultado, archivo.name, usuario, proveedor, modelo, stats)

    # Generar Excel profesional
    excel_bytes = _generar_excel_gap_analysis(resultado, stats, archivo.name, usuario, proveedor, modelo)

    return resultado, stats, pdf_bytes, excel_bytes, texto


def analyze_text_policy(texto, proveedor, modelo, temperatura, usar_rag, usuario, nombre_doc="texto_manual"):
    palabras = len(texto.split())
    if palabras == 0:
        return None, 0, "El texto está vacío."

    log_event("GAP_ANALYSIS_START", user=usuario, detail=f"Texto libre | {palabras} palabras | Modelo: {proveedor}/{modelo}")

    resultado = ejecutar_gap_analysis(texto, proveedor, modelo, temperatura, usar_rag)
    stats = guardar_reporte(usuario, "TEXTO", nombre_doc, proveedor, modelo, resultado)

    log_event("GAP_ANALYSIS_DONE", user=usuario, result="success",
              detail=f"Texto libre | Cumplimiento: {stats['cumplimiento_pct']}%")

    pdf_bytes = generar_pdf(resultado, nombre_doc, usuario, proveedor, modelo, stats)
    return resultado, stats, pdf_bytes


def chat_with_auditor(mensajes, proveedor, modelo, temperatura):
    return ejecutar_chat(mensajes, proveedor, modelo, temperatura)
