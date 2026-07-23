from views.common import *


def render(usuario, nombre, proveedor, modelo, temperatura, usar_rag, opcion_menu):
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 🧩 Matriz de Confluencia Normativa IA-SEC")
    st.markdown("**Herramienta de apoyo al auditor** — ISO/IEC 27001:2022 + ISO/IEC 42001:2023 + NIST AI RMF 1.0")
    st.markdown("*23 controles correlacionados · 9 dimensiones operativas · Escala MSPI 6 niveles*")

    # Selección de sector
    sectores_opciones = ["Selecciona un sector..."] + [f"{k}: {v['nombre']}" for k, v in SECTORES_IC.items()]
    sector_seleccionado = st.selectbox(
        "🏭 Sector de Infraestructura Crítica",
        sectores_opciones,
        help="Selecciona el sector específico para contextualizar la evaluación"
    )

    sector_id = None
    if sector_seleccionado != "Selecciona un sector...":
        sector_id = sector_seleccionado.split(":")[0]
        sector_info = obtener_sector(sector_id)
        if sector_info:
            st.info(f"**{sector_info['nombre']}**: {sector_info['descripcion']}")
            st.write(f"**Regulaciones adicionales:** {', '.join(sector_info['regulaciones_adicionales'])}")

    st.markdown("---")

    # Inicializar DataFrame en sesión si no existe
    if 'df_interop' not in st.session_state:
        st.session_state.df_interop = generar_df_interoperabilidad()

    # Mostrar métricas de resumen
    resumen = generar_resumen_interoperabilidad()
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Controles", resumen["total_controles"])
    with col2:
        st.metric("Controles Críticos", resumen["por_relevancia"]["Crítica"])
    
    # Calcular cumplimiento actual desde el DataFrame
    df_eval = st.session_state.df_interop
    df_eval['%'] = df_eval['Estado Actual'].apply(calcular_cumplimiento)
    avance_global = df_eval['%'].mean()
    
    with col3:
        st.metric("Dimensiones", len(resumen["dimensiones"]))
    with col4:
        st.metric("Avance Global", f"{avance_global:.1%}")

    st.markdown("---")

    # Editor de datos interactivo (Modo GRC Excel)
    st.markdown("#### 📊 Matriz de Evaluación GRC (Interactivo)")
    st.info("Utiliza esta tabla para evaluar el estado de implementación. Los cambios se guardan automáticamente en la sesión.")
    
    edited_df = st.data_editor(
        st.session_state.df_interop,
        column_config={
            "ID": st.column_config.TextColumn("ID", width="small", disabled=True),
            "Dimensión": st.column_config.TextColumn("Dimensión", width="medium", disabled=True),
            "ISO 27001": st.column_config.TextColumn("ISO 27001", width="medium", disabled=True),
            "ISO 42001": st.column_config.TextColumn("ISO 42001", width="medium", disabled=True),
            "NIST AI RMF": st.column_config.TextColumn("NIST AI RMF", width="medium", disabled=True),
            "Relevancia IC": st.column_config.TextColumn("Relevancia IC", width="small", disabled=True),
            "Estado Actual": st.column_config.SelectboxColumn(
                "Estado Actual",
                options=[
                    "No iniciado", 
                    "En proceso inicial", 
                    "Implementado parcialmente", 
                    "Mayormente implementado", 
                    "Cumplimiento total",
                    "No evaluado"
                ],
                required=True,
                width="medium"
            ),
            "Plan de Acción": st.column_config.TextColumn(
                "¿Qué voy a hacer para lograrlo?",
                width="large"
            )
        },
        hide_index=True,
        use_container_width=True,
        key="grc_editor"
    )
    
    # Actualizar sesión con los cambios
    st.session_state.df_interop = edited_df

    # Botón para descargar en Excel (Estilo GRC GAP)
    if 'df_interop' in st.session_state:
        excel_bytes = exportar_matriz_excel(st.session_state.df_interop)
        st.download_button(
            label="📥 Descargar Matriz de Cumplimiento GRC (Excel)",
            data=excel_bytes,
            file_name="Matriz_Cumplimiento_Integrada_GRC.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Descarga un reporte Excel profesional con pestañas separadas por norma, semáforos de cumplimiento y la tabla de valoración de madurez."
        )

    st.markdown("---")

    # Navegación por dimensiones para vista detallada
    dimensiones = obtener_dimensiones()
    dimension_seleccionada = st.selectbox(
        "🔍 Explorar Requisitos Detallados por Dimensión",
        ["Vista general"] + dimensiones,
        help="Selecciona una dimensión para ver los controles detallados y requisitos específicos de IC"
    )

    if dimension_seleccionada == "Vista general":
        # Vista general con estadísticas
        import pandas as pd

        # Crear tabla resumen por dimensión
        data_dimensiones = []
        for dim in dimensiones:
            controles = obtener_controles_por_dimension(dim)
            criticos = sum(1 for c in controles if c.relevancia_ic == "Crítica")
            muy_altos = sum(1 for c in controles if c.relevancia_ic == "Muy Alta")
            altos = sum(1 for c in controles if c.relevancia_ic == "Alta")
            data_dimensiones.append({
                "Dimensión": dim,
                "Total Controles": len(controles),
                "Críticos": criticos,
                "Muy Altos": muy_altos,
                "Altos": altos
            })

        df_dimensiones = pd.DataFrame(data_dimensiones)
        st.dataframe(df_dimensiones, use_container_width=True, hide_index=True)

        # Gráfico de distribución por relevancia
        fig_relevancia = go.Figure()
        relevancia_data = resumen["por_relevancia"]
        fig_relevancia.add_trace(go.Pie(
            labels=list(relevancia_data.keys()),
            values=list(relevancia_data.values()),
            marker_colors=["#dc2626", "#d97706", "#059669"],
            title="Distribución por relevancia en IC"
        ))
        fig_relevancia.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="white", showlegend=True,
            height=300,
        )
        st.plotly_chart(fig_relevancia, use_container_width=True)

        # Gráfico de barras de cumplimiento
        # Recalcular cobertura para el gráfico
        evals_dict = {row['ID']: row['Estado Actual'] for _, row in st.session_state.df_interop.iterrows()}
        cobertura = calcular_cobertura(evals_dict)

        st.markdown("#### 📈 Nivel de Cumplimiento por Dimensión")
        cols_cob = st.columns(len(cobertura))
        for i, (dim, stats) in enumerate(cobertura.items()):
            with cols_cob[i % len(cols_cob)]:
                st.metric(dim, f"{stats['porcentaje']}%")
        
        fig_cumplimiento = px.bar(
            x=list(cobertura.keys()),
            y=[s["porcentaje"] for s in cobertura.values()],
            labels={"x": "Dimensión", "y": "% Cumplimiento"},
            title="Porcentaje de Cumplimiento Acumulado",
            color_discrete_sequence=["#a78bfa"]
        )
        fig_cumplimiento.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="white", height=350
        )
        st.plotly_chart(fig_cumplimiento, use_container_width=True)

    else:
        # Vista detallada de una dimensión
        controles = obtener_controles_por_dimension(dimension_seleccionada)
        st.subheader(f"📋 {dimension_seleccionada}")

        # Controles en formato expandible
        for control in controles:
            relevancia_color = {
                "Crítica": "#dc2626",
                "Muy Alta": "#d97706",
                "Alta": "#059669"
            }.get(control.relevancia_ic, "#64748b")

            with st.expander(f"🔴 {control.id}: {control.iso_27001_ref} ({control.relevancia_ic})"):
                col_a, col_b = st.columns(2)

                with col_a:
                    st.markdown("**ISO/IEC 27001:2022**")
                    st.write(f"*{control.iso_27001_ref}*")
                    st.write(control.iso_27001_desc)

                    st.markdown("**ISO/IEC 42001:2023**")
                    st.write(f"*{control.iso_42001_ref}*")
                    st.write(control.iso_42001_desc)

                    st.markdown("**NIST AI RMF 1.0**")
                    st.write(f"*{control.nist_ai_rmf_ref}*")
                    st.write(control.nist_ai_rmf_desc)

                with col_b:
                    st.markdown(f"**Relevancia en IC:** <span style='color:{relevancia_color};font-weight:bold'>{control.relevancia_ic}</span>", unsafe_allow_html=True)
                    st.write("**Requisito específico IC:**")
                    st.write(control.requisito_ic)
                    
                    # Mostrar estado y plan desde la matriz GRC
                    row = st.session_state.df_interop[st.session_state.df_interop['ID'] == control.id]
                    if not row.empty:
                        estado = row['Estado Actual'].values[0]
                        plan = row['Plan de Acción'].values[0]
                        st.markdown(f"**Estado Evaluación:** `{estado}`")
                        if plan:
                            st.success(f"📌 **Plan:** {plan}")

                    st.write("**Riesgo sin control:**")
                    st.write(control.riesgo_sin_control)

    st.markdown("---")

    # Generación de propuestas
    if sector_id:
        st.markdown("### 🎯 Generar Propuesta de Implementación")
        st.markdown("""
        **Metodología:**  
        🔷 **ISO 19011:2018** — Las 5 fases del ciclo de auditoría guían la estructura de la propuesta  
        🔷 **ISO 23894:2023** — Los riesgos de IA (opacidad, deriva, adversarial, sesgo, privacidad, impacto social, seguridad física) se incorporan como catálogo de riesgos  
        🔷 **ISO 27001 + ISO 42001 + NIST AI RMF** — Controles correlacionados para la evaluación
        """)

        # Evaluación de controles
        st.markdown("#### 📝 Datos para la propuesta")
        st.info("La propuesta se basará en los datos ingresados en la Matriz de Evaluación GRC y el contexto organizacional.")

        # Recolectar evaluaciones del DataFrame de la sesión
        controles_evaluados = {row['ID']: row['Estado Actual'] for _, row in st.session_state.df_interop.iterrows()}

        # Contexto organizacional
        contexto_org = st.text_area(
            "📋 Contexto de la organización",
            placeholder="Describe brevemente la organización: tamaño, madurez en ciberseguridad, experiencia previa con IA, regulaciones aplicables, etc.",
            height=100
        )

        # Botón de generación
        if st.button("🚀 Generar Propuesta de Implementación", type="primary"):
            if not contexto_org.strip():
                st.error("Debes proporcionar contexto organizacional.")
            else:
                log_event("INTEROP_PROPOSAL_START", user=usuario, detail=f"Sector: {sector_id} | Controles evaluados: {len(controles_evaluados)}")

                with st.spinner("🤖 Generando propuesta personalizada... esto puede tardar 1-2 minutos."):
                    propuesta = generar_propuesta_ic(
                        sector_id, controles_evaluados, contexto_org,
                        proveedor, modelo, temperatura
                    )

                if propuesta.startswith("ERROR:"):
                    st.error(propuesta)
                else:
                    log_event("INTEROP_PROPOSAL_DONE", user=usuario, result="success", detail=f"Sector: {sector_id} | Propuesta generada")

                    st.success("✅ Propuesta generada exitosamente")

                    # Mostrar propuesta
                    st.markdown("---")
                    st.markdown("## 📄 Propuesta de Implementación")
                    st.markdown(propuesta)

                    # Opciones de descarga
                    col_md, col_xls, col_pdf = st.columns(3)
                    with col_md:
                        st.download_button(
                            "📄 Descargar (.md)",
                            data=propuesta,
                            file_name=f"Propuesta_Implementacion_{sector_id}.md",
                            mime="text/markdown",
                            key="download_md"
                        )
                    with col_xls:
                        try:
                            import tempfile, os
                            from openpyxl import Workbook
                            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                            from openpyxl.utils import get_column_letter

                            wb = Workbook()

                            # ── HOJA 1: PORTADA ──
                            ws0 = wb.active
                            ws0.title = "Portada"
                            ws0.sheet_properties.tabColor = "1a1a2e"

                            dark_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
                            accent_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
                            header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
                            white_font = Font(name="Century Gothic", size=11, bold=True, color="FFFFFF")
                            title_font = Font(name="Century Gothic", size=18, bold=True, color="5e6ad2")
                            sub_font = Font(name="Century Gothic", size=10, color="AAAAAA")
                            thin_border = Border(
                                left=Side(style="thin", color="334155"),
                                right=Side(style="thin", color="334155"),
                                top=Side(style="thin", color="334155"),
                                bottom=Side(style="thin", color="334155"),
                            )

                            ws0.merge_cells("A1:F1")
                            ws0["A1"] = "PROPUESTA DE IMPLEMENTACIÓN"
                            ws0["A1"].font = title_font
                            ws0["A1"].fill = dark_fill
                            ws0["A1"].alignment = Alignment(horizontal="center", vertical="center")
                            ws0.row_dimensions[1].height = 50

                            ws0.merge_cells("A2:F2")
                            ws0["A2"] = "Auditoría de Ciberseguridad e Inteligencia Artificial en Infraestructura Crítica"
                            ws0["A2"].font = sub_font
                            ws0["A2"].fill = dark_fill
                            ws0["A2"].alignment = Alignment(horizontal="center")

                            info_data = [
                                ("Sector", sector_info["nombre"] if sector_info else sector_id),
                                ("Fecha", datetime.now().strftime("%Y-%m-%d %H:%M")),
                                ("Auditor", nombre),
                                ("Usuario", usuario),
                                ("Proveedor IA", proveedor),
                                ("Modelo", modelo),
                                ("Controles Evaluados", str(len([v for v in controles_evaluados.values() if v != "No evaluado"]))),
                                ("Marcos Normativos", "ISO 27001:2022 · ISO 42001:2023 · NIST AI RMF 1.0 · ISO 19011:2018 · ISO 23894:2023"),
                            ]
                            for i, (label, val) in enumerate(info_data, 4):
                                lbl = ws0.cell(row=i, column=1, value=label)
                                lbl.font = Font(name="Century Gothic", size=9, bold=True, color="58C8F2")
                                lbl.fill = header_fill
                                lbl.border = thin_border
                                ws0.merge_cells(f"A{i}:B{i}")
                                val_cell = ws0.cell(row=i, column=3, value=val)
                                val_cell.font = Font(name="Century Gothic", size=9, color="E0E0E0")
                                val_cell.fill = accent_fill
                                val_cell.border = thin_border
                                ws0.merge_cells(f"C{i}:F{i}")
                                ws0.row_dimensions[i].height = 22

                            for col_l in ["A","B","C","D","E","F"]:
                                ws0.column_dimensions[col_l].width = 18

                            # ── HOJA 2: RESUMEN EJECUTIVO ──
                            ws1 = wb.create_sheet("Resumen Ejecutivo")
                            ws1.sheet_properties.tabColor = "a78bfa"

                            ws1.merge_cells("A1:F1")
                            ws1["A1"] = "RESUMEN EJECUTIVO — PROPUESTA DE IMPLEMENTACIÓN"
                            ws1["A1"].font = Font(name="Century Gothic", size=13, bold=True, color="FFFFFF")
                            ws1["A1"].fill = dark_fill
                            ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
                            ws1.row_dimensions[1].height = 35

                            ws1.merge_cells("A3:F3")
                            ws1["A3"] = "CONTEXTO ORGANIZACIONAL"
                            ws1["A3"].font = Font(name="Century Gothic", size=10, bold=True, color="58C8F2")
                            ws1.merge_cells("A4:F4")
                            ws1["A4"] = contexto_org
                            ws1["A4"].font = Font(name="Century Gothic", size=9, color="CCCCCC")
                            ws1["A4"].alignment = Alignment(wrap_text=True)
                            ws1.row_dimensions[4].height = 50

                            ws1.merge_cells("A6:F6")
                            ws1["A6"] = "RESUMEN DE EVALUACIÓN"
                            ws1["A6"].font = Font(name="Century Gothic", size=10, bold=True, color="58C8F2")

                            estados_val = list(controles_evaluados.values())
                            total_controles = len(estados_val)
                            cumplimiento = sum(1 for e in estados_val if e in ("Implementado parcialmente", "Mayormente implementado", "Cumplimiento total"))
                            pct_cumpl = round(cumplimiento / total_controles * 100) if total_controles else 0

                            resumen_kpis = [
                                ("Total Controles", str(total_controles)),
                                ("En Cumplimiento", str(cumplimiento)),
                                ("% Cumplimiento", f"{pct_cumpl}%"),
                                ("Nivel de Madurez", "BÁSICO" if pct_cumpl < 40 else "INTERMEDIO" if pct_cumpl < 70 else "AVANZADO"),
                            ]
                            for i, (label, val) in enumerate(resumen_kpis):
                                r = 8 + i
                                c_lbl = ws1.cell(row=r, column=1, value=label)
                                c_lbl.font = Font(name="Century Gothic", size=9, bold=True, color="E0E0E0")
                                c_lbl.fill = header_fill
                                c_lbl.border = thin_border
                                ws1.merge_cells(f"A{r}:B{r}")
                                c_val = ws1.cell(row=r, column=3, value=val)
                                c_val.font = Font(name="Century Gothic", size=11, bold=True, color="5e6ad2")
                                c_val.fill = accent_fill
                                c_val.border = thin_border
                                c_val.alignment = Alignment(horizontal="center")
                                ws1.merge_cells(f"C{r}:D{r}")

                            ws1.column_dimensions["A"].width = 18
                            ws1.column_dimensions["B"].width = 18
                            ws1.column_dimensions["C"].width = 18
                            ws1.column_dimensions["D"].width = 18
                            ws1.column_dimensions["E"].width = 18
                            ws1.column_dimensions["F"].width = 18

                            # ── HOJA 3: MATRIZ DE EVALUACIÓN ──
                            ws2 = wb.create_sheet("Matriz de Evaluación")
                            ws2.sheet_properties.tabColor = "dc2626"

                            eval_headers = ["#", "Control ID", "Dimensión", "ISO 27001", "ISO 42001", "NIST AI RMF", "Relevancia IC", "Estado Actual", "Plan de Acción"]
                            for col, h in enumerate(eval_headers, 1):
                                c = ws2.cell(row=1, column=col, value=h)
                                c.font = Font(name="Century Gothic", size=8, bold=True, color="FFFFFF")
                                c.fill = header_fill
                                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                c.border = thin_border
                            ws2.row_dimensions[1].height = 30

                            fills_estado = {
                                "No iniciado": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
                                "En proceso inicial": PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid"),
                                "Implementado parcialmente": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
                                "Mayormente implementado": PatternFill(start_color="44BB66", end_color="44BB66", fill_type="solid"),
                                "Cumplimiento total": PatternFill(start_color="228B22", end_color="228B22", fill_type="solid"),
                            }
                            data_font = Font(name="Century Gothic", size=8, color="E0E0E0")
                            data_fill_even = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

                            from interoperabilidad import obtener_controles_por_dimension, obtener_dimensiones
                            row_idx = 2
                            idx = 1
                            for dim in obtener_dimensiones():
                                controles = obtener_controles_por_dimension(dim)
                                for ctrl in controles:
                                    estado = controles_evaluados.get(ctrl.id, "No evaluado")
                                    plan = st.session_state.df_interop[
                                        st.session_state.df_interop['ID'] == ctrl.id
                                    ]['Plan de Acción'].values[0] if not st.session_state.df_interop[
                                        st.session_state.df_interop['ID'] == ctrl.id
                                    ].empty else ""
                                    row_data = [idx, ctrl.id, dim, ctrl.iso_27001_ref, ctrl.iso_42001_ref,
                                                ctrl.nist_ai_rmf_ref, ctrl.relevancia_ic, estado, plan]
                                    for col, val in enumerate(row_data, 1):
                                        c = ws2.cell(row=row_idx, column=col, value=val)
                                        c.font = data_font
                                        c.border = thin_border
                                        c.alignment = Alignment(wrap_text=True, vertical="top")
                                        c.fill = data_fill_even
                                        if col == 8:  # Color por estado
                                            estado_fill = fills_estado.get(val)
                                            if estado_fill:
                                                c.fill = estado_fill
                                                c.font = Font(name="Century Gothic", size=8, bold=True,
                                                             color="000000" if val in ("Implementado parcialmente",) else "FFFFFF")
                                    row_idx += 1
                                    idx += 1

                            eval_widths = {"A": 4, "B": 14, "C": 16, "D": 16, "E": 16, "F": 16, "G": 14, "H": 22, "I": 40}
                            for col_l, w in eval_widths.items():
                                ws2.column_dimensions[col_l].width = w

                            # ── HOJA 4: PROPUESTA COMPLETA ──
                            ws3 = wb.create_sheet("Propuesta Completa")
                            ws3.sheet_properties.tabColor = "58C8F2"

                            ws3.merge_cells("A1:B1")
                            ws3["A1"] = "PROPUESTA DE IMPLEMENTACIÓN — TEXTO COMPLETO"
                            ws3["A1"].font = Font(name="Century Gothic", size=12, bold=True, color="FFFFFF")
                            ws3["A1"].fill = dark_fill
                            ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
                            ws3.row_dimensions[1].height = 35

                            # Write the proposal text across merged cells
                            lines = propuesta.split("\n")
                            for i, line in enumerate(lines):
                                r = 3 + i
                                ws3.merge_cells(f"A{r}:B{r}")
                                cell = ws3.cell(row=r, column=1, value=line)
                                cell.font = Font(name="Century Gothic", size=9, color="E0E0E0" if not line.startswith("#") else "58C8F2",
                                                bold=line.startswith("#"))
                                cell.alignment = Alignment(wrap_text=True)
                                ws3.row_dimensions[r].height = 18 if not line.strip() else 22

                            ws3.column_dimensions["A"].width = 80
                            ws3.column_dimensions["B"].width = 40

                            # ── HOJA 5: PLAN DE ACCIÓN ──
                            ws4 = wb.create_sheet("Plan de Acción")
                            ws4.sheet_properties.tabColor = "10b981"

                            plan_headers = ["Prioridad", "Control ID", "Dimensión", "Estado Actual", "Acción Recomendada", "Responsable", "Plazo Estimado"]
                            for col, h in enumerate(plan_headers, 1):
                                c = ws4.cell(row=1, column=col, value=h)
                                c.font = Font(name="Century Gothic", size=9, bold=True, color="FFFFFF")
                                c.fill = header_fill
                                c.alignment = Alignment(horizontal="center", wrap_text=True)
                                c.border = thin_border
                            ws4.row_dimensions[1].height = 28

                            plan_font = Font(name="Century Gothic", size=8, color="E0E0E0")
                            row_p = 2
                            prioridad = 1
                            for dim in obtener_dimensiones():
                                controles = obtener_controles_por_dimension(dim)
                                for ctrl in controles:
                                    estado = controles_evaluados.get(ctrl.id, "No evaluado")
                                    plan = st.session_state.df_interop[
                                        st.session_state.df_interop['ID'] == ctrl.id
                                    ]['Plan de Acción'].values[0] if not st.session_state.df_interop[
                                        st.session_state.df_interop['ID'] == ctrl.id
                                    ].empty else ""
                                    if estado in ("No iniciado", "En proceso inicial", "Implementado parcialmente"):
                                        prioridad_label = f"P{prioridad}"
                                        prioridad += 1
                                        for col, val in enumerate([prioridad_label, ctrl.id, dim, estado, plan or "Definir plan de implementación", "Por asignar", "90 días"], 1):
                                            c = ws4.cell(row=row_p, column=col, value=val)
                                            c.font = plan_font
                                            c.border = thin_border
                                            c.alignment = Alignment(wrap_text=True)
                                        ws4.row_dimensions[row_p].height = 28
                                        row_p += 1

                            plan_widths = {"A": 12, "B": 14, "C": 16, "D": 22, "E": 45, "F": 18, "G": 16}
                            for col_l, w in plan_widths.items():
                                ws4.column_dimensions[col_l].width = w

                            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                            wb.save(tmp.name)
                            tmp.close()
                            with open(tmp.name, 'rb') as f:
                                excel_propuesta = f.read()
                            os.unlink(tmp.name)

                            st.download_button(
                                "📊 Descargar Propuesta (Excel)",
                                data=excel_propuesta,
                                file_name=f"Propuesta_Implementacion_{sector_id}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_xls",
                                type="primary",
                                use_container_width=True,
                            )
                        except Exception as e:
                            st.error(f"Error generando Excel: {e}")

                    with col_pdf:
                        sector_info = obtener_sector(sector_id)
                        sector_nombre = sector_info["nombre"] if sector_info else sector_id
                        pdf_bytes = generar_pdf_propuesta_interoperabilidad(
                            propuesta, sector_id, sector_nombre, controles_evaluados,
                            usuario, proveedor, modelo
                        )
                        st.download_button(
                            "📄 Descargar PDF (auditoría ISO 19011)",
                            data=pdf_bytes,
                            file_name=f"Informe_Auditoria_{sector_id}.pdf",
                            mime="application/pdf",
                            key="download_pdf",
                            use_container_width=True,
                        )

    # Descarga de la matriz completa
    st.markdown("---")
    st.markdown("### 📥 Descargas")
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        matriz_md = generar_markdown_matriz(sector_id)
        st.download_button(
            "📄 Matriz Interoperabilidad (.md)",
            data=matriz_md,
            file_name=f"matriz_interoperabilidad{'_' + sector_id if sector_id else ''}.md",
            mime="text/markdown",
            key="matriz_md"
        )
    with col_b:
        # Excel profesional con formato
        df_excel = generar_df_interoperabilidad()
        if sector_id:
            sector_info = obtener_sector(sector_id)
            sector_nombre = sector_info["nombre"] if sector_info else sector_id
        else:
            sector_nombre = "General"
        excel_bytes = exportar_matriz_excel(df_excel)
        st.download_button(
            "📊 Matriz Interoperabilidad (.xlsx)",
            data=excel_bytes,
            file_name=f"matriz_interoperabilidad{'_' + sector_id if sector_id else ''}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="matriz_xlsx",
            type="primary",
        )
    with col_c:
        st.info("💡 **Uso recomendado:** Comparte esta matriz con equipos técnicos para evaluar brechas antes de generar propuestas específicas.")

    st.markdown("</div>", unsafe_allow_html=True)

    # ── Tab Guía (Enriquecida con ISO 19011 e ISO 23894) ────────────────────────
