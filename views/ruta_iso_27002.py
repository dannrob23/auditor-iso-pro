from views.common import *


def render(usuario, nombre, proveedor, modelo, temperatura, usar_rag, opcion_menu):
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 📋 Ruta de Auditoría ISO 27002")
    st.markdown("Guía paso a paso desde el Gap Analysis hasta el establecimiento de controles ISO 27002, adaptada a cualquier empresa sin importar su tamaño o sector.")

    tab1, tab2, tab3 = st.tabs(["🏢 Configurar Empresa", "🗺️ Mapa de Etapas", "📋 Controles y Plan"])

    with tab1:
        # Seleccionar o crear auditoría
        auditorias = obtener_auditorias(usuario)
        opciones_aud = ["➕ Nueva auditoría"] + [f"#{a['id']} - {a.get('nombre_empresa', 'Sin nombre')} ({a.get('sector', '?')})" for a in auditorias]
        aud_seleccionada = st.selectbox("Auditoría", opciones_aud, key="aud_select", index=0)

        if aud_seleccionada == "➕ Nueva auditoría":
            if st.button("🚀 Iniciar nueva auditoría"):
                nombre_empresa = st.session_state.get("aud_nombre", "")
                sector = st.session_state.get("aud_sector", "")
                tamano = st.session_state.get("aud_tamano", "")
                if not nombre_empresa or not sector or not tamano:
                    st.warning("Completa todos los campos antes de iniciar.")
                else:
                    aid = crear_auditoria(usuario, sector, tamano, nombre_empresa)
                    st.session_state["auditoria_actual_id"] = aid
                    st.success(f"Auditoría #{aid} creada. Ve al mapa de etapas para comenzar.")
                    st.rerun()

            with st.form("form_auditoria"):
                st.text_input("Nombre de la empresa", key="aud_nombre")
                st.selectbox("Sector", SECTORES_EMPRESA, key="aud_sector", index=10)
                st.selectbox("Tamaño de la empresa", TAMANOS_EMPRESA, key="aud_tamano", index=0)
                st.form_submit_button("Guardar datos")
        else:
            # Cargar auditoría existente
            idx = int(aud_seleccionada.split(" - ")[0].lstrip("#"))
            st.session_state["auditoria_actual_id"] = idx
            aud = obtener_auditoria(idx)
            if aud:
                st.markdown(f"**Empresa:** {aud.get('nombre_empresa', '—')}")
                st.markdown(f"**Sector:** {aud.get('sector', '—')}")
                st.markdown(f"**Tamaño:** {aud.get('tamano_empresa', '—')}")
                st.markdown(f"**Etapa actual:** {aud.get('etapa_actual', 'gap_analysis')}")
                if st.button("🗑️ Eliminar esta auditoría"):
                    conn = sqlite3.connect("data/reportes.db")
                    conn.execute("DELETE FROM auditoria_progreso WHERE id = ?", (idx,))
                    conn.commit()
                    conn.close()
                    st.success("Auditoría eliminada.")
                    st.session_state.pop("auditoria_actual_id", None)
                    st.rerun()

    with tab2:
        aid = st.session_state.get("auditoria_actual_id")
        if not aid:
            st.info("Primero configura una empresa en la pestaña '🏢 Configurar Empresa'.")
        else:
            aud = obtener_auditoria(aid)
            if not aud:
                st.error("Auditoría no encontrada.")
            else:
                # Mostrar barra de progreso visual
                etapas_ids = [e["id"] for e in ETAPAS]
                etapa_actual = aud.get("etapa_actual", "gap_analysis")
                try:
                    progreso_actual = etapas_ids.index(etapa_actual)
                except ValueError:
                    progreso_actual = 0
                pct = (progreso_actual / (len(etapas_ids) - 1)) * 100

                st.markdown("### 📊 Progreso General")
                st.progress(int(pct), text=f"Etapa {progreso_actual + 1} de {len(etapas_ids)}")

                # Mostrar cada etapa con su estado
                st.markdown("### 🗺️ Etapas de la Auditoría")
                cols = st.columns(len(ETAPAS))
                for i, etapa in enumerate(ETAPAS):
                    icono = etapa_icono_estado(etapa["id"], aud)
                    with cols[i]:
                        st.markdown(f"#### {icono}")
                        st.markdown(f"**{etapa['nombre']}**")
                        st.caption(etapa["descripcion"])
                        if i < progreso_actual:
                            st.info("✅ Completado")
                        elif i == progreso_actual:
                            st.info("⏳ En curso")
                        else:
                            st.info("⬜ Pendiente")

                # Botón para avanzar etapa
                st.markdown("---")
                st.markdown(f"**Etapa actual:** {ETAPAS[progreso_actual]['nombre']}")

                if etapa_actual == "gap_analysis":
                    st.info("Ejecuta un Gap Analysis en las secciones 📄 Auditoría de IA o ✏️ Texto Libre para avanzar a la siguiente etapa.")
                    st.markdown("**Resultado del Gap Analysis:**")
                    reportes = listar_reportes(limit=5)
                    if reportes:
                        for r in reportes:
                            if st.button(f"Cargar reporte #{r['id']} ({r['nombre_doc']})", key=f"gap_load_{r['id']}"):
                                actualizar_auditoria(aid, gap_analysis_id=r["id"])
                                aud["gap_analysis_id"] = r["id"]
                                aud = avanzar_etapa(aud, "gap_analysis")
                                actualizar_auditoria(aid, etapa_actual=aud["etapa_actual"], estado_json=aud["estado_json"])
                                st.success("¡Gap Analysis completado! Avanza a la siguiente etapa.")
                                st.rerun()
                    else:
                        st.warning("No hay reportes de Gap Analysis aún. Usa 📄 Auditoría de IA o ✏️ Texto Libre para crear uno.")
                else:
                    if st.button("✅ Marcar etapa como completada y avanzar"):
                        aud = avanzar_etapa(aud, etapa_actual)
                        actualizar_auditoria(aid, etapa_actual=aud["etapa_actual"], estado_json=aud["estado_json"])
                        st.success("¡Etapa completada! Avanza a la siguiente.")
                        st.rerun()

    with tab3:
        aid = st.session_state.get("auditoria_actual_id")
        if not aid:
            st.info("Primero configura una empresa y avanza a la etapa de controles.")
        else:
            aud = obtener_auditoria(aid)
            if not aud:
                st.error("Auditoría no encontrada.")
            else:
                # Cargar o generar controles recomendados
                if "controles_actuales" not in st.session_state or st.session_state.get("controles_aid") != aid:
                    sector = aud.get("sector", "Otro")
                    tamano = aud.get("tamano_empresa", "PYME (< 50 empleados)")
                    brechas = []
                    gap_id = aud.get("gap_analysis_id")
                    if gap_id:
                        r = obtener_reporte(gap_id)
                        if r and r.get("resultado"):
                            for line in r["resultado"].splitlines():
                                line_s = line.strip()
                                if line_s.startswith("|") and "❌" in line_s:
                                    cells = [c.strip() for c in line_s.split("|") if c.strip()]
                                    if cells:
                                        brechas.append(cells[0])
                    ctrls = obtener_controles_recomendados(sector, tamano, brechas if brechas else None)
                    st.session_state["controles_actuales"] = ctrls
                    st.session_state["controles_aid"] = aid
                    plan = generar_plan_implementacion(ctrls)
                    st.session_state["plan_actual"] = plan
                    # Persistir
                    actualizar_auditoria(aid, controles_json=ctrls, plan_json=plan)

                ctrls = st.session_state.get("controles_actuales", [])
                plan = st.session_state.get("plan_actual", {})

                st.markdown("### 🎯 Controles Recomendados")
                st.markdown(f"**Total:** {len(ctrls)} controles | "
                            f"🔴 Alta: {plan.get('alta_prioridad', 0)} | "
                            f"🟡 Media: {plan.get('media_prioridad', 0)} | "
                            f"🟢 Baja: {plan.get('baja_prioridad', 0)}")

                # Mostrar controles en un data_editor editable
                df_ctrls = pd.DataFrame(ctrls)
                if not df_ctrls.empty:
                    edited = st.data_editor(
                        df_ctrls,
                        column_config={
                            "id": st.column_config.TextColumn("Control", width="small", disabled=True),
                            "nombre": st.column_config.TextColumn("Nombre", width="medium", disabled=True),
                            "categoria": st.column_config.TextColumn("Categoría", width="small", disabled=True),
                            "prioridad": st.column_config.TextColumn("Prioridad", width="small", disabled=True),
                            "tiempo_estimado": st.column_config.TextColumn("Tiempo", width="small", disabled=True),
                            "descripcion": st.column_config.TextColumn("Descripción", width="large", disabled=True),
                            "responsable_sugerido": st.column_config.TextColumn("Responsable", width="medium"),
                            "estado": st.column_config.SelectboxColumn(
                                "Estado",
                                options=["pendiente", "en_progreso", "completado"],
                                width="small",
                            ),
                        },
                        hide_index=True,
                        use_container_width=True,
                        key="ctrl_editor",
                    )
                    # Guardar ediciones
                    ctrls_editados = edited.to_dict("records")
                    st.session_state["controles_actuales"] = ctrls_editados
                    actualizar_auditoria(aid, controles_json=ctrls_editados)

                # Generar plan de implementación
                st.markdown("---")
                st.markdown("### 📋 Plan de Implementación")
                for fase in plan.get("fases", []):
                    with st.expander(f"📁 {fase['fase']} — {fase['duracion_estimada']}"):
                        for ctrl in fase.get("controles", []):
                            icon_prioridad = {"alta": "🔴", "media": "🟡", "baja": "🟢"}.get(ctrl["prioridad"], "⚪")
                            st.markdown(f"{icon_prioridad} **{ctrl['id']}**: {ctrl['nombre']} — {ctrl['tiempo_estimado']}")
                            st.caption(f"Responsable sugerido: {ctrl.get('responsable_sugerido', '')}")

                # Exportar reporte
                    st.markdown("---")
                if st.button("📄 Generar reporte Markdown"):
                    md = generar_reporte_markdown(aud, ctrls if isinstance(ctrls, list) else [],
                                                  plan, aud.get("nombre_empresa", ""),
                                                  aud.get("sector", ""), aud.get("tamano_empresa", ""))
                    st.download_button("⬇️ (.md)", data=md,
                                       file_name=f"plan_implementacion_{aud.get('nombre_empresa', 'empresa')}.md",
                                       mime="text/markdown")
                    etapas_ids = [e["id"] for e in ETAPAS]
                    etapa_actual = aud.get("etapa_actual", "gap_analysis")
                    try:
                        progreso_actual = etapas_ids.index(etapa_actual)
                    except ValueError:
                        progreso_actual = 0
                    pct = (progreso_actual / (len(etapas_ids) - 1)) * 100
                    from pdf_export import generar_pdf
                    pdf_bytes = generar_pdf(md, "Plan Implementación", usuario, "", "ISO 27002", {"cumplimiento_pct": pct})
                    st.download_button("📄 PDF", data=pdf_bytes,
                                       file_name=f"plan_implementacion_{aud.get('nombre_empresa', 'empresa')}.pdf",
                                       mime="application/pdf")
                    import io, openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    font_family = "Arial"
                    thin_border = Border(
                        left=Side(style="thin", color="CCCCCC"),
                        right=Side(style="thin", color="CCCCCC"),
                        top=Side(style="thin", color="CCCCCC"),
                        bottom=Side(style="thin", color="CCCCCC"),
                    )
                    purple_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
                    white_font = Font(name=font_family, bold=True, color="FFFFFF", size=11)
                    header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
                    alt_fill = PatternFill(start_color="F3F0FF", end_color="F3F0FF", fill_type="solid")

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Plan Implementación"

                    # ── Portada / Título ──────────────────────────────────────────
                    ws.merge_cells("A1:E1")
                    ws["A1"] = "PLAN DE IMPLEMENTACIÓN ISO 27002"
                    ws["A1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
                    ws["A1"].fill = purple_fill
                    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 40

                    # ── Metadatos ─────────────────────────────────────────────────
                    row = 3
                    ws.cell(row=row, column=1, value="Empresa:").font = Font(name=font_family, bold=True, size=10)
                    ws.cell(row=row, column=2, value=aud.get("nombre_empresa", "")).font = Font(name=font_family, size=10)
                    row += 1
                    ws.cell(row=row, column=1, value="Sector:").font = Font(name=font_family, bold=True, size=10)
                    ws.cell(row=row, column=2, value=aud.get("sector", "")).font = Font(name=font_family, size=10)
                    row += 1
                    ws.cell(row=row, column=1, value="Tamaño:").font = Font(name=font_family, bold=True, size=10)
                    ws.cell(row=row, column=2, value=aud.get("tamano_empresa", "")).font = Font(name=font_family, size=10)

                    # ── Resumen ────────────────────────────────────────────────────
                    row += 2
                    ws.cell(row=row, column=1, value="RESUMEN").font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
                    ws.cell(row=row, column=1).fill = purple_fill
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                    for c in range(1, 6):
                        ws.cell(row=row, column=c).fill = purple_fill
                    row += 1
                    for label, key in [("Total controles", "total_controles"), ("Alta prioridad", "alta_prioridad"),
                                       ("Media prioridad", "media_prioridad"), ("Baja prioridad", "baja_prioridad")]:
                        ws.cell(row=row, column=1, value=label).font = Font(name=font_family, bold=True, size=10)
                        ws.cell(row=row, column=2, value=plan.get(key, 0)).font = Font(name=font_family, size=10)
                        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
                        row += 1

                    # ── Fases ─────────────────────────────────────────────────────
                    row += 1
                    for fase in plan.get("fases", []):
                        ws.cell(row=row, column=1, value=fase.get("fase", "")).font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                        for c in range(1, 6):
                            ws.cell(row=row, column=c).fill = header_fill
                        row += 1
                        headers = ["Control", "Nombre", "Prioridad", "Tiempo", "Responsable"]
                        for ci, h in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=ci, value=h)
                            cell.font = Font(name=font_family, bold=True, color="FFFFFF", size=9)
                            cell.fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        row += 1
                        for idx, item in enumerate(fase.get("controles", [])):
                            for ci, key in enumerate(["id", "nombre", "prioridad", "tiempo_estimado", "responsable_sugerido"], 1):
                                val = item.get(key, "TI") if key != "responsable_sugerido" else item.get("responsable_sugerido", "TI")
                                cell = ws.cell(row=row, column=ci, value=val)
                                cell.font = Font(name=font_family, size=9)
                                cell.border = thin_border
                                if idx % 2 == 1:
                                    cell.fill = alt_fill
                            ws.row_dimensions[row].height = 20
                            row += 1
                        row += 1

                    # ── Ancho de columnas ─────────────────────────────────────────
                    ws.column_dimensions["A"].width = 18
                    ws.column_dimensions["B"].width = 35
                    ws.column_dimensions["C"].width = 14
                    ws.column_dimensions["D"].width = 16
                    ws.column_dimensions["E"].width = 22

                    bus = io.BytesIO()
                    wb.save(bus)
                    bus.seek(0)
                    st.download_button("📊 Excel", data=bus,
                                       file_name=f"plan_implementacion_{aud.get('nombre_empresa', 'empresa')}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # ── Descargar plantilla de informe profesional ─────────────────
                st.markdown("---")
                st.markdown("### 📑 Plantilla de Informe de Auditoría")
                st.caption("Descarga una plantilla reutilizable para informes de auditoría ISO 27001.")
                col_md_t, col_xl_t = st.columns(2)
                with col_md_t:
                    if st.button("📝 Plantilla (.md)", key="btn_plantilla"):
                        from pdf_export import generar_plantilla_informe_markdown
                        plantilla = generar_plantilla_informe_markdown(
                            aud.get("nombre_empresa", ""),
                            aud.get("sector", ""),
                            aud.get("tamano_empresa", ""),
                            usuario, "", "ISO 27002",
                        )
                        st.download_button("⬇️ .md", data=plantilla,
                                           file_name="plantilla_informe_auditoria_iso27001.md", mime="text/markdown",
                                           key="dl_plantilla")
                with col_xl_t:
                    if st.button("📊 Plantilla (.xlsx)", key="btn_plantilla_xl"):
                        import io, openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        font_family = "Arial"
                        thin_border = Border(
                            left=Side(style="thin", color="CCCCCC"),
                            right=Side(style="thin", color="CCCCCC"),
                            top=Side(style="thin", color="CCCCCC"),
                            bottom=Side(style="thin", color="CCCCCC"),
                        )
                        purple_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
                        header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
                        alt_fill = PatternFill(start_color="F3F0FF", end_color="F3F0FF", fill_type="solid")

                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Plantilla Informe"

                        ws.merge_cells("A1:F1")
                        ws["A1"] = "PLANTILLA INFORME DE AUDITORÍA ISO/IEC 27001:2022"
                        ws["A1"].font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
                        ws["A1"].fill = purple_fill
                        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                        ws.row_dimensions[1].height = 35

                        row = 3
                        meta = [
                            ("Organización:", aud.get("nombre_empresa", "")),
                            ("Sector:", aud.get("sector", "")),
                            ("Tamaño:", aud.get("tamano_empresa", "")),
                            ("Auditor:", usuario),
                            ("Fecha:", datetime.now().strftime("%d/%m/%Y")),
                        ]
                        for label, val in meta:
                            ws.cell(row=row, column=1, value=label).font = Font(name=font_family, bold=True, size=10)
                            ws.cell(row=row, column=2, value=val).font = Font(name=font_family, size=10)
                            row += 1

                        row += 1
                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                        ws.cell(row=row, column=1, value="CONTROLES RECOMENDADOS").font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
                        for c in range(1, 7):
                            ws.cell(row=row, column=c).fill = header_fill
                        row += 1

                        ctrl_headers = ["Control", "Nombre", "Prioridad", "Tiempo", "Responsable", "Estado"]
                        for ci, h in enumerate(ctrl_headers, 1):
                            cell = ws.cell(row=row, column=ci, value=h)
                            cell.font = Font(name=font_family, bold=True, color="FFFFFF", size=9)
                            cell.fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        row += 1

                        for idx, c in enumerate(ctrls):
                            vals = [c.get(k, "") for k in ["id", "nombre", "prioridad", "tiempo_estimado", "responsable_sugerido", "estado"]]
                            for ci, v in enumerate(vals, 1):
                                cell = ws.cell(row=row, column=ci, value=v)
                                cell.font = Font(name=font_family, size=9)
                                cell.border = thin_border
                                if idx % 2 == 1:
                                    cell.fill = alt_fill
                            row += 1

                        row += 1
                        for fase in plan.get("fases", []):
                            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                            ws.cell(row=row, column=1, value=fase.get("fase", "")).font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
                            for c in range(1, 7):
                                ws.cell(row=row, column=c).fill = header_fill
                            row += 1
                            fase_headers = ["Control", "Nombre", "Prioridad", "Tiempo", "Responsable", ""]
                            for ci, h in enumerate(fase_headers, 1):
                                cell = ws.cell(row=row, column=ci, value=h)
                                cell.font = Font(name=font_family, bold=True, color="FFFFFF", size=9)
                                cell.fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
                                cell.border = thin_border
                            row += 1
                            for item in fase.get("controles", []):
                                vals = [item.get(k, "") for k in ["id", "nombre", "prioridad", "tiempo_estimado", "responsable_sugerido", ""]]
                                for ci, v in enumerate(vals, 1):
                                    ws.cell(row=row, column=ci, value=v).font = Font(name=font_family, size=9)
                                    ws.cell(row=row, column=ci).border = thin_border
                                row += 1
                            row += 1

                        ws.column_dimensions["A"].width = 14
                        ws.column_dimensions["B"].width = 32
                        ws.column_dimensions["C"].width = 14
                        ws.column_dimensions["D"].width = 14
                        ws.column_dimensions["E"].width = 20
                        ws.column_dimensions["F"].width = 14

                        buf = io.BytesIO()
                        wb.save(buf)
                        buf.seek(0)
                        st.download_button("⬇️ .xlsx", data=buf,
                                           file_name=f"plantilla_informe_{aud.get('nombre_empresa', 'empresa')}.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key="dl_plantilla_xl")

    st.markdown("</div>", unsafe_allow_html=True)

    # ── Tab RAG (NotebookLM) ──────────────────────────────────────────────────────
