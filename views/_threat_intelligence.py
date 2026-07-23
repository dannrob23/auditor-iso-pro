from views.common import *


def render(usuario, nombre, proveedor, modelo, temperatura, usar_rag, opcion_menu):
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 🛡️ Threat Intelligence — Análisis ISO 27001 de Ransomware")
    st.markdown("""
    **Feed automatizado de ransomware + Matriz de Riesgos ISO 27001:2022**
    
    Cada ataque se analiza como Auditor ISO 27001: se asignan probabilidad, severidad, nivel de riesgo
    y **controles ISO 27002:2022** que lo habrían prevenido.
    
    Fuente: [ransomware.live](https://www.ransomware.live/) · 29.520+ víctimas · 357 grupos
    """)
    
    # ── Stats rápidos ──
    try:
        from ingestor.ransomware_feed import consultar_estadisticas as ce
        stats = ce()
        if stats:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("🌍 Víctimas totales", f"{stats.get('victims',0):,}")
            c2.metric("👥 Grupos activos", stats.get('groups',0))
            c3.metric("📰 Cobertura prensa", stats.get('press',0))
            c4.metric("🕐 Última actualización", "Hoy")
    except:
        pass
    
    st.markdown("---")
    
    # ── Filtros ──
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    with col_f1:
        solo_co = st.checkbox("🇨🇴 Solo Colombia", value=True, key="ransom_co")
    with col_f2:
        sectores_opts = ["Todos", "Servicios Financieros", "Salud", "Manufactura",
                         "Tecnología", "Gobierno", "Energía", "Comercio", "Educación",
                         "Transporte", "Legal", "Construcción", "Agricultura y Alimentos"]
        sector_filtro = st.selectbox("Sector", sectores_opts, key="ransom_sector")
    with col_f3:
        grupo_buscar = st.text_input("Grupo específico (opcional)", placeholder="qilin, lockbit...", key="ransom_grupo")
    with col_f4:
        cantidad = st.number_input("Resultados", 1, 50, 10, key="ransom_cant")
    
    # ── Botón consultar ──
    if st.button("🔄 Consultar amenazas actuales", type="primary", use_container_width=True):
        with st.spinner("🔍 Consultando ransomware.live..."):
            import ingestor.ransomware_feed as rf
            
            if grupo_buscar:
                info_grupo = rf.consultar_grupo(grupo_buscar)
                if info_grupo:
                    st.success(f"✅ Inteligencia del grupo **{grupo_buscar}** obtenida")
                    with st.expander("📋 Detalles del grupo", expanded=True):
                        st.json(info_grupo)
                else:
                    st.warning(f"No se encontró información para el grupo '{grupo_buscar}'")
            
            mapa_sectores = {
                "Servicios Financieros": "Financial Services",
                "Salud": "Healthcare", "Manufactura": "Manufacturing",
                "Tecnología": "Technology", "Gobierno": "Government",
                "Energía": "Energy", "Comercio": "Retail", "Educación": "Education",
                "Transporte": "Transportation", "Legal": "Legal",
                "Construcción": "Construction", "Agricultura y Alimentos": "Agriculture and Food Production",
            }
            sector_api = mapa_sectores.get(sector_filtro, sector_filtro)
            
            if solo_co and sector_api and sector_filtro != "Todos":
                import urllib.parse
                url = f"{rf.API_BASE}/victims/?country=CO&sector={urllib.parse.quote(sector_api)}"
                req = urllib.request.Request(url, headers=rf.API_HEADERS)
                try:
                    resp = urllib.request.urlopen(req, timeout=15)
                    data = json.loads(resp.read().decode())
                    victimas = data.get("victims", [])
                    ataques = []
                    for item in victimas[:cantidad]:
                        ataques.append({
                            "grupo": item.get("group", "Desconocido"),
                            "victima": item.get("victim", ""),
                            "pais": "CO",
                            "sector": item.get("activity", "N/A"),
                            "fecha": item.get("discovered", "")[:10],
                            "website": item.get("website", ""),
                            "data_size": str(item.get("data_size") or ""),
                            "descripcion": item.get("description", "")[:200],
                        })
                except:
                    ataques = []
            elif solo_co:
                ataques = rf.consultar_por_pais("CO", cantidad)
            elif sector_filtro != "Todos" and sector_api:
                ataques = rf.consultar_por_sector(sector_api, cantidad)
            else:
                ataques = rf.consultar_victimas_recientes(cantidad)
            
            if not ataques:
                st.info("No se encontraron víctimas con esos filtros.")
            else:
                solo_sector = None if sector_filtro == "Todos" else sector_api
                riesgos = rf.generar_matriz_riesgos(ataques, solo_colombia=False, solo_sector=solo_sector)
                st.session_state["riesgos_ransom"] = riesgos
                st.session_state["sector_filtro_ransom"] = sector_filtro
                st.session_state["controles_ransom"] = rf.sugerir_controles_prioritarios(
                    sector_api if sector_filtro != "Todos" else "Financial Services"
                )
                st.rerun()
    
    # ── Mostrar resultados ──
    if "riesgos_ransom" in st.session_state:
        riesgos = st.session_state["riesgos_ransom"]
        sector_filtro_mostrar = st.session_state.get("sector_filtro_ransom", "Todos")
        controles_rec = st.session_state.get("controles_ransom", [])
        from ingestor.ransomware_feed import estadisticas_por_sector as es
        
        st.success(f"✅ {len(riesgos)} ataques analizados con metodología ISO 27001")
        
        # KPIs
        niveles = [r["nivel_riesgo"] for r in riesgos]
        extremos = sum(1 for n in niveles if n == "EXTREMO")
        altos = sum(1 for n in niveles if n == "ALTO")
        medios = sum(1 for n in niveles if n == "MEDIO")
        score_prom = sum(r["score"] for r in riesgos) / len(riesgos) if riesgos else 0
        
        mk1, mk2, mk3, mk4 = st.columns(4)
        mk1.metric("🔴 Extremos", extremos)
        mk2.metric("🟠 Altos", altos)
        mk3.metric("🟡 Medios", medios)
        mk4.metric("📊 Score Promedio", f"{score_prom:.0f}/25")
        
        st.markdown("#### 📋 Tabla de Riesgos ISO 27001")
        df_data = []
        for r in riesgos:
            df_data.append({
                "Grupo": r["grupo"], "Víctima": r["victima"], "País": r["pais"],
                "Sector": r["sector"], "Fecha": r["fecha"],
                "Probabilidad": r["probabilidad"], "Severidad": r["severidad"],
                "Score": f'{r["score"]}/25', "Nivel": r["nivel_riesgo"],
                "Descripción del Ataque": r.get("descripcion_ataque", r.get("descripcion", "No disponible")),
                "Análisis ISO 27001": r.get("justificacion", "No disponible"),
            })
        st.dataframe(pd.DataFrame(df_data), use_container_width=True, hide_index=True, column_config={
            "Descripción del Ataque": st.column_config.TextColumn(width="medium"),
            "Análisis ISO 27001": st.column_config.TextColumn(width="large"),
        })
        
        st.markdown("#### 📊 Afectación por sector")
        stats_sector = es(riesgos)
        if stats_sector:
            cols = st.columns(min(len(stats_sector), 4))
            for i, (sector, data) in enumerate(sorted(stats_sector.items(), key=lambda x: -x[1]["conteo"])):
                with cols[i % 4]:
                    color = "🔴" if data["nivel_promedio"] == "EXTREMO" else "🟠" if data["nivel_promedio"] == "ALTO" else "🟡"
                    st.metric(f"{color} {sector}", f'{data["conteo"]} ataques', data["nivel_promedio"])
        
        st.markdown("#### 🛡️ Controles ISO 27002:2022 Recomendados")
        if controles_rec:
            cols_iso = st.columns(len(controles_rec))
            for i, c in enumerate(controles_rec):
                with cols_iso[i]:
                    st.markdown(f"""
                    <div style="background:rgba(88,200,242,0.08);border:1px solid rgba(88,200,242,0.15);
                                border-radius:8px;padding:12px;text-align:center;">
                        <div style="color:#58C8F2;font-size:20px;font-weight:800;">{c['codigo']}</div>
                        <div style="color:rgba(255,255,255,0.6);font-size:11px;margin-top:4px;">{c['nombre']}</div>
                    </div>
                    """, unsafe_allow_html=True)
        
        # ── Botón Generar Excel Profesional ──
        st.markdown("---")
        col_ex1, col_ex2, col_ex3 = st.columns([2, 1, 1])
        with col_ex1:
            if st.button("📊 Generar Informe ISO 27001 (Excel profesional)", type="primary", key="btn_gen_excel"):
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
                from openpyxl.utils import get_column_letter
                import tempfile, os
                
                wb = Workbook()
                
                # ── Hoja 1: RESUMEN EJECUTIVO ──
                ws_res = wb.active
                ws_res.title = "Resumen Ejecutivo"
                ws_res.sheet_properties.tabColor = "1a1a2e"
                
                title_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
                title_font = Font(name="Century Gothic", size=16, bold=True, color="FFFFFF")
                sub_font = Font(name="Century Gothic", size=10, italic=True, color="888888")
                
                ws_res.merge_cells("A1:F1")
                ws_res["A1"] = "INFORME DE THREAT INTELLIGENCE — ANÁLISIS ISO 27001"
                ws_res["A1"].font = title_font
                ws_res["A1"].fill = title_fill
                ws_res["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws_res.row_dimensions[1].height = 40
                
                ws_res.merge_cells("A2:F2")
                ws_res["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} · Fuente: ransomware.live · {len(riesgos)} ataques analizados"
                ws_res["A2"].font = sub_font
                ws_res["A2"].alignment = Alignment(horizontal="center")
                
                # Métricas clave
                kpi_data = [
                    ("Ataques Analizados", len(riesgos), "🔴" if extremos > 0 else "🟠"),
                    ("Nivel Extremo", extremos, "🔴" if extremos > 0 else "✅"),
                    ("Nivel Alto", altos, "🟠" if altos > 0 else "✅"),
                    ("Score Promedio", f"{score_prom:.0f}/25", "📊"),
                    ("Controles Recomendados", len(controles_rec), "🛡️"),
                ]
                kpi_header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
                kpi_header_font = Font(name="Century Gothic", size=10, bold=True, color="FFFFFF")
                kpi_font = Font(name="Century Gothic", size=12, bold=True, color="E0E0E0")
                kpi_label_font = Font(name="Century Gothic", size=9, color="AAAAAA")
                thin_border = Border(left=Side(style="thin", color="334155"), right=Side(style="thin", color="334155"),
                                     top=Side(style="thin", color="334155"), bottom=Side(style="thin", color="334155"))
                
                for i, (label, value, icon) in enumerate(kpi_data):
                    col = i + 1
                    c_label = ws_res.cell(row=4, column=col, value=f"{icon} {label}")
                    c_label.font = kpi_header_font
                    c_label.fill = kpi_header_fill
                    c_label.alignment = Alignment(horizontal="center", vertical="center")
                    c_label.border = thin_border
                    
                    c_val = ws_res.cell(row=5, column=col, value=value)
                    c_val.font = kpi_font
                    c_val.alignment = Alignment(horizontal="center", vertical="center")
                    c_val.fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
                    c_val.border = thin_border
                    ws_res.row_dimensions[5].height = 35
                
                # Conclusión
                ws_res.merge_cells("A7:F7")
                ws_res["A7"] = "CONCLUSIÓN DEL ANÁLISIS"
                ws_res["A7"].font = Font(name="Century Gothic", size=11, bold=True, color="58C8F2")
                
                conclusion = (
                    f"Se analizaron {len(riesgos)} ataques de ransomware bajo la metodología ISO 27001:2022. "
                    f"De estos, {extremos} fueron clasificados como EXTREMOS, {altos} como ALTOS y {medios} como MEDIOS. "
                    f"El score de riesgo promedio es {score_prom:.0f}/25, lo que indica un nivel de amenaza "
                    f"{'CRÍTICO' if score_prom > 18 else 'ALTO' if score_prom > 12 else 'MODERADO'} para el sector evaluado. "
                    f"Se recomienda implementar prioritariamente los controles ISO 27002:2022 listados en la hoja 'Controles ISO 27002'."
                )
                ws_res.merge_cells("A8:F8")
                ws_res["A8"] = conclusion
                ws_res["A8"].font = Font(name="Century Gothic", size=9, color="CCCCCC")
                ws_res["A8"].alignment = Alignment(wrap_text=True, vertical="top")
                ws_res.row_dimensions[8].height = 60
                
                for col_let in ["A","B","C","D","E","F"]:
                    ws_res.column_dimensions[col_let].width = 22
                
                # ── Hoja 2: MATRIZ DE RIESGOS ──
                ws = wb.create_sheet("Matriz de Riesgos")
                ws.sheet_properties.tabColor = "dc2626"
                
                headers_r = ["#", "Grupo", "Víctima", "País", "Sector", "Fecha",
                             "Probabilidad\n(1-5)", "Severidad\n(1-5)", "Score\n/25", "Nivel de Riesgo",
                             "Descripción del Ataque", "Análisis ISO 27001", "Controles Aplicables"]
                header_fill = PatternFill(start_color="1B273D", end_color="1B273D", fill_type="solid")
                header_font = Font(name="Century Gothic", size=9, bold=True, color="FFFFFF")
                
                for col, h in enumerate(headers_r, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border = thin_border
                ws.row_dimensions[1].height = 35
                
                fills_nivel = {
                    "EXTREMO": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
                    "ALTO": PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid"),
                    "MEDIO": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
                    "BAJO": PatternFill(start_color="44BB66", end_color="44BB66", fill_type="solid"),
                }
                data_fill_even = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
                data_fill_odd = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
                data_font = Font(name="Century Gothic", size=8, color="E0E0E0")
                
                for row_idx, r in enumerate(riesgos, 2):
                    row_data = [
                        row_idx - 1, r["grupo"], r["victima"], r["pais"], r["sector"], r["fecha"],
                        r["probabilidad"], r["severidad"], f'{r["score"]}/25', r["nivel_riesgo"],
                        r.get("descripcion_ataque", r.get("descripcion", "No disponible")),
                        r.get("justificacion", "No disponible"),
                        "; ".join([c["codigo"] for c in controles_rec[:4]])
                    ]
                    for col, val in enumerate(row_data, 1):
                        c = ws.cell(row=row_idx, column=col, value=val)
                        c.font = data_font
                        c.border = thin_border
                        c.alignment = Alignment(wrap_text=True, vertical="top")
                        c.fill = data_fill_even if row_idx % 2 == 0 else data_fill_odd
                        # Color por nivel de riesgo
                        if col == 10:
                            nivel_fill = fills_nivel.get(val)
                            if nivel_fill:
                                c.fill = nivel_fill
                                c.font = Font(name="Century Gothic", size=9, bold=True, color="000000" if val in ("MEDIO","BAJO") else "FFFFFF")
                    ws.row_dimensions[row_idx].height = 30
                
                anchors = {"A": 6, "B": 22, "C": 28, "D": 8, "E": 22, "F": 14, "G": 12, "H": 12, "I": 10, "J": 16, "K": 45, "L": 55, "M": 30}
                for col_let, w in anchors.items():
                    ws.column_dimensions[col_let].width = w
                
                # Leyenda en Matriz
                leg_row = len(riesgos) + 4
                ws.cell(row=leg_row, column=1, value="LEYENDA DE RIESGO ISO 27001:").font = Font(bold=True, color="58C8F2")
                for i, (nivel, fill) in enumerate(fills_nivel.items()):
                    c = ws.cell(row=leg_row + 1, column=i * 2 + 1, value=nivel)
                    c.fill = fill
                    c.font = Font(bold=True, size=9, color="000000" if nivel in ("MEDIO","BAJO") else "FFFFFF")
                    c.border = thin_border
                
                # ── Hoja 3: CONTROLES ISO 27002 ──
                ws3 = wb.create_sheet("Controles ISO 27002")
                ws3.sheet_properties.tabColor = "58C8F2"
                
                ctrl_header_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
                ctrl_headers = ["Código ISO 27002", "Nombre del Control", "Dominio", "Descripción", "Prioridad", "Sector Recomendado"]
                for col, h in enumerate(ctrl_headers, 1):
                    c = ws3.cell(row=1, column=col, value=h)
                    c.fill = ctrl_header_fill
                    c.font = Font(name="Century Gothic", size=9, bold=True, color="FFFFFF")
                    c.alignment = Alignment(horizontal="center", wrap_text=True)
                    c.border = thin_border
                ws3.row_dimensions[1].height = 30
                
                prioridad = 1
                for i, ctrl in enumerate(controles_rec, 2):
                    ws3.cell(row=i, column=1, value=ctrl["codigo"]).border = thin_border
                    ws3.cell(row=i, column=1).font = Font(name="Century Gothic", size=9, bold=True, color="58C8F2")
                    ws3.cell(row=i, column=2, value=ctrl["nombre"]).border = thin_border
                    ws3.cell(row=i, column=2).font = data_font
                    ws3.cell(row=i, column=3, value="Seguridad de la Información").border = thin_border
                    ws3.cell(row=i, column=3).font = data_font
                    ws3.cell(row=i, column=4, value="Control prioritario para mitigar ataques de ransomware en el sector evaluado.").border = thin_border
                    ws3.cell(row=i, column=4).font = data_font
                    ws3.cell(row=i, column=5, value=f"P{i-1}").border = thin_border
                    ws3.cell(row=i, column=5).font = Font(name="Century Gothic", size=10, bold=True, color="FF8800" if i <= 3 else "E0E0E0")
                    ws3.cell(row=i, column=6, value=sector_filtro_mostrar).border = thin_border
                    ws3.cell(row=i, column=6).font = data_font
                    ws3.row_dimensions[i].height = 25
                    prioridad += 1
                
                ctrl_widths = {"A": 16, "B": 30, "C": 24, "D": 45, "E": 12, "F": 22}
                for col_let, w in ctrl_widths.items():
                    ws3.column_dimensions[col_let].width = w
                
                # ── Hoja 4: ESTADÍSTICAS ──
                ws4 = wb.create_sheet("Estadísticas")
                ws4.sheet_properties.tabColor = "a78bfa"
                
                ws4.merge_cells("A1:D1")
                ws4["A1"] = "DISTRIBUCIÓN DE ATAQUES POR NIVEL DE RIESGO"
                ws4["A1"].font = Font(name="Century Gothic", size=11, bold=True, color="FFFFFF")
                ws4["A1"].fill = header_fill
                ws4["A1"].alignment = Alignment(horizontal="center")
                
                est_data = [
                    ("EXTREMO", extremos),
                    ("ALTO", altos),
                    ("MEDIO", medios),
                    ("BAJO", sum(1 for n in niveles if n == "BAJO")),
                    ("TOTAL", len(riesgos)),
                ]
                for i, (nivel, count) in enumerate(est_data, 3):
                    ws4.cell(row=i, column=1, value=nivel).font = Font(bold=True, color="E0E0E0")
                    ws4.cell(row=i, column=1).fill = fills_nivel.get(nivel, data_fill_even)
                    ws4.cell(row=i, column=1).border = thin_border
                    ws4.cell(row=i, column=2, value=count).font = Font(bold=True, size=14)
                    ws4.cell(row=i, column=2).fill = fills_nivel.get(nivel, data_fill_even)
                    ws4.cell(row=i, column=2).border = thin_border
                    pct = f"{count/len(riesgos)*100:.1f}%" if riesgos else "0%"
                    ws4.cell(row=i, column=3, value=pct).font = data_font
                    ws4.cell(row=i, column=3).fill = fills_nivel.get(nivel, data_fill_even)
                    ws4.cell(row=i, column=3).border = thin_border
                    ws4.row_dimensions[i].height = 28
                
                ws4.column_dimensions["A"].width = 20
                ws4.column_dimensions["B"].width = 15
                ws4.column_dimensions["C"].width = 15
                ws4.column_dimensions["D"].width = 20
                
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                wb.save(tmp.name)
                tmp.close()
                with open(tmp.name, 'rb') as f:
                    st.session_state["excel_datos"] = f.read()
                os.unlink(tmp.name)
                st.rerun()
        
        with col_ex2:
            if "excel_datos" in st.session_state:
                st.download_button(
                    "⬇️ Descargar Informe ISO 27001 (.xlsx)",
                    data=st.session_state["excel_datos"],
                    file_name=f"Threat_Intelligence_ISO27001_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_ransom_excel",
                    type="primary",
                    use_container_width=True,
                )
            else:
                st.info("Pulsa 'Generar Informe'\ny luego descarga aquí")
        with col_ex3:
            with st.expander("📖 Metodología"):
                st.markdown("""
                **ISO 27001:2022 — Matriz de Riesgos 5x5**
                
                - **Probabilidad:** 1 (Rara) a 5 (Muy frecuente)
                - **Severidad:** 1 (Bajo) a 5 (Crítico)
                - **Score:** Probabilidad × Severidad (1-25)
                - **Niveles:** BAJO (1-4), MEDIO (5-9), ALTO (10-15), EXTREMO (16-25)
                - **Controles:** ISO 27002:2022 mapeados a cada ataque
                """)
    
    st.markdown("</div>", unsafe_allow_html=True)

    # ── Tab Autodiagnóstico de Madurez ───────────────────────────────────────
