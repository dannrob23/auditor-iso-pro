import streamlit as st
import yaml
import os
import sys
import json as _json
import subprocess
import traceback
from pathlib import Path
from yaml.loader import SafeLoader
from dotenv import load_dotenv

st.set_page_config(
    page_title="IA-SEC Auditor Tool",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)
import streamlit_authenticator as stauth
from langchain_openai import ChatOpenAI
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage, SystemMessage, AIMessage
import pypdf
import io
from audit_logger import log_event, get_log_entries
from database import (
    guardar_reporte, listar_reportes, obtener_reporte, stats_globales, init_db,
    guardar_lote_vulnerabilidades, listar_vulnerabilidades, stats_vulnerabilidades,
    guardar_vulnerabilidad, limpiar_vulnerabilidades_antiguas, exportar_excel_respaldo,
    crear_auditoria, obtener_auditorias, obtener_auditoria, actualizar_auditoria,
)
from pdf_export import (
    generar_pdf,
    generar_pdf_propuesta_interoperabilidad,
)
from guia_auditoria import (
    ETAPAS, SECTORES_EMPRESA, TAMANOS_EMPRESA,
    obtener_controles_recomendados, generar_plan_implementacion,
    generar_reporte_markdown, etapa_icono_estado, avanzar_etapa,
)

try:
    from rag_knowledge import (
        obtener_contexto,
        indexar_documentos,
        base_conocimiento_activa,
        indexar_documentos_temporales,
        listar_uploads,
        limpiar_uploads,
        base_uploads_activa,
    )
    _RAG_AVAILABLE = True
except Exception:
    _RAG_AVAILABLE = False

    def obtener_contexto(query: str, k: int = 4) -> str:
        return ""

    def indexar_documentos():
        return 0, "RAG no disponible: faltan dependencias en el entorno."

    def base_conocimiento_activa() -> bool:
        return False
from interoperabilidad import (
    MATRIZ_INTEROPERABILIDAD, SECTORES_IC, obtener_dimensiones,
    obtener_controles_por_dimension, calcular_cobertura, obtener_sector,
    generar_resumen_interoperabilidad, generar_propuesta_ic, generar_markdown_matriz,
    generar_df_interoperabilidad, calcular_cumplimiento, exportar_matriz_excel
)
from ingestor import (
    fetch_all, fetch_avid_reports, fetch_mitre_atlas, fetch_mit_risk_repo,
    analizar_vulnerabilidad, normalize_batch, set_provider, MODEL_MAP,
    calcular_riesgo, criticidad_compuesta, resumen_estadistico, clasificar_tactica,
    CUESTIONARIO_CONTEXTO, generar_prompt_contexto,
    exportar_matriz_riesgos_ia,
)
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
from datetime import datetime, timezone, timedelta

# ── Paleta de colores de riesgo (sistema único) ──────────────────────────────
RIESGO_COLORS = {
    "EXTREMO": {"bg": "rgba(239,68,68,0.18)",  "text": "#f87171", "border": "rgba(239,68,68,0.45)",  "hex": "#f87171"},
    "ALTO":    {"bg": "rgba(234,88,12,0.18)",  "text": "#fb923c", "border": "rgba(234,88,12,0.45)",  "hex": "#fb923c"},
    "MEDIO":   {"bg": "rgba(234,179,8,0.18)",  "text": "#fbbf24", "border": "rgba(234,179,8,0.45)",  "hex": "#fbbf24"},
    "BAJO":    {"bg": "rgba(52,211,153,0.18)", "text": "#34d399", "border": "rgba(52,211,153,0.45)", "hex": "#34d399"},
}

# Zona horaria Bogotá, Colombia (UTC-5)
BOGOTA_OFFSET = timedelta(hours=-5)

def ahora_bogota():
    return datetime.now(timezone.utc) + BOGOTA_OFFSET

init_db()
load_dotenv()

def _obtener_kb_dir():
    return Path("knowledge_base")

def _auto_descargar_fuentes():
    kb = _obtener_kb_dir()
    if kb.exists() and any(kb.glob("*.pdf")):
        return
    # Intentar desde repo externo (variable KB_REPO)
    from rag_knowledge import sincronizar_fuentes_oficiales
    n = sincronizar_fuentes_oficiales()
    if n > 0:
        return
    # Fallback: ZIP en GitHub Releases
    url = "https://github.com/dannrob23/auditor-iso-pro/releases/download/knowledge/knowledge_base_pdfs.zip"
    try:
        import requests
    except Exception:
        return
    try:
        with st.spinner("Descargando fuentes oficiales..."):
            r = requests.get(url, timeout=120)
            if r.status_code == 200:
                import zipfile, io
                kb.mkdir(exist_ok=True)
                with zipfile.ZipFile(io.BytesIO(r.content)) as z:
                    for member in z.namelist():
                        if member.endswith(".pdf"):
                            target = kb / member.split("/")[-1]
                            target.write_bytes(z.read(member))
    except Exception:
        pass

_auto_descargar_fuentes()

# ── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="CID INFOSEC | Auditor de Ciberseguridad e IA",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS personalizado — Sistema de diseño CID INFOSEC ────────────────────────
# Inspirado en Linear: dark-mode-native, Inter, bordes sutiles, acento indigo
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

:root {
  --bg-canvas: #08090a;
  --bg-panel: #0f1011;
  --bg-surface: #191a1b;
  --bg-elevated: #28282c;
  --text-primary: #f7f8f8;
  --text-secondary: #d0d6e0;
  --text-tertiary: #8a8f98;
  --text-muted: #62666d;
  --accent: #5e6ad2;
  --accent-hover: #7170ff;
  --accent-light: #828fff;
  --border-subtle: rgba(255,255,255,0.05);
  --border-standard: rgba(255,255,255,0.08);
  --success: #10b981;
  --warning: #fbbf24;
  --danger: #f87171;
}

html, body, [class*="css"] {
  font-family: 'Inter', system-ui, -apple-system, 'Segoe UI', Roboto, sans-serif;
  font-feature-settings: "cv01", "ss03";
}

/* ── Fondo general ── */
.stApp {
  background: var(--bg-canvas);
  min-height: 100vh;
}

/* ── Sidebar ── */
[data-testid="stSidebar"] {
  background: var(--bg-panel);
  border-right: 1px solid var(--border-subtle);
}

/* ── Tarjetas ── */
.card {
  background: rgba(255,255,255,0.02);
  border: 1px solid var(--border-standard);
  border-radius: 8px;
  padding: 24px;
  margin-bottom: 16px;
  animation: fadeIn 0.3s ease;
}
@keyframes fadeIn {
  from { opacity: 0; transform: translateY(6px); }
  to   { opacity: 1; transform: translateY(0); }
}

/* ── Onboarding ── */
.onboarding-card {
  background: rgba(94,106,210,0.08);
  border: 1px solid rgba(94,106,210,0.25);
  border-radius: 8px;
  padding: 20px 24px;
  margin-bottom: 20px;
}
.onboarding-card h3 { color: var(--accent-light); margin-bottom: 8px; font-size: 1.1rem; font-weight: 600; }
.onboarding-step {
  display: flex; align-items: center; gap: 10px;
  padding: 6px 0; font-size: 0.88rem; color: var(--text-secondary);
}
.step-num {
  background: rgba(94,106,210,0.2); color: var(--accent-light);
  border-radius: 50%; width: 22px; height: 22px;
  display: flex; align-items: center; justify-content: center;
  font-size: 0.72rem; font-weight: 600; flex-shrink: 0;
}

/* ── Health bar ── */
.health-bar {
  display: flex; gap: 16px; flex-wrap: wrap;
  background: var(--bg-surface);
  border: 1px solid var(--border-subtle);
  border-radius: 6px;
  padding: 8px 16px;
  margin-bottom: 16px;
  align-items: center;
}
.health-item {
  display: flex; align-items: center; gap: 6px;
  font-size: 0.8rem; color: var(--text-tertiary);
}
.health-dot {
  width: 7px; height: 7px; border-radius: 50%;
  display: inline-block;
  box-shadow: 0 0 4px currentColor;
}
.dot-green  { background: var(--success); color: var(--success); }
.dot-yellow { background: var(--warning); color: var(--warning); }
.dot-red    { background: var(--danger); color: var(--danger); }

/* ── Header CID INFOSEC ── */
.cid-header {
  display: flex; align-items: center; gap: 12px;
  padding: 1.2rem 0 0.6rem 0;
}
.cid-logo {
  width: 36px; height: 36px;
  background: linear-gradient(135deg, var(--accent), var(--accent-hover));
  border-radius: 6px;
  display: flex; align-items: center; justify-content: center;
  font-size: 1.1rem; flex-shrink: 0;
  box-shadow: 0 2px 12px rgba(94,106,210,0.35);
}
.cid-title {
  font-size: 1.4rem; font-weight: 600;
  color: var(--text-primary);
  letter-spacing: -0.5px;
  margin: 0;
}
.cid-subtitle {
  font-size: 0.82rem; color: var(--text-tertiary);
  margin: 0; letter-spacing: 0.02em;
}

/* ── Badges ── */
.badge {
  display: inline-block; padding: 2px 8px;
  border-radius: 999px; font-size: 0.74rem; font-weight: 500;
}
.badge-red    { background: rgba(248,113,113,0.12); color: var(--danger); border: 1px solid rgba(248,113,113,0.3); }
.badge-yellow { background: rgba(251,191,36,0.12);  color: var(--warning); border: 1px solid rgba(251,191,36,0.3); }
.badge-green  { background: rgba(16,185,129,0.12);  color: var(--success); border: 1px solid rgba(16,185,129,0.3); }
.badge-orange { background: rgba(251,146,60,0.12);  color: #fb923c; border: 1px solid rgba(251,146,60,0.3); }

/* ── Tabla de resultados ── */
.result-table { width: 100%; border-collapse: collapse; margin-top: 16px; }
.result-table th {
  background: var(--bg-surface); color: var(--text-tertiary);
  padding: 10px 14px; text-align: left;
  font-size: 0.8rem; font-weight: 500;
  border-bottom: 1px solid var(--border-standard);
  letter-spacing: 0.02em;
}
.result-table td {
  padding: 10px 14px; border-bottom: 1px solid var(--border-subtle);
  color: var(--text-secondary); font-size: 0.85rem; vertical-align: top;
}
.result-table tr:hover td { background: rgba(255,255,255,0.02); }

/* ── Métricas ── */
.metric-box {
  background: var(--bg-surface);
  border: 1px solid var(--border-subtle);
  border-radius: 6px; padding: 16px; text-align: center;
  transition: border-color 0.2s ease;
}
.metric-box:hover { border-color: rgba(94,106,210,0.3); }
.metric-box .metric-value { font-size: 1.8rem; font-weight: 600; color: var(--accent-light); }
.metric-box .metric-label { font-size: 0.78rem; color: var(--text-muted); margin-top: 4px; }

/* ── Paginación ── */
.pag-info { color: var(--text-muted); font-size: 0.82rem; text-align: center; padding: 8px; }

/* ── Botones ── */
div.stButton > button {
  background: var(--accent);
  color: #ffffff; border: none; border-radius: 6px;
  padding: 8px 16px; font-weight: 500; font-size: 0.88rem;
  transition: all 0.2s ease; width: 100%;
}
div.stButton > button:hover {
  background: var(--accent-hover);
  transform: translateY(-1px);
  box-shadow: 0 4px 16px rgba(94,106,210,0.3);
}

/* ── Ocultar elementos nativos innecesarios ── */
#MainMenu, footer { visibility: hidden; }

/* ── Sidebar: permitir colapso nativo ── */
[data-testid="stSidebar"] {
  min-width: 280px !important;
  max-width: 280px !important;
}
[data-testid="stSidebarContent"] {
  overflow-y: auto !important;
  padding: 1rem !important;
}

/* ── Radio buttons del sidebar (navegación) ── */
[data-testid="stSidebar"] [data-testid="stRadio"] {
  gap: 0 !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] > div {
  gap: 2px !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label {
  font-size: 0.84rem !important;
  color: var(--text-tertiary) !important;
  padding: 6px 10px !important;
  border-radius: 4px !important;
  transition: all 0.15s ease !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
  background: rgba(255,255,255,0.04) !important;
  color: var(--text-secondary) !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-checked="true"] {
  background: rgba(94,106,210,0.12) !important;
  color: var(--accent-light) !important;
}

/* ── Selectbox estilo minimalista ── */
[data-testid="stSidebar"] [data-baseweb="select"] > div {
  background: var(--bg-surface) !important;
  border: 1px solid var(--border-standard) !important;
  border-radius: 6px !important;
}

/* ── Login screen ── */
.cid-login-container {
  max-width: 420px;
  margin: 4rem auto;
  padding: 2.5rem;
  background: var(--bg-panel);
  border: 1px solid var(--border-standard);
  border-radius: 12px;
  text-align: center;
}
.cid-login-logo {
  width: 56px; height: 56px;
  background: linear-gradient(135deg, var(--accent), var(--accent-hover));
  border-radius: 10px;
  display: flex; align-items: center; justify-content: center;
  font-size: 1.6rem;
  margin: 0 auto 1.2rem auto;
  box-shadow: 0 4px 24px rgba(94,106,210,0.3);
}
.cid-login-title {
  font-size: 1.5rem; font-weight: 600;
  color: var(--text-primary);
  letter-spacing: -0.5px;
  margin-bottom: 0.4rem;
}
.cid-login-subtitle {
  font-size: 0.85rem; color: var(--text-tertiary);
  margin-bottom: 2rem;
}
.cid-login-frameworks {
  font-size: 0.76rem; color: var(--text-muted);
  letter-spacing: 0.03em;
  margin-top: 1.5rem; padding-top: 1.2rem;
  border-top: 1px solid var(--border-subtle);
}

/* ── Chat input ── */
[data-testid="stChatInput"] textarea {
  background: var(--bg-surface) !important;
  border: 1px solid var(--border-standard) !important;
  border-radius: 6px !important;
  color: var(--text-primary) !important;
}

/* ── Expander ── */
[data-testid="stExpander"] {
  background: var(--bg-surface) !important;
  border: 1px solid var(--border-subtle) !important;
  border-radius: 6px !important;
}

/* ── Footer CID INFOSEC ── */
.cid-footer {
  text-align: center;
  padding: 2rem 0 1rem 0;
  margin-top: 3rem;
  border-top: 1px solid var(--border-subtle);
  color: var(--text-muted);
  font-size: 0.78rem;
  letter-spacing: 0.02em;
}
.cid-footer a {
  color: var(--accent-light);
  text-decoration: none;
}

/* ── Breadcrumb ── */
.breadcrumb {
  display: flex; align-items: center; gap: 6px;
  padding: 0.3rem 0 0.8rem 0;
  font-size: 0.78rem;
}
.breadcrumb-item {
  color: var(--text-muted);
  font-weight: 400;
}
.breadcrumb-item.active {
  color: var(--text-secondary);
  font-weight: 500;
}
.breadcrumb-sep {
  color: rgba(255,255,255,0.12);
  font-weight: 300;
}

/* ── Botones diferenciados ── */
/* Botón primario (default — ya definido arriba) */

/* Botón secundario */
div.stButton.secondary > button,
button[id*="secondary"] {
  background: rgba(255,255,255,0.04) !important;
  color: var(--text-secondary) !important;
  border: 1px solid var(--border-standard) !important;
}
div.stButton.secondary > button:hover,
button[id*="secondary"]:hover {
  background: rgba(255,255,255,0.08) !important;
  border-color: rgba(255,255,255,0.12) !important;
  box-shadow: none !important;
}

/* Botón peligro */
div.stButton.danger > button,
button[id*="danger"] {
  background: rgba(248,113,113,0.12) !important;
  color: var(--danger) !important;
  border: 1px solid rgba(248,113,113,0.25) !important;
}
div.stButton.danger > button:hover,
button[id*="danger"]:hover {
  background: rgba(248,113,113,0.2) !important;
  border-color: rgba(248,113,113,0.4) !important;
  box-shadow: 0 2px 10px rgba(248,113,113,0.2) !important;
}

/* ── KPI Cards mejoradas ── */
.kpi-card {
  background: var(--bg-surface);
  border: 1px solid var(--border-subtle);
  border-radius: 8px;
  padding: 18px 16px;
  text-align: left;
  transition: all 0.2s ease;
  position: relative;
  overflow: hidden;
}
.kpi-card:hover {
  border-color: rgba(94,106,210,0.3);
  transform: translateY(-1px);
  box-shadow: 0 4px 12px rgba(0,0,0,0.3);
}
.kpi-card::before {
  content: "";
  position: absolute;
  top: 0; left: 0;
  width: 3px; height: 100%;
  background: var(--accent);
  border-radius: 3px 0 0 3px;
  opacity: 0.5;
}
.kpi-icon {
  font-size: 1.3rem;
  margin-bottom: 6px;
}
.kpi-value {
  font-size: 1.8rem;
  font-weight: 600;
  color: var(--text-primary);
  letter-spacing: -0.5px;
  line-height: 1.1;
}
.kpi-label {
  font-size: 0.74rem;
  color: var(--text-muted);
  margin-top: 2px;
  letter-spacing: 0.02em;
}

/* ── Responsive ── */
@media (max-width: 1024px) {
  [data-testid="stSidebar"] {
    min-width: 260px !important;
    max-width: 260px !important;
  }
  section[data-testid="stSidebar"] + section {
    margin-left: 260px !important;
  }
}
@media (max-width: 768px) {
  [data-testid="stSidebar"] {
    min-width: 100% !important;
    max-width: 100% !important;
  }
  .cid-title {
    font-size: 1.1rem !important;
  }
  .cid-subtitle {
    font-size: 0.72rem !important;
  }
  .kpi-card {
    padding: 12px !important;
  }
  .kpi-value {
    font-size: 1.3rem !important;
  }
  .card {
    padding: 14px !important;
  }
}
@media (max-width: 480px) {
  .health-bar {
    flex-direction: column;
    align-items: flex-start;
  }
  .breadcrumb {
    flex-wrap: wrap;
  }
}
</style>
""", unsafe_allow_html=True)

# ── Autenticación ────────────────────────────────────────────────────────────
with open("config.yaml") as f:
    config = yaml.load(f, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config["credentials"],
    config["cookie"]["name"],
    config["cookie"]["key"],
    config["cookie"]["expiry_days"],
)

authenticator.login()

if st.session_state.get("authentication_status") is False:
    log_event("LOGIN_FAILED", user=st.session_state.get("username", "unknown"), result="failure", detail="Credenciales incorrectas")
    st.error("⛔ Credenciales incorrectas. Acceso denegado.")
    st.stop()

if st.session_state.get("authentication_status") is None:
    st.markdown("""
    <div class='cid-login-container'>
        <div class='cid-login-logo'>🛡️</div>
        <h1 class='cid-login-title'>CID INFOSEC</h1>
        <p class='cid-login-subtitle'>Plataforma de Auditoría de Ciberseguridad e Inteligencia Artificial</p>
        <div class='cid-login-frameworks'>
            ISO 27001 · ISO 42001 · NIST AI RMF · ISO 19011 · ISO 23894
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Usuario autenticado ──────────────────────────────────────────────────────
nombre = st.session_state.get("name", "Auditor")
usuario = st.session_state.get("username", "unknown")

# Registrar inicio de sesión exitoso (solo una vez por sesión)
if not st.session_state.get("_login_logged"):
    log_event("LOGIN_SUCCESS", user=usuario, result="success", detail=f"Sesión iniciada por {nombre}")
    st.session_state["_login_logged"] = True

# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div style='display:flex;align-items:center;gap:10px;margin-bottom:0.4rem;'>
        <div style='width:28px;height:28px;background:linear-gradient(135deg,#5e6ad2,#7170ff);border-radius:5px;display:flex;align-items:center;justify-content:center;font-size:0.9rem;flex-shrink:0;'>🛡️</div>
        <span style='font-size:0.95rem;font-weight:600;color:#f7f8f8;letter-spacing:-0.3px;'>CID INFOSEC</span>
    </div>
    <p style='font-size:0.76rem;color:#8a8f98;margin:0 0 0.8rem 0;'>Auditoría de Ciberseguridad e IA</p>
    """, unsafe_allow_html=True)

    st.markdown(f"**{nombre}**")
    st.markdown(f"<span style='font-size:0.76rem;color:#62666d;'>{usuario}</span>", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<p style='font-size:0.74rem;color:#62666d;letter-spacing:0.05em;text-transform:uppercase;font-weight:500;margin-bottom:0.5rem;'>Navegación</p>", unsafe_allow_html=True)
    opcion_menu = st.radio(
        "Navegación",
        [
            "Chat Auditor",
            "Auditoría de IA",
            "Texto Libre",
            "Ruta ISO 27002",
            "Base RAG",
            "Confluencia",
            "Interoperabilidad",
            "Matriz Riesgos IA",
            "🔴 Threat Intelligence",
            "Autodiagnóstico",
            "Dashboard",
            "Historial",
            "Guía",
            "Rol Auditor",
            "Logs",
        ],
        index=0,
        key="nav_modulo",
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.markdown("<p style='font-size:0.74rem;color:#62666d;letter-spacing:0.05em;text-transform:uppercase;font-weight:500;margin-bottom:0.5rem;'>Entregables</p>", unsafe_allow_html=True)
    try:
        ruta_confluencia = Path("Matriz_Confluencia_Auditoria_IA.xlsx")
        if ruta_confluencia.exists():
            with open(ruta_confluencia, "rb") as f:
                st.download_button(
                    "📊 Base: Matriz de Confluencia",
                    data=f,
                    file_name="Matriz_Confluencia_Auditoria_IA.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    except Exception as e:
        st.caption(f"⚠️ No disponible: {e}")

    # ── Botón de descarga del Excel de respaldo ──
    try:
        excel_respaldo = Path("data") / "Matriz_Riesgos_IA.xlsx"
        if excel_respaldo.exists():
            with open(excel_respaldo, "rb") as f:
                st.download_button(
                    "🛡️ Matriz de Riesgos IA (Excel)",
                    data=f,
                    file_name="Matriz_Riesgos_IA.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="sidebar_excel_riesgos",
                )
    except Exception as e:
        st.caption(f"⚠️ No disponible: {e}")
    st.markdown("---")
    with st.expander("⚙️ Config. Asistente IA", expanded=False):
        proveedor = st.selectbox(
            "Proveedor de IA",
            ["DeepSeek", "Groq (Gratis)", "OpenAI", "Google (Gemini)"],
            index=0,
        )

        if proveedor == "Google (Gemini)":
            modelo = st.selectbox("Modelo", ["gemini-1.5-pro", "gemini-1.5-flash"], index=0)
        elif proveedor == "DeepSeek":
            modelo = st.selectbox("Modelo", ["deepseek-chat", "deepseek-coder"], index=0)
        elif proveedor == "Groq (Gratis)":
            modelo = st.selectbox("Modelo", ["llama-3.3-70b-versatile", "llama-3.1-8b-instant", "mixtral-8x7b-32768"], index=0)
        else:
            modelo = st.selectbox("Modelo", ["gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"], index=0)

        temperatura = st.slider("Temperatura", 0.0, 1.0, 0.1, 0.05)

        usar_rag = st.checkbox("📚 Usar Base de Conocimiento RAG", value=True,
                              help="Desactiva si experimentas problemas con la base de conocimiento")

        try:
            rag_activo = base_conocimiento_activa()
            st.caption("🟢 Base RAG operativa" if rag_activo else "🟡 Base RAG inactiva")
        except Exception:
            st.caption("🔴 Error en Base RAG")

    st.markdown("---")
    if st.session_state.get("authentication_status"):
        authenticator.logout("🚪 Cerrar sesión", location="sidebar")

# ── Encabezado principal ─────────────────────────────────────────────────────
st.markdown("""
<div class='cid-header'>
    <div class='cid-logo'>🛡️</div>
    <div>
        <h1 class='cid-title'>CID INFOSEC</h1>
        <p class='cid-subtitle'>Plataforma de Auditoría · ISO 27001 · ISO 42001 · NIST AI RMF · ISO 19011 · ISO 23894</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Breadcrumb dinámico ─────────────────────────────────────────────────────
st.markdown(f"""
<div class='breadcrumb'>
  <span class='breadcrumb-item'>CID INFOSEC</span>
  <span class='breadcrumb-sep'>/</span>
  <span class='breadcrumb-item active'>{opcion_menu}</span>
</div>
""", unsafe_allow_html=True)

# ── Health Bar del sistema ────────────────────────────────────────────────────
_sv = stats_vulnerabilidades()
_total_vulns = int(_sv.get("total") or 0)
_sg = stats_globales()
_total_analisis = int(_sg.get("total_reportes") or 0)
try:
    _rag_ok = base_conocimiento_activa()
except Exception:
    _rag_ok = False
_rag_dot = "dot-green" if _rag_ok else "dot-yellow"
_rag_label = "RAG Activo" if _rag_ok else "RAG Inactivo"
_vuln_dot = "dot-green" if _total_vulns > 0 else "dot-yellow"
st.markdown(f"""
<div class='health-bar'>
  <span style='font-size:0.78rem;color:rgba(255,255,255,0.4);font-weight:600;letter-spacing:0.05em'>SISTEMA</span>
  <span class='health-item'><span class='health-dot {_rag_dot}'></span>{_rag_label}</span>
  <span class='health-item'><span class='health-dot {_vuln_dot}'></span>{_total_vulns} vulnerabilidades IA</span>
  <span class='health-item'><span class='health-dot dot-green'></span>{_total_analisis} análisis realizados</span>
  <span class='health-item'><span class='health-dot dot-green'></span>Modelo: {proveedor}</span>
</div>
""", unsafe_allow_html=True)

# ── Onboarding (primer uso) ───────────────────────────────────────────────────
if not st.session_state.get("_onboarding_done") and _total_analisis == 0:
    st.markdown("""
    <div class='onboarding-card'>
      <h3>👋 Bienvenido al IA-SEC Auditor Tool</h3>
      <p style='color:rgba(255,255,255,0.65);font-size:0.88rem;margin-bottom:12px'>Sigue estos pasos para realizar tu primer análisis:</p>
      <div class='onboarding-step'><span class='step-num'>1</span> Configura el proveedor de IA en el panel lateral</div>
      <div class='onboarding-step'><span class='step-num'>2</span> Ve a <b>📄 Auditoría de IA</b> y sube un PDF de política de seguridad</div>
      <div class='onboarding-step'><span class='step-num'>3</span> Haz clic en <b>🔍 Ejecutar Gap Analysis</b> y espera el resultado</div>
      <div class='onboarding-step'><span class='step-num'>4</span> Descarga el reporte en PDF o Markdown</div>
    </div>
    """, unsafe_allow_html=True)
    if st.button("✅ Entendido, no mostrar de nuevo", key="btn_onboarding_ok"):
        st.session_state["_onboarding_done"] = True
        st.rerun()

# ── Función: extraer texto del PDF ───────────────────────────────────────────
def extraer_texto_pdf(archivo) -> str:
    lector = pypdf.PdfReader(io.BytesIO(archivo.read()))
    texto = ""
    for pagina in lector.pages:
        texto += pagina.extract_text() or ""
    return texto.strip()

# ── Función: llamar al LLM ───────────────────────────────────────────────────
def ejecutar_gap_analysis(texto_politica: str, proveedor: str, modelo_llm: str, temp: float, usar_rag: bool = True) -> str:
    if proveedor == "Google (Gemini)":
        api_key = os.getenv("GOOGLE_API_KEY")
        if not api_key:
            return "ERROR: No se encontró GOOGLE_API_KEY en el archivo .env"
        llm = ChatGoogleGenerativeAI(model=modelo_llm, temperature=temp, google_api_key=api_key)
    
    elif proveedor == "DeepSeek":
        api_key = os.getenv("DEEPSEEK_API_KEY")
        if not api_key:
            return "ERROR: No se encontró DEEPSEEK_API_KEY en el archivo .env"
        llm = ChatOpenAI(
            model=modelo_llm,
            temperature=temp,
            api_key=api_key,
            base_url="https://api.deepseek.com/v1",
        )
    
    elif proveedor == "Groq (Gratis)":
        api_key = os.getenv("GROQ_API_KEY")
        if not api_key:
            return "ERROR: No se encontró GROQ_API_KEY en el archivo .env"
        llm = ChatOpenAI(
            model=modelo_llm,
            temperature=temp,
            api_key=api_key,
            base_url="https://api.groq.com/openai/v1",
        )
    
    else: # OpenAI / Google
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return "ERROR: No se encontró OPENAI_API_KEY en el archivo .env"
        llm = ChatOpenAI(model=modelo_llm, temperature=temp, api_key=api_key)

    system_prompt = """Eres un Auditor Líder experto en Ciberseguridad e Inteligencia Artificial, con dominio profundo de ISO/IEC 27001:2022, NIST AI RMF 1.0 e ISO/IEC 42001:2023.
Tu misión es ejecutar una auditoría técnica y un análisis de confluencia normativa sobre sistemas de IA desplegados en INFRAESTRUCTURA CRÍTICA.

Evalúa el documento proporcionado considerando la convergencia de estos tres marcos:
1. ISO 27001: Seguridad de la información base (Controles de acceso, logs, vulnerabilidades).
2. NIST AI RMF: Gestión de riesgos específicos de IA (Gobernanza, Robustez Adversarial, Explicabilidad).
3. ISO 42001: Sistema de Gestión de IA (Responsabilidad, transparencia, supervisión humana).

Criterios de evaluación:
1. EXISTENCIA: ¿El control o requisito está presente?
2. RIGOR EN INFRAESTRUCTURA CRÍTICA: ¿El control es suficiente para sectores de alto riesgo (Finanzas, Salud, Energía)?
3. CONFLUENCIA: ¿El documento articula los requisitos de ciberseguridad con los de IA confiable?

Estados:
- ✅ CONFORME: Cumple los requisitos de los tres marcos con rigor.
- ⚠️ PARCIAL: Falta alineación con alguno de los marcos o falta rigor para infraestructura crítica.
- ❌ NO CONFORME: El control no existe o es crítico en sistemas de alto riesgo.

INSTRUCCIONES DE FORMATO:
- Responde ÚNICAMENTE con una tabla Markdown con exactamente 4 columnas:
  | Control / Requisito (Marco Confluente) | Estado | Observación Técnica (Infraestructura Crítica) | Acción de Mejora |
- Al final, agrega un RESUMEN EJECUTIVO de confluencia."""

    # ── RAG: Extraer contexto de la norma oficial ──
    contexto_oficial = ""
    rag_funciona = False

    if usar_rag:
        try:
            # Verificación rápida de que RAG está disponible
            from rag_knowledge import base_conocimiento_activa
            if base_conocimiento_activa():
                contexto_oficial = obtener_contexto(texto_politica[:1500], k=4, incluir_uploads=True)
                rag_funciona = bool(contexto_oficial)
            else:
                contexto_oficial = ""
                rag_funciona = False
        except Exception:
            # Cualquier error - desactivar RAG silenciosamente
            contexto_oficial = ""
            rag_funciona = False

    if contexto_oficial:
        contexto_oficial_texto = f"\n\n--- INICIO CONTEXTO NORMA OFICIAL (RAG) ---\n{contexto_oficial}\n--- FIN CONTEXTO ---\nUtiliza este contexto para evaluar la política."
    else:
        if usar_rag and not rag_funciona:
            contexto_oficial_texto = "\n\nNota: Base de conocimiento RAG no disponible. Análisis realizado con conocimiento general del LLM."
        else:
            contexto_oficial_texto = "\n\nNota: Análisis realizado sin base de conocimiento RAG (deshabilitado por configuración)."

    user_prompt = f"""Ejecuta el Gap Analysis sobre el siguiente documento de política de seguridad:

---INICIO DEL DOCUMENTO DE POLÍTICA---
{texto_politica}
---FIN DEL DOCUMENTO DE POLÍTICA---{contexto_oficial_texto}

Genera la tabla de Gap Analysis ISO/IEC 27001:2022 completa."""

    mensajes = [
        SystemMessage(content=system_prompt),
        HumanMessage(content=user_prompt),
    ]
    respuesta = llm.invoke(mensajes)
    return respuesta.content

# ── Función: Chat Interactivo ────────────────────────────────────────────────
def _construir_llm(proveedor: str, modelo_llm: str, temp: float):
    """Construye el objeto LLM según proveedor. Lanza ValueError si falta API key."""
    if proveedor == "Google (Gemini)":
        api_key = os.getenv("GOOGLE_API_KEY")
        if not api_key:
            raise ValueError("No se encontró GOOGLE_API_KEY en .env")
        return ChatGoogleGenerativeAI(model=modelo_llm, temperature=temp, google_api_key=api_key)
    elif proveedor == "DeepSeek":
        api_key = os.getenv("DEEPSEEK_API_KEY")
        if not api_key:
            raise ValueError("No se encontró DEEPSEEK_API_KEY en .env")
        return ChatOpenAI(model=modelo_llm, temperature=temp, api_key=api_key, base_url="https://api.deepseek.com/v1")
    elif proveedor == "Groq (Gratis)":
        api_key = os.getenv("GROQ_API_KEY")
        if not api_key:
            raise ValueError("No se encontró GROQ_API_KEY en .env")
        return ChatOpenAI(model=modelo_llm, temperature=temp, api_key=api_key, base_url="https://api.groq.com/openai/v1")
    else:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("No se encontró OPENAI_API_KEY en .env")
        return ChatOpenAI(model=modelo_llm, temperature=temp, api_key=api_key)


def ejecutar_chat(mensajes_historial: list, proveedor: str, modelo_llm: str, temp: float):
    try:
        llm = _construir_llm(proveedor, modelo_llm, temp)
    except ValueError as e:
        return f"⚠️ Error de configuración: {e}"

    system_prompt = SystemMessage(content="Eres un Auditor experto en ISO 27001, NIST y ISO 42001. Responde de forma profesional y técnica.")
    formatted_messages = [system_prompt]
    for m in mensajes_historial:
        if m["role"] == "user":
            formatted_messages.append(HumanMessage(content=m["content"]))
        else:
            formatted_messages.append(AIMessage(content=m["content"]))
    try:
        respuesta = llm.invoke(formatted_messages)
        return respuesta.content
    except Exception as e:
        return f"⚠️ El asistente no pudo responder: {e}"

# ── Tabs principales ─────────────────────────────────────────────────────────
# Eliminamos st.tabs y usamos opcion_menu del sidebar

# ── Tab 0: Chat Auditor ──────────────────────────────────────────────────────
if opcion_menu == "Chat Auditor":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 💬 Chat Interactivo con el Auditor")
    st.info(f"Modelo actual: **{proveedor} / {modelo}**")

    if "messages" not in st.session_state:
        st.session_state.messages = []

    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    if prompt := st.chat_input("¿Qué deseas consultar sobre la normativa?"):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        with st.chat_message("assistant"):
            with st.spinner("Pensando..."):
                full_response = ejecutar_chat(st.session_state.messages, proveedor, modelo, temperatura)
                st.markdown(full_response)
        st.session_state.messages.append({"role": "assistant", "content": full_response})
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab 1: Subir PDF ─────────────────────────────────────────────────────────
if opcion_menu == "Auditoría de IA":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 📂 Sube tu política de seguridad (PDF)")
    archivo = st.file_uploader(
        "Arrastra o selecciona un archivo PDF",
        type=["pdf"],
        help="Sube el documento de política de seguridad que deseas auditar",
    )

    if archivo:
        with st.spinner("Extrayendo texto del PDF..."):
            texto = extraer_texto_pdf(archivo)
        palabras = len(texto.split())
        st.success(f"✅ Documento cargado — {palabras:,} palabras extraídas")

        col1, col2 = st.columns([3, 1])
        with col1:
            with st.expander("👁️ Vista previa del texto extraído"):
                st.text_area("", texto[:3000] + ("..." if len(texto) > 3000 else ""), height=200, disabled=True)
        with col2:
            st.markdown("<div class='metric-box'>", unsafe_allow_html=True)
            st.markdown(f"<div class='metric-value'>{palabras:,}</div><div class='metric-label'>Palabras analizadas</div>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("</div>", unsafe_allow_html=True)

        if st.button("🔍 Ejecutar Gap Analysis", key="btn_pdf"):
            log_event("GAP_ANALYSIS_START", user=usuario, detail=f"Archivo: {archivo.name} | Modelo: {proveedor}/{modelo}")
            with st.spinner("🤖 El Auditor IA está analizando tu política... esto puede tardar 30-60 segundos."):
                resultado = ejecutar_gap_analysis(texto, proveedor, modelo, temperatura, usar_rag)
            stats = guardar_reporte(usuario, "PDF", archivo.name, proveedor, modelo, resultado)
            log_event("GAP_ANALYSIS_DONE", user=usuario, result="success", detail=f"Archivo: {archivo.name} | Cumplimiento: {stats['cumplimiento_pct']}%")

            st.markdown("---")
            st.markdown("## 📊 Resultados del Gap Analysis")
            st.markdown(resultado)
            col_md, col_pdf, col_pro = st.columns(3)
            with col_md:
                st.download_button("⬇️ Descargar (.md)", data=resultado,
                    file_name=f"gap_analysis_{archivo.name.replace('.pdf','')}.md", mime="text/markdown")
            with col_pdf:
                pdf_bytes = generar_pdf(resultado, archivo.name, usuario, proveedor, modelo, stats)
                st.download_button("📄 PDF Simple", data=pdf_bytes,
                    file_name=f"gap_analysis_{archivo.name.replace('.pdf','')}.pdf", mime="application/pdf")
            with col_pro:
                # Obtener datos de auditoría para el informe profesional
                aud_id = st.session_state.get("auditoria_actual_id")
                aud_data = obtener_auditoria(aud_id) if aud_id else None
                ctrls = st.session_state.get("controles_actuales", []) if st.session_state.get("controles_aid") == aud_id else None
                plan = st.session_state.get("plan_actual", {}) if st.session_state.get("controles_aid") == aud_id else None
                prof_bytes = generar_informe_auditoria_profesional(
                    resultado_md=resultado,
                    nombre_empresa=aud_data.get("nombre_empresa", archivo.name) if aud_data else archivo.name,
                    sector=aud_data.get("sector", "") if aud_data else "",
                    tamano_empresa=aud_data.get("tamano_empresa", "") if aud_data else "",
                    nombre_doc=archivo.name,
                    usuario=usuario,
                    proveedor=proveedor,
                    modelo=modelo,
                    stats=stats,
                    controles_seleccionados=ctrls if ctrls and aud_id else None,
                    plan_implementacion=plan if plan and aud_id else None,
                )
                st.download_button("📄 PDF Informe Profesional", data=prof_bytes,
                    file_name=f"informe_auditoria_{archivo.name.replace('.pdf','')}.pdf", mime="application/pdf")
    else:
        st.markdown("</div>", unsafe_allow_html=True)

# ── Tab 2: Texto libre ───────────────────────────────────────────────────────
if opcion_menu == "Texto Libre":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### ✏️ Pega el texto de tu política directamente")
    texto_manual = st.text_area(
        "Texto de la política",
        placeholder="Pega aquí el contenido de tu política de seguridad de la información...",
        height=300,
    )
    st.markdown("</div>", unsafe_allow_html=True)

    if texto_manual.strip():
        if st.button("🔍 Ejecutar Gap Analysis", key="btn_texto"):
            log_event("GAP_ANALYSIS_START", user=usuario, detail=f"Texto libre | Modelo: {proveedor}/{modelo} | Chars: {len(texto_manual)}")
            with st.spinner("🤖 Analizando... por favor espera."):
                resultado = ejecutar_gap_analysis(texto_manual, proveedor, modelo, temperatura, usar_rag)
            stats = guardar_reporte(usuario, "Texto", "Análisis manual", proveedor, modelo, resultado)
            log_event("GAP_ANALYSIS_DONE", user=usuario, result="success", detail=f"Texto libre | Cumplimiento: {stats['cumplimiento_pct']}%")

            st.markdown("---")
            st.markdown("## 📊 Resultados del Gap Analysis")
            st.markdown(resultado)
            col_md2, col_pdf2, col_pro2 = st.columns(3)
            with col_md2:
                st.download_button("⬇️ (.md)", data=resultado,
                    file_name="gap_analysis_resultado.md", mime="text/markdown")
            with col_pdf2:
                pdf_bytes = generar_pdf(resultado, "Análisis manual", usuario, proveedor, modelo, stats)
                st.download_button("📄 PDF Simple", data=pdf_bytes,
                    file_name="gap_analysis_resultado.pdf", mime="application/pdf")
            with col_pro2:
                from pdf_export import generar_informe_auditoria_profesional
                aud_id = st.session_state.get("auditoria_actual_id")
                aud_data = obtener_auditoria(aud_id) if aud_id else None
                ctrls = st.session_state.get("controles_actuales", []) if st.session_state.get("controles_aid") == aud_id else None
                plan = st.session_state.get("plan_actual", {}) if st.session_state.get("controles_aid") == aud_id else None
                prof_bytes = generar_informe_auditoria_profesional(
                    resultado_md=resultado,
                    nombre_empresa=aud_data.get("nombre_empresa", "Análisis manual") if aud_data else "Análisis manual",
                    sector=aud_data.get("sector", "") if aud_data else "",
                    tamano_empresa=aud_data.get("tamano_empresa", "") if aud_data else "",
                    nombre_doc="Análisis manual",
                    usuario=usuario,
                    proveedor=proveedor,
                    modelo=modelo,
                    stats=stats,
                    controles_seleccionados=ctrls if ctrls and aud_id else None,
                    plan_implementacion=plan if plan and aud_id else None,
                )
                st.download_button("📄 PDF Profesional", data=prof_bytes,
                    file_name="informe_auditoria_manual.pdf", mime="application/pdf")

# ── Tab Ruta de Auditoría ISO 27002 ──────────────────────────────────────────
if opcion_menu == "Ruta ISO 27002":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 📋 Ruta de Auditoría ISO 27002")
    st.markdown("Guía paso a paso desde el Gap Analysis hasta el establecimiento de controles ISO 27002, adaptada a cualquier empresa sin importar su tamaño o sector.")

    tab1, tab2, tab3 = st.tabs(["🏢 Configurar Empresa", "🗺️ Mapa de Etapas", "📋 Controles y Plan"])

    with tab1:
        # Seleccionar o crear auditoría
        auditorias = obtener_auditorias(usuario)
        opciones_aud = ["➕ Nueva auditoría"] + [f"#{a['id']} - {a.get('nombre_empresa', 'Sin nombre')} ({a.get('sector', '?')})" for a in auditorias]
        aud_seleccionada = st.selectbox("Auditoría", opciones_aud, key="aud_select", index=0)

        if aud_seleccionada == "➕ Nueva auditoría":
            if st.button("🚀 Iniciar nueva auditoría"):
                nombre_empresa = st.session_state.get("aud_nombre", "")
                sector = st.session_state.get("aud_sector", "")
                tamano = st.session_state.get("aud_tamano", "")
                if not nombre_empresa or not sector or not tamano:
                    st.warning("Completa todos los campos antes de iniciar.")
                else:
                    aid = crear_auditoria(usuario, sector, tamano, nombre_empresa)
                    st.session_state["auditoria_actual_id"] = aid
                    st.success(f"Auditoría #{aid} creada. Ve al mapa de etapas para comenzar.")
                    st.rerun()

            with st.form("form_auditoria"):
                st.text_input("Nombre de la empresa", key="aud_nombre")
                st.selectbox("Sector", SECTORES_EMPRESA, key="aud_sector", index=10)
                st.selectbox("Tamaño de la empresa", TAMANOS_EMPRESA, key="aud_tamano", index=0)
                st.form_submit_button("Guardar datos")
        else:
            # Cargar auditoría existente
            idx = int(aud_seleccionada.split(" - ")[0].lstrip("#"))
            st.session_state["auditoria_actual_id"] = idx
            aud = obtener_auditoria(idx)
            if aud:
                st.markdown(f"**Empresa:** {aud.get('nombre_empresa', '—')}")
                st.markdown(f"**Sector:** {aud.get('sector', '—')}")
                st.markdown(f"**Tamaño:** {aud.get('tamano_empresa', '—')}")
                st.markdown(f"**Etapa actual:** {aud.get('etapa_actual', 'gap_analysis')}")
                if st.button("🗑️ Eliminar esta auditoría"):
                    conn = __import__("sqlite3", fromlist=["connect"]).connect("data/reportes.db")
                    conn.execute("DELETE FROM auditoria_progreso WHERE id = ?", (idx,))
                    conn.commit()
                    conn.close()
                    st.success("Auditoría eliminada.")
                    st.session_state.pop("auditoria_actual_id", None)
                    st.rerun()

    with tab2:
        aid = st.session_state.get("auditoria_actual_id")
        if not aid:
            st.info("Primero configura una empresa en la pestaña '🏢 Configurar Empresa'.")
        else:
            aud = obtener_auditoria(aid)
            if not aud:
                st.error("Auditoría no encontrada.")
            else:
                # Mostrar barra de progreso visual
                etapas_ids = [e["id"] for e in ETAPAS]
                etapa_actual = aud.get("etapa_actual", "gap_analysis")
                try:
                    progreso_actual = etapas_ids.index(etapa_actual)
                except ValueError:
                    progreso_actual = 0
                pct = (progreso_actual / (len(etapas_ids) - 1)) * 100

                st.markdown("### 📊 Progreso General")
                st.progress(int(pct), text=f"Etapa {progreso_actual + 1} de {len(etapas_ids)}")

                # Mostrar cada etapa con su estado
                st.markdown("### 🗺️ Etapas de la Auditoría")
                cols = st.columns(len(ETAPAS))
                for i, etapa in enumerate(ETAPAS):
                    icono = etapa_icono_estado(etapa["id"], aud)
                    with cols[i]:
                        st.markdown(f"#### {icono}")
                        st.markdown(f"**{etapa['nombre']}**")
                        st.caption(etapa["descripcion"])
                        if i < progreso_actual:
                            st.info("✅ Completado")
                        elif i == progreso_actual:
                            st.info("⏳ En curso")
                        else:
                            st.info("⬜ Pendiente")

                # Botón para avanzar etapa
                st.markdown("---")
                st.markdown(f"**Etapa actual:** {ETAPAS[progreso_actual]['nombre']}")

                if etapa_actual == "gap_analysis":
                    st.info("Ejecuta un Gap Analysis en las secciones 📄 Auditoría de IA o ✏️ Texto Libre para avanzar a la siguiente etapa.")
                    st.markdown("**Resultado del Gap Analysis:**")
                    reportes = listar_reportes(limit=5)
                    if reportes:
                        for r in reportes:
                            if st.button(f"Cargar reporte #{r['id']} ({r['nombre_doc']})", key=f"gap_load_{r['id']}"):
                                actualizar_auditoria(aid, gap_analysis_id=r["id"])
                                aud["gap_analysis_id"] = r["id"]
                                aud = avanzar_etapa(aud, "gap_analysis")
                                actualizar_auditoria(aid, etapa_actual=aud["etapa_actual"], estado_json=aud["estado_json"])
                                st.success("¡Gap Analysis completado! Avanza a la siguiente etapa.")
                                st.rerun()
                    else:
                        st.warning("No hay reportes de Gap Analysis aún. Usa 📄 Auditoría de IA o ✏️ Texto Libre para crear uno.")
                else:
                    if st.button("✅ Marcar etapa como completada y avanzar"):
                        aud = avanzar_etapa(aud, etapa_actual)
                        actualizar_auditoria(aid, etapa_actual=aud["etapa_actual"], estado_json=aud["estado_json"])
                        st.success("¡Etapa completada! Avanza a la siguiente.")
                        st.rerun()

    with tab3:
        aid = st.session_state.get("auditoria_actual_id")
        if not aid:
            st.info("Primero configura una empresa y avanza a la etapa de controles.")
        else:
            aud = obtener_auditoria(aid)
            if not aud:
                st.error("Auditoría no encontrada.")
            else:
                # Cargar o generar controles recomendados
                if "controles_actuales" not in st.session_state or st.session_state.get("controles_aid") != aid:
                    sector = aud.get("sector", "Otro")
                    tamano = aud.get("tamano_empresa", "PYME (< 50 empleados)")
                    brechas = []
                    gap_id = aud.get("gap_analysis_id")
                    if gap_id:
                        r = obtener_reporte(gap_id)
                        if r and r.get("resultado"):
                            for line in r["resultado"].splitlines():
                                line_s = line.strip()
                                if line_s.startswith("|") and "❌" in line_s:
                                    cells = [c.strip() for c in line_s.split("|") if c.strip()]
                                    if cells:
                                        brechas.append(cells[0])
                    ctrls = obtener_controles_recomendados(sector, tamano, brechas if brechas else None)
                    st.session_state["controles_actuales"] = ctrls
                    st.session_state["controles_aid"] = aid
                    plan = generar_plan_implementacion(ctrls)
                    st.session_state["plan_actual"] = plan
                    # Persistir
                    actualizar_auditoria(aid, controles_json=ctrls, plan_json=plan)

                ctrls = st.session_state.get("controles_actuales", [])
                plan = st.session_state.get("plan_actual", {})

                st.markdown("### 🎯 Controles Recomendados")
                st.markdown(f"**Total:** {len(ctrls)} controles | "
                            f"🔴 Alta: {plan.get('alta_prioridad', 0)} | "
                            f"🟡 Media: {plan.get('media_prioridad', 0)} | "
                            f"🟢 Baja: {plan.get('baja_prioridad', 0)}")

                # Mostrar controles en un data_editor editable
                df_ctrls = __import__("pandas", fromlist=["DataFrame"]).DataFrame(ctrls)
                if not df_ctrls.empty:
                    edited = st.data_editor(
                        df_ctrls,
                        column_config={
                            "id": st.column_config.TextColumn("Control", width="small", disabled=True),
                            "nombre": st.column_config.TextColumn("Nombre", width="medium", disabled=True),
                            "categoria": st.column_config.TextColumn("Categoría", width="small", disabled=True),
                            "prioridad": st.column_config.TextColumn("Prioridad", width="small", disabled=True),
                            "tiempo_estimado": st.column_config.TextColumn("Tiempo", width="small", disabled=True),
                            "descripcion": st.column_config.TextColumn("Descripción", width="large", disabled=True),
                            "responsable_sugerido": st.column_config.TextColumn("Responsable", width="medium"),
                            "estado": st.column_config.SelectboxColumn(
                                "Estado",
                                options=["pendiente", "en_progreso", "completado"],
                                width="small",
                            ),
                        },
                        hide_index=True,
                        use_container_width=True,
                        key="ctrl_editor",
                    )
                    # Guardar ediciones
                    ctrls_editados = edited.to_dict("records")
                    st.session_state["controles_actuales"] = ctrls_editados
                    actualizar_auditoria(aid, controles_json=ctrls_editados)

                # Generar plan de implementación
                st.markdown("---")
                st.markdown("### 📋 Plan de Implementación")
                for fase in plan.get("fases", []):
                    with st.expander(f"📁 {fase['fase']} — {fase['duracion_estimada']}"):
                        for ctrl in fase.get("controles", []):
                            icon_prioridad = {"alta": "🔴", "media": "🟡", "baja": "🟢"}.get(ctrl["prioridad"], "⚪")
                            st.markdown(f"{icon_prioridad} **{ctrl['id']}**: {ctrl['nombre']} — {ctrl['tiempo_estimado']}")
                            st.caption(f"Responsable sugerido: {ctrl.get('responsable_sugerido', '')}")

                # Exportar reporte
                    st.markdown("---")
                if st.button("📄 Generar reporte Markdown"):
                    md = generar_reporte_markdown(aud, ctrls if isinstance(ctrls, list) else [],
                                                  plan, aud.get("nombre_empresa", ""),
                                                  aud.get("sector", ""), aud.get("tamano_empresa", ""))
                    st.download_button("⬇️ (.md)", data=md,
                                       file_name=f"plan_implementacion_{aud.get('nombre_empresa', 'empresa')}.md",
                                       mime="text/markdown")
                    from pdf_export import generar_pdf
                    pdf_bytes = generar_pdf(md, "Plan Implementación", usuario, "", "ISO 27002", {"cumplimiento_pct": pct})
                    st.download_button("📄 PDF", data=pdf_bytes,
                                       file_name=f"plan_implementacion_{aud.get('nombre_empresa', 'empresa')}.pdf",
                                       mime="application/pdf")
                    import io, openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    font_family = "Arial"
                    thin_border = Border(
                        left=Side(style="thin", color="CCCCCC"),
                        right=Side(style="thin", color="CCCCCC"),
                        top=Side(style="thin", color="CCCCCC"),
                        bottom=Side(style="thin", color="CCCCCC"),
                    )
                    purple_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
                    white_font = Font(name=font_family, bold=True, color="FFFFFF", size=11)
                    header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
                    alt_fill = PatternFill(start_color="F3F0FF", end_color="F3F0FF", fill_type="solid")

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Plan Implementación"

                    # ── Portada / Título ──────────────────────────────────────────
                    ws.merge_cells("A1:E1")
                    ws["A1"] = "PLAN DE IMPLEMENTACIÓN ISO 27002"
                    ws["A1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
                    ws["A1"].fill = purple_fill
                    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 40

                    # ── Metadatos ─────────────────────────────────────────────────
                    row = 3
                    ws.cell(row=row, column=1, value="Empresa:").font = Font(name=font_family, bold=True, size=10)
                    ws.cell(row=row, column=2, value=aud.get("nombre_empresa", "")).font = Font(name=font_family, size=10)
                    row += 1
                    ws.cell(row=row, column=1, value="Sector:").font = Font(name=font_family, bold=True, size=10)
                    ws.cell(row=row, column=2, value=aud.get("sector", "")).font = Font(name=font_family, size=10)
                    row += 1
                    ws.cell(row=row, column=1, value="Tamaño:").font = Font(name=font_family, bold=True, size=10)
                    ws.cell(row=row, column=2, value=aud.get("tamano_empresa", "")).font = Font(name=font_family, size=10)

                    # ── Resumen ────────────────────────────────────────────────────
                    row += 2
                    ws.cell(row=row, column=1, value="RESUMEN").font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
                    ws.cell(row=row, column=1).fill = purple_fill
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                    for c in range(1, 6):
                        ws.cell(row=row, column=c).fill = purple_fill
                    row += 1
                    for label, key in [("Total controles", "total_controles"), ("Alta prioridad", "alta_prioridad"),
                                       ("Media prioridad", "media_prioridad"), ("Baja prioridad", "baja_prioridad")]:
                        ws.cell(row=row, column=1, value=label).font = Font(name=font_family, bold=True, size=10)
                        ws.cell(row=row, column=2, value=plan.get(key, 0)).font = Font(name=font_family, size=10)
                        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
                        row += 1

                    # ── Fases ─────────────────────────────────────────────────────
                    row += 1
                    for fase in plan.get("fases", []):
                        ws.cell(row=row, column=1, value=fase.get("fase", "")).font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                        for c in range(1, 6):
                            ws.cell(row=row, column=c).fill = header_fill
                        row += 1
                        headers = ["Control", "Nombre", "Prioridad", "Tiempo", "Responsable"]
                        for ci, h in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=ci, value=h)
                            cell.font = Font(name=font_family, bold=True, color="FFFFFF", size=9)
                            cell.fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        row += 1
                        for idx, item in enumerate(fase.get("controles", [])):
                            for ci, key in enumerate(["id", "nombre", "prioridad", "tiempo_estimado", "responsable_sugerido"], 1):
                                val = item.get(key, "TI") if key != "responsable_sugerido" else item.get("responsable_sugerido", "TI")
                                cell = ws.cell(row=row, column=ci, value=val)
                                cell.font = Font(name=font_family, size=9)
                                cell.border = thin_border
                                if idx % 2 == 1:
                                    cell.fill = alt_fill
                            ws.row_dimensions[row].height = 20
                            row += 1
                        row += 1

                    # ── Ancho de columnas ─────────────────────────────────────────
                    ws.column_dimensions["A"].width = 18
                    ws.column_dimensions["B"].width = 35
                    ws.column_dimensions["C"].width = 14
                    ws.column_dimensions["D"].width = 16
                    ws.column_dimensions["E"].width = 22

                    bus = io.BytesIO()
                    wb.save(bus)
                    bus.seek(0)
                    st.download_button("📊 Excel", data=bus,
                                       file_name=f"plan_implementacion_{aud.get('nombre_empresa', 'empresa')}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # ── Descargar plantilla de informe profesional ─────────────────
                st.markdown("---")
                st.markdown("### 📑 Plantilla de Informe de Auditoría")
                st.caption("Descarga una plantilla reutilizable para informes de auditoría ISO 27001.")
                col_md_t, col_xl_t = st.columns(2)
                with col_md_t:
                    if st.button("📝 Plantilla (.md)", key="btn_plantilla"):
                        from pdf_export import generar_plantilla_informe_markdown
                        plantilla = generar_plantilla_informe_markdown(
                            aud.get("nombre_empresa", ""),
                            aud.get("sector", ""),
                            aud.get("tamano_empresa", ""),
                            usuario, "", "ISO 27002",
                        )
                        st.download_button("⬇️ .md", data=plantilla,
                                           file_name="plantilla_informe_auditoria_iso27001.md", mime="text/markdown",
                                           key="dl_plantilla")
                with col_xl_t:
                    if st.button("📊 Plantilla (.xlsx)", key="btn_plantilla_xl"):
                        import io, openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        font_family = "Arial"
                        thin_border = Border(
                            left=Side(style="thin", color="CCCCCC"),
                            right=Side(style="thin", color="CCCCCC"),
                            top=Side(style="thin", color="CCCCCC"),
                            bottom=Side(style="thin", color="CCCCCC"),
                        )
                        purple_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
                        header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
                        alt_fill = PatternFill(start_color="F3F0FF", end_color="F3F0FF", fill_type="solid")

                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Plantilla Informe"

                        ws.merge_cells("A1:F1")
                        ws["A1"] = "PLANTILLA INFORME DE AUDITORÍA ISO/IEC 27001:2022"
                        ws["A1"].font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
                        ws["A1"].fill = purple_fill
                        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                        ws.row_dimensions[1].height = 35

                        row = 3
                        meta = [
                            ("Organización:", aud.get("nombre_empresa", "")),
                            ("Sector:", aud.get("sector", "")),
                            ("Tamaño:", aud.get("tamano_empresa", "")),
                            ("Auditor:", usuario),
                            ("Fecha:", __import__("datetime").datetime.now().strftime("%d/%m/%Y")),
                        ]
                        for label, val in meta:
                            ws.cell(row=row, column=1, value=label).font = Font(name=font_family, bold=True, size=10)
                            ws.cell(row=row, column=2, value=val).font = Font(name=font_family, size=10)
                            row += 1

                        row += 1
                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                        ws.cell(row=row, column=1, value="CONTROLES RECOMENDADOS").font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
                        for c in range(1, 7):
                            ws.cell(row=row, column=c).fill = header_fill
                        row += 1

                        ctrl_headers = ["Control", "Nombre", "Prioridad", "Tiempo", "Responsable", "Estado"]
                        for ci, h in enumerate(ctrl_headers, 1):
                            cell = ws.cell(row=row, column=ci, value=h)
                            cell.font = Font(name=font_family, bold=True, color="FFFFFF", size=9)
                            cell.fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        row += 1

                        for idx, c in enumerate(ctrls):
                            vals = [c.get(k, "") for k in ["id", "nombre", "prioridad", "tiempo_estimado", "responsable_sugerido", "estado"]]
                            for ci, v in enumerate(vals, 1):
                                cell = ws.cell(row=row, column=ci, value=v)
                                cell.font = Font(name=font_family, size=9)
                                cell.border = thin_border
                                if idx % 2 == 1:
                                    cell.fill = alt_fill
                            row += 1

                        row += 1
                        for fase in plan.get("fases", []):
                            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                            ws.cell(row=row, column=1, value=fase.get("fase", "")).font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
                            for c in range(1, 7):
                                ws.cell(row=row, column=c).fill = header_fill
                            row += 1
                            fase_headers = ["Control", "Nombre", "Prioridad", "Tiempo", "Responsable", ""]
                            for ci, h in enumerate(fase_headers, 1):
                                cell = ws.cell(row=row, column=ci, value=h)
                                cell.font = Font(name=font_family, bold=True, color="FFFFFF", size=9)
                                cell.fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
                                cell.border = thin_border
                            row += 1
                            for item in fase.get("controles", []):
                                vals = [item.get(k, "") for k in ["id", "nombre", "prioridad", "tiempo_estimado", "responsable_sugerido", ""]]
                                for ci, v in enumerate(vals, 1):
                                    ws.cell(row=row, column=ci, value=v).font = Font(name=font_family, size=9)
                                    ws.cell(row=row, column=ci).border = thin_border
                                row += 1
                            row += 1

                        ws.column_dimensions["A"].width = 14
                        ws.column_dimensions["B"].width = 32
                        ws.column_dimensions["C"].width = 14
                        ws.column_dimensions["D"].width = 14
                        ws.column_dimensions["E"].width = 20
                        ws.column_dimensions["F"].width = 14

                        buf = io.BytesIO()
                        wb.save(buf)
                        buf.seek(0)
                        st.download_button("⬇️ .xlsx", data=buf,
                                           file_name=f"plantilla_informe_{aud.get('nombre_empresa', 'empresa')}.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key="dl_plantilla_xl")

    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab RAG (NotebookLM) ──────────────────────────────────────────────────────
if opcion_menu == "Base RAG":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 📚 Base de Conocimiento Oficial (Modo NotebookLM)")
    st.markdown("Sube los PDFs oficiales de la norma ISO/IEC 27001 a la carpeta `knowledge_base/` para que la IA ancle sus respuestas a la realidad.")
    
    estado_rag = "✅ Activa" if base_conocimiento_activa() else "❌ Inactiva / Vacía"
    st.markdown(f"**Estado actual de la base RAG:** {estado_rag}")

    # Información adicional sobre el estado
    try:
        from pathlib import Path
        data_dir = Path("data")
        faiss_dir = data_dir / "faiss_index"
        knowledge_dir = Path("knowledge_base")

        num_docs = len(list(knowledge_dir.glob("*.pdf"))) + len(list(knowledge_dir.glob("*.txt"))) + len(list(knowledge_dir.glob("*.docx")))

        if faiss_dir.exists():
            faiss_size = (faiss_dir / "index.faiss").stat().st_size if (faiss_dir / "index.faiss").exists() else 0
            st.info(f"📊 {num_docs} documentos indexados | Índice: {faiss_size:,} bytes")
        else:
            st.warning("📊 Base de datos FAISS no encontrada")

    except Exception as e:
        st.error(f"Error verificando estado: {e}")
    
    col_idx1, col_idx2 = st.columns(2)
    with col_idx1:
        if st.button("🔄 Indexar documentos de la norma"):
            with st.spinner("Fragmentando y vectorizando documentos con modelo local (esto puede tardar la primera vez)..."):
                num_chunks, msg = indexar_documentos()
                if num_chunks > 0:
                    st.success(msg)
                    log_event("RAG_INDEX_UPDATE", user=usuario, result="success", detail=msg)
                else:
                    st.warning(msg)

    with col_idx2:
        if st.button("🔧 Diagnosticar/Reconstruir índice", help="Usa esto si experimentas errores con RAG"):
            with st.spinner("Diagnosticando y reconstruyendo índice..."):
                try:
                    # Importar aquí para evitar problemas si no está disponible
                    import subprocess
                    result = subprocess.run([sys.executable, "regenerar_indice.py"],
                                          capture_output=True, text=True, cwd=os.getcwd())
                    if result.returncode == 0:
                        st.success("Índice reconstruido exitosamente")
                        st.rerun()  # Recargar para actualizar estado
                    else:
                        st.error(f"Error reconstruyendo índice: {result.stderr}")
                except Exception as e:
                    st.error(f"Error ejecutando diagnóstico: {e}")
                

    # Uploader temporal aislado: no toca knowledge_base/
    st.markdown("---")
    st.markdown("#### 📤 Documentos temporales (carga puntual)")
    st.caption("Estos archivos solo se usan para la sesión actual y no alteran la base oficial de conocimiento.")
    archivos_subidos = st.file_uploader(
        "Arrastra PDFs, TXTs o DOCXs aquí",
        type=["pdf", "txt", "docx"],
        accept_multiple_files=True,
        key="rag_uploader_temporal",
    )
    if archivos_subidos:
        from pathlib import Path as _Path
        UPLOADS_DIR = _Path("uploads")
        UPLOADS_DIR.mkdir(exist_ok=True)
        guardados = []
        for archivo in archivos_subidos:
            ruta_destino = UPLOADS_DIR / archivo.name
            with open(ruta_destino, "wb") as f:
                f.write(archivo.getbuffer())
            guardados.append(archivo.name)
        st.success(f"✅ {len(guardados)} archivos subidos: {', '.join(guardados)}")

    col_up1, col_up2, col_up3 = st.columns(3)
    with col_up1:
        if st.button("🔄 Indexar uploads", key="btn_indexar_uploads"):
            with st.spinner("Indexando documentos temporales..."):
                try:
                    num_chunks, msg = indexar_documentos_temporales()
                    if num_chunks > 0:
                        st.success(msg)
                    else:
                        st.warning(msg)
                except Exception as e:
                    st.error(f"No se pudo indexar uploads en este entorno: {e}")
    with col_up2:
        if st.button("🧹 Limpiar uploads", key="btn_limpiar_uploads"):
            try:
                limpiar_uploads()
                st.success("Uploads temporales eliminados.")
                st.rerun()
            except Exception as e:
                st.error(f"No se pudo limpiar uploads: {e}")
    with col_up3:
        try:
            if base_uploads_activa():
                st.markdown("**Uploads activos:** ✅")
            else:
                st.markdown("**Uploads activos:** —")
        except Exception:
            st.markdown("**Uploads activos:** —")
    st.info("💡 **Cómo funciona:** El sistema busca en el 'Marco de Gobernanza' (NIST/ISO42001) y en los estándares de ciberseguridad (ISO27001) para anclar la evaluación a requisitos reales de infraestructura crítica.")
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Confluencia ──────────────────────────────────────────────────────────
if opcion_menu == "Confluencia":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 🔗 Confluencia Normativa IA-SEC")
    st.markdown("**ISO 27001 + NIST AI RMF + ISO 42001 + ISO 19011 + ISO 23894** — 5 marcos integrados")
    st.markdown("""
    Esta sección muestra cómo se integran los marcos fundamentales para la auditoría de IA en infraestructura crítica.
    
    | Dimensión | ISO/IEC 27001:2022 | NIST AI RMF 1.0 | ISO/IEC 42001:2023 | ISO 19011:2018 | ISO 23894:2023 |
    | :--- | :--- | :--- | :--- | :--- | :--- |
    | **Gobernanza** | Cláusula 5: Liderazgo | Función GOVERN | Cláusula 5: Liderazgo AIMS | 6.2: Iniciación | 4: Principios |
    | **Riesgos** | 6.1.2: Evaluación | MAP/MEASURE | 6.1.2: Riesgos IA | 6.3: Preparación | 6.1: Proceso completo |
    | **Operación** | Anexo A: Controles | MANAGE | Anexo A: Controles IA | 6.4: Ejecución | 6.1.5: Tratamiento |
    | **Monitoreo** | A.8.16: Seguimiento | MEASURE | Cláusula 9: Evaluación | 6.5: Informe | 6.2: Monitoreo |
    | **Auditoría** | 9.2: Auditoría interna | — | 9.2: Auditoría interna | 6: Todo el ciclo | — |
    | **Privacidad** | A.8.10: Datos | Trustworthy AI | Anexo B: Impacto social | — | 6.2: Privacidad |
    | **Técnico** | A.8.8: Vulnerabilidades | Robustness & Safety | A.7: Desarrollo seguro | 6.4: Evidencia | 5.5: Adversarial |
    """)
    
    st.markdown("---")
    st.markdown("#### 📥 Entregable Académico")
    st.info("Descarga la matriz detallada de confluencia que articula los 5 marcos normativos para tu propuesta de auditoría de IA en infraestructura crítica.")
    with open("Matriz_Confluencia_Auditoria_IA.xlsx", "rb") as f:
        st.download_button(
            "⬇️ Descargar Matriz de Confluencia (Excel)",
            data=f,
            file_name="Matriz_Confluencia_Auditoria_IA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_matriz_conf"
        )
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Interoperabilidad ─────────────────────────────────────────────────────
if opcion_menu == "Interoperabilidad":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 🧩 Matriz de Confluencia Normativa IA-SEC")
    st.markdown("**Herramienta de apoyo al auditor** — ISO/IEC 27001:2022 + ISO/IEC 42001:2023 + NIST AI RMF 1.0")
    st.markdown("*23 controles correlacionados · 9 dimensiones operativas · Escala MSPI 6 niveles*")

    # Selección de sector
    sectores_opciones = ["Selecciona un sector..."] + [f"{k}: {v['nombre']}" for k, v in SECTORES_IC.items()]
    sector_seleccionado = st.selectbox(
        "🏭 Sector de Infraestructura Crítica",
        sectores_opciones,
        help="Selecciona el sector específico para contextualizar la evaluación"
    )

    sector_id = None
    if sector_seleccionado != "Selecciona un sector...":
        sector_id = sector_seleccionado.split(":")[0]
        sector_info = obtener_sector(sector_id)
        if sector_info:
            st.info(f"**{sector_info['nombre']}**: {sector_info['descripcion']}")
            st.write(f"**Regulaciones adicionales:** {', '.join(sector_info['regulaciones_adicionales'])}")

    st.markdown("---")

    # Inicializar DataFrame en sesión si no existe
    if 'df_interop' not in st.session_state:
        st.session_state.df_interop = generar_df_interoperabilidad()

    # Mostrar métricas de resumen
    resumen = generar_resumen_interoperabilidad()
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Controles", resumen["total_controles"])
    with col2:
        st.metric("Controles Críticos", resumen["por_relevancia"]["Crítica"])
    
    # Calcular cumplimiento actual desde el DataFrame
    df_eval = st.session_state.df_interop
    df_eval['%'] = df_eval['Estado Actual'].apply(calcular_cumplimiento)
    avance_global = df_eval['%'].mean()
    
    with col3:
        st.metric("Dimensiones", len(resumen["dimensiones"]))
    with col4:
        st.metric("Avance Global", f"{avance_global:.1%}")

    st.markdown("---")

    # Editor de datos interactivo (Modo GRC Excel)
    st.markdown("#### 📊 Matriz de Evaluación GRC (Interactivo)")
    st.info("Utiliza esta tabla para evaluar el estado de implementación. Los cambios se guardan automáticamente en la sesión.")
    
    edited_df = st.data_editor(
        st.session_state.df_interop,
        column_config={
            "ID": st.column_config.TextColumn("ID", width="small", disabled=True),
            "Dimensión": st.column_config.TextColumn("Dimensión", width="medium", disabled=True),
            "ISO 27001": st.column_config.TextColumn("ISO 27001", width="medium", disabled=True),
            "ISO 42001": st.column_config.TextColumn("ISO 42001", width="medium", disabled=True),
            "NIST AI RMF": st.column_config.TextColumn("NIST AI RMF", width="medium", disabled=True),
            "Relevancia IC": st.column_config.TextColumn("Relevancia IC", width="small", disabled=True),
            "Estado Actual": st.column_config.SelectboxColumn(
                "Estado Actual",
                options=[
                    "No iniciado", 
                    "En proceso inicial", 
                    "Implementado parcialmente", 
                    "Mayormente implementado", 
                    "Cumplimiento total",
                    "No evaluado"
                ],
                required=True,
                width="medium"
            ),
            "Plan de Acción": st.column_config.TextColumn(
                "¿Qué voy a hacer para lograrlo?",
                width="large"
            )
        },
        hide_index=True,
        use_container_width=True,
        key="grc_editor"
    )
    
    # Actualizar sesión con los cambios
    st.session_state.df_interop = edited_df

    # Botón para descargar en Excel (Estilo GRC GAP)
    if 'df_interop' in st.session_state:
        excel_bytes = exportar_matriz_excel(st.session_state.df_interop)
        st.download_button(
            label="📥 Descargar Matriz de Cumplimiento GRC (Excel)",
            data=excel_bytes,
            file_name="Matriz_Cumplimiento_Integrada_GRC.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Descarga un reporte Excel profesional con pestañas separadas por norma, semáforos de cumplimiento y la tabla de valoración de madurez."
        )

    st.markdown("---")

    # Navegación por dimensiones para vista detallada
    dimensiones = obtener_dimensiones()
    dimension_seleccionada = st.selectbox(
        "🔍 Explorar Requisitos Detallados por Dimensión",
        ["Vista general"] + dimensiones,
        help="Selecciona una dimensión para ver los controles detallados y requisitos específicos de IC"
    )

    if dimension_seleccionada == "Vista general":
        # Vista general con estadísticas
        import pandas as pd

        # Crear tabla resumen por dimensión
        data_dimensiones = []
        for dim in dimensiones:
            controles = obtener_controles_por_dimension(dim)
            criticos = sum(1 for c in controles if c.relevancia_ic == "Crítica")
            muy_altos = sum(1 for c in controles if c.relevancia_ic == "Muy Alta")
            altos = sum(1 for c in controles if c.relevancia_ic == "Alta")
            data_dimensiones.append({
                "Dimensión": dim,
                "Total Controles": len(controles),
                "Críticos": criticos,
                "Muy Altos": muy_altos,
                "Altos": altos
            })

        df_dimensiones = pd.DataFrame(data_dimensiones)
        st.dataframe(df_dimensiones, use_container_width=True, hide_index=True)

        # Gráfico de distribución por relevancia
        fig_relevancia = go.Figure()
        relevancia_data = resumen["por_relevancia"]
        fig_relevancia.add_trace(go.Pie(
            labels=list(relevancia_data.keys()),
            values=list(relevancia_data.values()),
            marker_colors=["#dc2626", "#d97706", "#059669"],
            title="Distribución por relevancia en IC"
        ))
        fig_relevancia.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="white", showlegend=True,
            height=300,
        )
        st.plotly_chart(fig_relevancia, use_container_width=True)

        # Gráfico de barras de cumplimiento
        # Recalcular cobertura para el gráfico
        evals_dict = {row['ID']: row['Estado Actual'] for _, row in st.session_state.df_interop.iterrows()}
        cobertura = calcular_cobertura(evals_dict)

        st.markdown("#### 📈 Nivel de Cumplimiento por Dimensión")
        cols_cob = st.columns(len(cobertura))
        for i, (dim, stats) in enumerate(cobertura.items()):
            with cols_cob[i % len(cols_cob)]:
                st.metric(dim, f"{stats['porcentaje']}%")
        
        fig_cumplimiento = px.bar(
            x=list(cobertura.keys()),
            y=[s["porcentaje"] for s in cobertura.values()],
            labels={"x": "Dimensión", "y": "% Cumplimiento"},
            title="Porcentaje de Cumplimiento Acumulado",
            color_discrete_sequence=["#a78bfa"]
        )
        fig_cumplimiento.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="white", height=350
        )
        st.plotly_chart(fig_cumplimiento, use_container_width=True)

    else:
        # Vista detallada de una dimensión
        controles = obtener_controles_por_dimension(dimension_seleccionada)
        st.subheader(f"📋 {dimension_seleccionada}")

        # Controles en formato expandible
        for control in controles:
            relevancia_color = {
                "Crítica": "#dc2626",
                "Muy Alta": "#d97706",
                "Alta": "#059669"
            }.get(control.relevancia_ic, "#64748b")

            with st.expander(f"🔴 {control.id}: {control.iso_27001_ref} ({control.relevancia_ic})"):
                col_a, col_b = st.columns(2)

                with col_a:
                    st.markdown("**ISO/IEC 27001:2022**")
                    st.write(f"*{control.iso_27001_ref}*")
                    st.write(control.iso_27001_desc)

                    st.markdown("**ISO/IEC 42001:2023**")
                    st.write(f"*{control.iso_42001_ref}*")
                    st.write(control.iso_42001_desc)

                    st.markdown("**NIST AI RMF 1.0**")
                    st.write(f"*{control.nist_ai_rmf_ref}*")
                    st.write(control.nist_ai_rmf_desc)

                with col_b:
                    st.markdown(f"**Relevancia en IC:** <span style='color:{relevancia_color};font-weight:bold'>{control.relevancia_ic}</span>", unsafe_allow_html=True)
                    st.write("**Requisito específico IC:**")
                    st.write(control.requisito_ic)
                    
                    # Mostrar estado y plan desde la matriz GRC
                    row = st.session_state.df_interop[st.session_state.df_interop['ID'] == control.id]
                    if not row.empty:
                        estado = row['Estado Actual'].values[0]
                        plan = row['Plan de Acción'].values[0]
                        st.markdown(f"**Estado Evaluación:** `{estado}`")
                        if plan:
                            st.success(f"📌 **Plan:** {plan}")

                    st.write("**Riesgo sin control:**")
                    st.write(control.riesgo_sin_control)

    st.markdown("---")

    # Generación de propuestas
    if sector_id:
        st.markdown("### 🎯 Generar Propuesta de Implementación")
        st.markdown("""
        **Metodología:**  
        🔷 **ISO 19011:2018** — Las 5 fases del ciclo de auditoría guían la estructura de la propuesta  
        🔷 **ISO 23894:2023** — Los riesgos de IA (opacidad, deriva, adversarial, sesgo, privacidad, impacto social, seguridad física) se incorporan como catálogo de riesgos  
        🔷 **ISO 27001 + ISO 42001 + NIST AI RMF** — Controles correlacionados para la evaluación
        """)

        # Evaluación de controles
        st.markdown("#### 📝 Datos para la propuesta")
        st.info("La propuesta se basará en los datos ingresados en la Matriz de Evaluación GRC y el contexto organizacional.")

        # Recolectar evaluaciones del DataFrame de la sesión
        controles_evaluados = {row['ID']: row['Estado Actual'] for _, row in st.session_state.df_interop.iterrows()}

        # Contexto organizacional
        contexto_org = st.text_area(
            "📋 Contexto de la organización",
            placeholder="Describe brevemente la organización: tamaño, madurez en ciberseguridad, experiencia previa con IA, regulaciones aplicables, etc.",
            height=100
        )

        # Botón de generación
        if st.button("🚀 Generar Propuesta de Implementación", type="primary"):
            if not contexto_org.strip():
                st.error("Debes proporcionar contexto organizacional.")
            else:
                log_event("INTEROP_PROPOSAL_START", user=usuario, detail=f"Sector: {sector_id} | Controles evaluados: {len(controles_evaluados)}")

                with st.spinner("🤖 Generando propuesta personalizada... esto puede tardar 1-2 minutos."):
                    propuesta = generar_propuesta_ic(
                        sector_id, controles_evaluados, contexto_org,
                        proveedor, modelo, temperatura
                    )

                if propuesta.startswith("ERROR:"):
                    st.error(propuesta)
                else:
                    log_event("INTEROP_PROPOSAL_DONE", user=usuario, result="success", detail=f"Sector: {sector_id} | Propuesta generada")

                    st.success("✅ Propuesta generada exitosamente")

                    # Mostrar propuesta
                    st.markdown("---")
                    st.markdown("## 📄 Propuesta de Implementación")
                    st.markdown(propuesta)

                    # Opciones de descarga
                    col_md, col_pdf = st.columns(2)
                    with col_md:
                        st.download_button(
                            "⬇️ Descargar (.md)",
                            data=propuesta,
                            file_name=f"propuesta_interoperabilidad_{sector_id}.md",
                            mime="text/markdown",
                            key="download_md"
                        )
                    with col_pdf:
                        # Usar función especializada para propuestas de interoperabilidad
                        sector_info = obtener_sector(sector_id)
                        sector_nombre = sector_info["nombre"] if sector_info else sector_id
                        pdf_bytes = generar_pdf_propuesta_interoperabilidad(
                            propuesta, sector_id, sector_nombre, controles_evaluados,
                            usuario, proveedor, modelo
                        )
                        st.download_button(
                            "📄 Descargar PDF (auditoría ISO 19011)",
                            data=pdf_bytes,
                            file_name=f"informe_auditoria_{sector_id}.pdf",
                            mime="application/pdf",
                            key="download_pdf"
                        )

    # Descarga de la matriz completa
    st.markdown("---")
    st.markdown("### 📥 Descargas")
    col_a, col_b = st.columns(2)
    with col_a:
        matriz_md = generar_markdown_matriz(sector_id)
        st.download_button(
            "⬇️ Matriz Interoperabilidad (.md)",
            data=matriz_md,
            file_name=f"matriz_interoperabilidad{'_' + sector_id if sector_id else ''}.md",
            mime="text/markdown",
            key="matriz_md"
        )
    with col_b:
        st.info("💡 **Uso recomendado:** Comparte esta matriz con equipos técnicos para evaluar brechas antes de generar propuestas específicas.")

    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Guía (Enriquecida con ISO 19011 e ISO 23894) ────────────────────────
if opcion_menu == "Guía":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    from guia_auditor import mostrar_guia_mejorada
    mostrar_guia_mejorada()
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Rol del Auditor (ISO 19011 + ISO 23894 + Base de Conocimiento) ─────
if opcion_menu == "Rol Auditor":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    from guia_auditor import mostrar_rol_auditor
    mostrar_rol_auditor()
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Matriz de Riesgos IA ──────────────────────────────────────────────
if opcion_menu == "Matriz Riesgos IA":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 🛡️ Matriz de Riesgos de Inteligencia Artificial")
    st.markdown("""
    **Feed automatizado de vulnerabilidades IA + Matriz de Riesgos personalizada**
    Consulta en tiempo real: AVID (avidml.org) · MITRE ATLAS (atlas.mitre.org) · MIT AI Risk Repository (airisk.mit.edu)
    """)

    # ── Actualización automática al cargar ───────────────────────────────
    if "feed_inicializado" not in st.session_state:
        st.session_state["feed_inicializado"] = False

    if not st.session_state["feed_inicializado"]:
        with st.spinner("🔄 Verificando base de conocimiento de vulnerabilidades IA..."):
            try:
                from database import stats_vulnerabilidades

                stats = stats_vulnerabilidades()
                ahora = ahora_bogota()
                ultima_act = stats.get("ultima_actualizacion")
                total = stats.get("total", 0)

                necesita_actualizar = total == 0 or total is None
                if not necesita_actualizar and ultima_act:
                    try:
                        ultima_dt = datetime.fromisoformat(ultima_act)
                        if (ahora_bogota() - ultima_dt).total_seconds() > 86400:
                            necesita_actualizar = True
                    except (ValueError, TypeError):
                        necesita_actualizar = True

                if necesita_actualizar:
                    from ingestor.sources import fetch_all as _fetch_all_src
                    raw = _fetch_all_src(limit_per_source=9999, include_nvd=False)
                    if raw:
                        normalized = normalize_batch(raw)
                        for v in normalized:
                            p = v.get('probabilidad_base', 3)
                            s = v.get('severidad_tecnica', 3)
                            riesgo = calcular_riesgo(p, s)
                            v['riesgo_valor'] = riesgo['riesgo_valor']
                            v['nivel_riesgo'] = riesgo['nivel_riesgo']
                            v['criticidad'] = criticidad_compuesta(p, s)
                        insertados, errores = guardar_lote_vulnerabilidades(normalized)
                        limpiar_vulnerabilidades_antiguas(dias=90)
                        exportar_excel_respaldo()
                        st.success(f"✅ Base actualizada: {insertados} nuevas vulnerabilidades")
                else:
                    st.info(f"📊 Base de datos actual: {int(total or 0)} vulnerabilidades")
            except Exception as e:
                st.warning(f"⚠️ No se pudo actualizar automáticamente: {e}")
        st.session_state["feed_inicializado"] = True

    # ── Panel de control de actualización ────────────────────────────────
    with st.expander("⚙️ **Panel de control del feed**", expanded=False):
        col_s1, col_s2, col_s3, col_s4 = st.columns(4)

        with col_s1:
            if st.button("🔄 Actualizar ahora", key="btn_feed_now", type="primary"):
                st.session_state["forzar_actualizacion"] = True

        with col_s2:
            intervalo = st.selectbox(
                "Actualización automática",
                ["Manual", "Cada 1 hora", "Cada 6 horas", "Cada 12 horas", "Cada 24 horas"],
                index=3,
                key="intervalo_feed",
            )

        with col_s3:
            incluir_nvd = st.checkbox("Incluir NVD (CVEs de IA)", value=False, key="feed_nvd")

        with col_s4:
            proveedor_normalizador = st.selectbox(
                "Proveedor IA",
                ["deepseek", "google"],
                index=0,
                key="feed_proveedor",
            )
            modelo_normalizador = st.selectbox(
                "Modelo",
                MODEL_MAP.get(proveedor_normalizador, ["deepseek-chat"]),
                key="feed_modelo",
            )

        # Última actualización
        try:
            from database import stats_vulnerabilidades
            sv = stats_vulnerabilidades()
            ult = sv.get("ultima_actualizacion", "Nunca")
            if ult and ult != "Nunca":
                st.caption(f"📅 Última actualización: {ult[:19].replace('T', ' ')} UTC")
        except Exception:
            st.caption("📅 Última actualización: Desconocida")

    # ── Forzar actualización ─────────────────────────────────────────────
    if st.session_state.get("forzar_actualizacion", False):
        st.session_state["forzar_actualizacion"] = False
        with st.spinner("🔄 Consultando MITRE ATLAS, AVID, MIT AI Risk Repository..."):
            try:
                from ingestor.sources import fetch_all as _fetch_all_src
                raw = _fetch_all_src(limit_per_source=9999, include_nvd=incluir_nvd)
                if not raw:
                    st.error("No se obtuvieron datos de las fuentes.")
                else:
                    proveedor = st.session_state.get("feed_proveedor", "deepseek")
                    modelo = st.session_state.get("feed_modelo", "deepseek-chat")
                    set_provider(proveedor, modelo)

                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    normalized = []
                    for idx, report in enumerate(raw):
                        status_text.text(f"Normalizando {idx+1}/{len(raw)}: {report.get('id', report.get('title',''))[:60]}...")
                        result = analizar_vulnerabilidad(report.get("source", "DESCONOCIDA"), report)
                        normalized.append(result)
                        progress_bar.progress((idx + 1) / len(raw))
                    progress_bar.empty()
                    status_text.empty()

                    for v in normalized:
                        p = v.get('probabilidad_base', 3)
                        s = v.get('severidad_tecnica', 3)
                        riesgo = calcular_riesgo(p, s)
                        v['riesgo_valor'] = riesgo['riesgo_valor']
                        v['nivel_riesgo'] = riesgo['nivel_riesgo']
                        v['criticidad'] = criticidad_compuesta(p, s)

                    insertados, errores = guardar_lote_vulnerabilidades(normalized)
                    limpiadas = limpiar_vulnerabilidades_antiguas(dias=90)
                    ruta_excel = exportar_excel_respaldo()

                    st.session_state["ultimas_vulns"] = normalized
                    st.session_state["vuln_stats"] = resumen_estadistico(normalized)

                    # ── definir msgs ANTES de usarlo ──
                    msgs = []
                    if insertados > 0:
                        msgs.append(f"✅ {insertados} vulnerabilidades nuevas insertadas")
                    if errores > 0:
                        msgs.append(f"⚠️ {errores} duplicados ignorados")
                    if limpiadas > 0:
                        msgs.append(f"🗑️ {limpiadas} vulnerabilidades antiguas desactivadas")
                    if ruta_excel:
                        msgs.append(f"📁 Excel respaldo generado")
                    st.success(" | ".join(msgs) if msgs else "✅ Base de datos actualizada")
                    st.rerun()
            except Exception as e:
                st.error(f"Error durante la actualización: {e}")
                st.code(traceback.format_exc())

    # ── Paso 1: Cuestionario contextual ────────────────────────────────────
    with st.expander("📋 **Paso 1: Contexto de la organización** (completar una vez)", expanded="contexto_ok" not in st.session_state):
        st.markdown("Responde estas preguntas para personalizar la matriz de riesgos:")

        respuestas = {}
        cols_preg = st.columns(2)
        for i, q in enumerate(CUESTIONARIO_CONTEXTO):
            with cols_preg[i % 2]:
                key = f"ctx_{q['id']}"
                if q["tipo"] == "select":
                    respuestas[q["id"]] = st.selectbox(
                        q["pregunta"],
                        q["opciones"],
                        index=0,
                        key=key,
                    )
                elif q["tipo"] == "multiselect":
                    respuestas[q["id"]] = st.multiselect(
                        q["pregunta"],
                        q["opciones"],
                        key=key,
                    )

        col_ctx1, col_ctx2 = st.columns([1, 4])
        with col_ctx1:
            if st.button("✅ Guardar Contexto", key="btn_guardar_ctx", type="primary"):
                st.session_state["contexto_ok"] = True
                st.session_state["respuestas_contexto"] = respuestas
                st.session_state["prompt_contexto"] = generar_prompt_contexto(respuestas)
                st.rerun()
        with col_ctx2:
            if "contexto_ok" in st.session_state:
                st.success("✅ Contexto guardado. Ve al Paso 2.")

    # ── Paso 2: Consultar y normalizar fuentes ────────────────────────────
    with st.expander("🔍 **Paso 2: Consultar fuentes de vulnerabilidades**", expanded=True):
        col_f1, col_f2, col_f3 = st.columns([1, 1, 1])
        with col_f1:
            limit_per_source = st.number_input("Vulnerabilidades por fuente", min_value=1, max_value=20, value=5, key="vuln_limit")
        with col_f2:
            incluir_nvd = st.checkbox("Incluir NVD (CVEs de IA)", value=False, key="vuln_nvd")
        with col_f3:
            usar_deepseek = st.checkbox("Usar DeepSeek para normalización", value=True, key="vuln_ds")

        if st.button("🚀 Consultar y Normalizar Fuentes", key="btn_fetch_all", type="primary"):
            with st.spinner("Consultando AVID, MITRE ATLAS, MIT AI Risk Repository..."):
                from ingestor.sources import fetch_all as _fetch_all_src
                raw_reports = _fetch_all_src(limit_per_source=limit_per_source, include_nvd=incluir_nvd)
            st.info(f"📥 {len(raw_reports)} reportes obtenidos de las fuentes.")

            progress_bar = st.progress(0)
            status_text = st.empty()
            normalized = []
            proveedor = st.session_state.get("feed_proveedor", "deepseek")
            modelo = st.session_state.get("feed_modelo", "deepseek-chat")
            set_provider(proveedor, modelo)
            for idx, report in enumerate(raw_reports):
                fuente = report.get("source", "DESCONOCIDA")
                status_text.text(f"Normalizando {idx+1}/{len(raw_reports)}: {report.get('id', report.get('title', ''))[:60]}...")
                result = analizar_vulnerabilidad(fuente, report)
                normalized.append(result)
                progress_bar.progress((idx + 1) / len(raw_reports))
            progress_bar.empty()
            status_text.empty()
            st.success(f"✅ {len(normalized)} vulnerabilidades normalizadas.")

            # Calcular riesgo compuesto para cada una
            for v in normalized:
                p = v.get("probabilidad_base", 3)
                s = v.get("severidad_tecnica", 3)
                riesgo = calcular_riesgo(p, s)
                v["riesgo_valor"] = riesgo["riesgo_valor"]
                v["nivel_riesgo"] = riesgo["nivel_riesgo"]
                v["criticidad"] = criticidad_compuesta(p, s)
                if "contexto_org" not in v and "prompt_contexto" in st.session_state:
                    v["contexto_org"] = st.session_state["prompt_contexto"]

            # Guardar en base de datos
            insertados, errores = guardar_lote_vulnerabilidades(normalized)
            st.session_state["ultimas_vulns"] = normalized
            st.session_state["vuln_stats"] = resumen_estadistico(normalized)

            if errores > 0:
                st.warning(f"⚠️ {insertados} insertadas, {errores} con errores (posibles duplicados)")
            else:
                st.success(f"✅ {insertados} vulnerabilidades guardadas en base de datos.")
            st.rerun()

    # ── Paso 3: Visualizar resultados ─────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📊 Vulnerabilidades Activas")

    vulns = st.session_state.get("ultimas_vulns", None)
    if vulns is None:
        vulns = listar_vulnerabilidades(limit=9999)

    if not vulns:
        st.info("Aún no hay vulnerabilidades consultadas. Ejecuta el Paso 2 para poblar la matriz.")
    else:
        stats_v = st.session_state.get("vuln_stats", resumen_estadistico(vulns))

        # Métricas rápidas
        mcols = st.columns(6)
        with mcols[0]:
            st.metric("Total", stats_v["total"])
        with mcols[1]:
            st.metric("🔴 Extremos", stats_v["por_nivel"].get("EXTREMO", 0))
        with mcols[2]:
            st.metric("🟠 Altos", stats_v["por_nivel"].get("ALTO", 0))
        with mcols[3]:
            st.metric("🟡 Medios", stats_v["por_nivel"].get("MEDIO", 0))
        with mcols[4]:
            st.metric("🟢 Bajos", stats_v["por_nivel"].get("BAJO", 0))
        with mcols[5]:
            st.metric("Criticidad Prom.", f"{stats_v['criticidad_promedio']}%")

        # Tabla de vulnerabilidades
        df_vulns = []
        for v in vulns:
            df_vulns.append({
                "ID": v.get("id_vulnerabilidad", ""),
                "Fuente": v.get("fuente_origen", ""),
                "Resumen (ES)": v.get("resumen_es", "")[:120],
                "Táctica MITRE": v.get("tactica_mitre", ""),
                "Tipo": v.get("clasificacion_tipo", ""),
                "Prob": v.get("probabilidad_base", 3),
                "Sev": v.get("severidad_tecnica", 3),
                "Riesgo": v.get("riesgo_valor", 9),
                "Nivel": v.get("nivel_riesgo", "MEDIO"),
            })

        df = pd.DataFrame(df_vulns)

        # Color por nivel de riesgo
        def color_nivel(val):
            colors = {"EXTREMO": "background-color: #FF0000; color: white",
                      "ALTO": "background-color: #FF6600; color: black",
                      "MEDIO": "background-color: #FFD700; color: black",
                      "BAJO": "background-color: #92D050; color: black"}
            return colors.get(val, "")

        styled_df = df.style.map(color_nivel, subset=["Nivel"])
        st.dataframe(styled_df, use_container_width=True, hide_index=True,
                     column_config={
                         "Riesgo": st.column_config.NumberColumn(format="%d"),
                         "Prob": st.column_config.NumberColumn(format="%d"),
                         "Sev": st.column_config.NumberColumn(format="%d"),
                     })

        # ── Expandir detalle de cada vulnerabilidad ──
        with st.expander("🔬 Ver detalle completo de vulnerabilidades"):
            for v in vulns:
                nivel = v.get("nivel_riesgo", "MEDIO")
                emoji = {"EXTREMO": "🔴", "ALTO": "🟠", "MEDIO": "🟡", "BAJO": "🟢"}
                with st.container():
                    st.markdown(f"""**{emoji.get(nivel, '⚪')} {v.get('id_vulnerabilidad', 'N/A')}** — *{v.get('fuente_origen', '?')}*
- **Resumen (ES):** {v.get('resumen_es', 'N/A')}
- **Causa Raíz:** {v.get('causa_raiz', 'N/A')}
- **Remediación:** {v.get('accion_remediacion', 'N/A')}
- **Táctica MITRE:** {v.get('tactica_mitre', 'N/A')} → {clasificar_tactica(v.get('tactica_mitre', ''))}
- **Probabilidad:** {v.get('probabilidad_base', '?')}/5 | **Severidad:** {v.get('severidad_tecnica', '?')}/5 | **Riesgo:** {v.get('riesgo_valor', '?')}/25
- **Normas:** {', '.join(v.get('normas_aplicables', [])) if isinstance(v.get('normas_aplicables'), list) else v.get('normas_aplicables', 'N/A')}
""")
                    st.markdown("---")

        # ── Gráficos ──
        col_g1, col_g2 = st.columns(2)

        with col_g1:
            st.markdown("**Distribución por nivel de riesgo**")
            niveles_data = stats_v["por_nivel"]
            fig_niveles = go.Figure(data=[
                go.Pie(
                    labels=list(niveles_data.keys()),
                    values=list(niveles_data.values()),
                    marker_colors=["#FF0000", "#FF6600", "#FFD700", "#92D050"],
                    hole=0.4,
                )
            ])
            fig_niveles.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font_color="white", height=300, showlegend=True,
            )
            st.plotly_chart(fig_niveles, use_container_width=True)

        with col_g2:
            st.markdown("**Distribución por clasificación**")
            clasif_data = stats_v["por_clasificacion"]
            if clasif_data:
                fig_clasif = go.Figure(data=[
                    go.Bar(
                        x=list(clasif_data.keys()),
                        y=list(clasif_data.values()),
                        marker_color="#a78bfa",
                    )
                ])
                fig_clasif.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="white", height=300,
                )
                st.plotly_chart(fig_clasif, use_container_width=True)

        # ── Exportar Excel ───────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📥 Exportar Matriz de Riesgos IA")

        col_e1, col_e2 = st.columns([1, 3])
        with col_e1:
            if st.button("📊 Generar Excel Matriz de Riesgos", key="btn_export_excel", type="primary"):
                with st.spinner("Generando Excel profesional..."):
                    contexto = st.session_state.get("prompt_contexto", "")
                    excel_bytes = exportar_matriz_riesgos_ia(
                        vulns,
                        contexto_organizacion=contexto,
                        respuestas_cuestionario=st.session_state.get("respuestas_contexto"),
                    )
                    st.session_state["excel_riesgos_ia"] = excel_bytes
                st.success("✅ Excel generado. Descárgalo abajo.")

        if "excel_riesgos_ia" in st.session_state:
            with col_e2:
                st.download_button(
                    "⬇️ Descargar Matriz de Riesgos IA (Excel)",
                    data=st.session_state["excel_riesgos_ia"],
                    file_name=f"Matriz_Riesgos_IA_{ahora_bogota().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_riesgos",
                )
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Threat Intelligence — Ransomware Live ────────────────────────────
if opcion_menu == "🔴 Threat Intelligence":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 🔴 Threat Intelligence — Ransomware Live")
    st.markdown("""
    **Alimenta tu matriz de riesgos con ataques reales.**  
    Cada víctima de ransomware se traduce a un riesgo en formato ISO 27001 5x5,
    con los controles **ISO 27002:2022** que la habrían evitado.
    
    Fuente: [ransomware.live](https://www.ransomware.live/) — 29.520+ víctimas documentadas, 357 grupos rastreados.
    """)
    
    # ── Stats rápidos ──
    try:
        from ingestor.ransomware_feed import consultar_estadisticas as ce
        stats = ce()
        if stats:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("🌍 Víctimas totales", f"{stats.get('victims',0):,}")
            c2.metric("👥 Grupos activos", stats.get('groups',0))
            c3.metric("📰 Cobertura prensa", stats.get('press',0))
            c4.metric("🕐 Última actualización", "Hoy")
    except:
        pass
    
    st.markdown("---")
    
    # ── Filtros ──
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    with col_f1:
        solo_co = st.checkbox("🇨🇴 Solo Colombia", value=True, key="ransom_co")
    with col_f2:
        sectores_opts = ["Todos", "Servicios Financieros", "Salud", "Manufactura",
                         "Tecnología", "Gobierno", "Energía", "Comercio", "Educación",
                         "Transporte", "Legal", "Construcción", "Agricultura y Alimentos"]
        sector_filtro = st.selectbox("Sector", sectores_opts, key="ransom_sector")
    with col_f3:
        grupo_buscar = st.text_input("Grupo específico (opcional)", placeholder="qilin, lockbit...", key="ransom_grupo")
    with col_f4:
        cantidad = st.number_input("Resultados", 1, 50, 10, key="ransom_cant")
    
    # ── Botón consultar ──
    if st.button("🔄 Consultar amenazas actuales", type="primary", use_container_width=True):
        with st.spinner("🔍 Consultando ransomware.live..."):
            import ingestor.ransomware_feed as rf
            
            if grupo_buscar:
                info_grupo = rf.consultar_grupo(grupo_buscar)
                if info_grupo:
                    st.success(f"✅ Inteligencia del grupo **{grupo_buscar}** obtenida")
                    with st.expander("📋 Detalles del grupo", expanded=True):
                        st.json(info_grupo)
                else:
                    st.warning(f"No se encontró información para el grupo '{grupo_buscar}'")
            
            mapa_sectores = {
                "Servicios Financieros": "Financial Services",
                "Salud": "Healthcare", "Manufactura": "Manufacturing",
                "Tecnología": "Technology", "Gobierno": "Government",
                "Energía": "Energy", "Comercio": "Retail", "Educación": "Education",
                "Transporte": "Transportation", "Legal": "Legal",
                "Construcción": "Construction", "Agricultura y Alimentos": "Agriculture and Food Production",
            }
            sector_api = mapa_sectores.get(sector_filtro, sector_filtro)
            
            if solo_co and sector_api and sector_filtro != "Todos":
                # Ambos filtros: país + sector (usar consultar_por_pais y filtrar por sector en la API)
                # La API soporta múltiples filtros: /victims/?country=CO&sector=Healthcare
                import urllib.parse
                url = f"{rf.API_BASE}/victims/?country=CO&sector={urllib.parse.quote(sector_api)}"
                req = urllib.request.Request(url, headers=rf.API_HEADERS)
                try:
                    resp = urllib.request.urlopen(req, timeout=15)
                    data = json.loads(resp.read().decode())
                    victimas = data.get("victims", [])
                    ataques = []
                    for item in victimas[:cantidad]:
                        ataques.append({
                            "grupo": item.get("group", "Desconocido"),
                            "victima": item.get("victim", ""),
                            "pais": "CO",
                            "sector": item.get("activity", "N/A"),
                            "fecha": item.get("discovered", "")[:10],
                            "website": item.get("website", ""),
                            "data_size": str(item.get("data_size") or ""),
                            "descripcion": item.get("description", "")[:200],
                        })
                except:
                    ataques = []
            elif solo_co:
                ataques = rf.consultar_por_pais("CO", cantidad)
            elif sector_filtro != "Todos" and sector_api:
                ataques = rf.consultar_por_sector(sector_api, cantidad)
            else:
                ataques = rf.consultar_victimas_recientes(cantidad)
            
            if not ataques:
                st.info("No se encontraron víctimas con esos filtros.")
            else:
                solo_sector = None if sector_filtro == "Todos" else sector_api
                riesgos = rf.generar_matriz_riesgos(ataques, solo_colombia=False, solo_sector=solo_sector)
                st.session_state["riesgos_ransom"] = riesgos
                st.session_state["sector_filtro_ransom"] = sector_filtro
                st.session_state["controles_ransom"] = rf.sugerir_controles_prioritarios(
                    sector_api if sector_filtro != "Todos" else "Financial Services"
                )
                st.rerun()
    
    # ── Mostrar resultados si hay datos en sesión ──
    if "riesgos_ransom" in st.session_state:
        riesgos = st.session_state["riesgos_ransom"]
        sector_filtro_mostrar = st.session_state.get("sector_filtro_ransom", "Todos")
        controles_rec = st.session_state.get("controles_ransom", [])
        from ingestor.ransomware_feed import estadisticas_por_sector as es
        
        st.success(f"✅ {len(riesgos)} ataques encontrados")
        
        st.markdown("#### 📋 Víctimas recientes")
        df_data = []
        for r in riesgos:
            df_data.append({
                "Grupo": r["grupo"], "Víctima": r["victima"], "País": r["pais"],
                "Sector": r["sector"], "Fecha": r["fecha"],
                "Nivel Riesgo": r["nivel_riesgo"], "Score": f'{r["score"]}/25',
                "Descripción del Ataque": r.get("descripcion_ataque", r.get("descripcion", "No disponible")),
                "Justificación Controles": r.get("justificacion", "No disponible"),
            })
        st.dataframe(pd.DataFrame(df_data), use_container_width=True, hide_index=True, column_config={
            "Descripción del Ataque": st.column_config.TextColumn(width="large"),
            "Justificación Controles": st.column_config.TextColumn(width="large"),
        })
        
        st.markdown("#### 📊 Afectación por sector")
        stats_sector = es(riesgos)
        if stats_sector:
            cols = st.columns(min(len(stats_sector), 4))
            for i, (sector, data) in enumerate(sorted(stats_sector.items(), key=lambda x: -x[1]["conteo"])):
                with cols[i % 4]:
                    color = "🔴" if data["nivel_promedio"] == "EXTREMO" else "🟠" if data["nivel_promedio"] == "ALTO" else "🟡"
                    st.metric(f"{color} {sector}", f'{data["conteo"]} ataques', data["nivel_promedio"])
        
        st.markdown("#### 🛡️ Controles ISO 27002 recomendados")
        if controles_rec:
            cols_iso = st.columns(len(controles_rec))
            for i, c in enumerate(controles_rec):
                with cols_iso[i]:
                    st.markdown(f"""
                    <div style="background:rgba(88,200,242,0.08);border:1px solid rgba(88,200,242,0.15);
                                border-radius:8px;padding:12px;text-align:center;">
                        <div style="color:#58C8F2;font-size:20px;font-weight:800;">{c['codigo']}</div>
                        <div style="color:rgba(255,255,255,0.6);font-size:11px;margin-top:4px;">{c['nombre']}</div>
                    </div>
                    """, unsafe_allow_html=True)
        
        # ── Botón para generar y descargar Excel ──
        st.markdown("---")
        
        col_ex1, col_ex2 = st.columns([2, 1])
        with col_ex1:
            if st.button("📥 Generar Excel de Riesgos (con datos)", type="primary", key="btn_gen_excel"):
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from datetime import datetime
                import tempfile, os
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Riesgos Ransomware"
                
                hf = PatternFill(start_color="1B273D", end_color="1B273D", fill_type="solid")
                hfn = Font(color="FFFFFF", bold=True, size=11)
                bd = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                            top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
                fe = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
                fa = PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid")
                fm = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                
                headers = ["Grupo", "Víctima", "País", "Sector", "Fecha", "Probabilidad", "Severidad", "Score", "Nivel Riesgo", "Descripción del Ataque", "Justificación de Controles ISO 27002"]
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.fill = hf; c.font = hfn; c.alignment = Alignment(horizontal='center'); c.border = bd
                
                for row, r in enumerate(riesgos, 2):
                    dr = [r["grupo"], r["victima"], r["pais"], r["sector"], r["fecha"],
                          r["probabilidad"], r["severidad"], f'{r["score"]}/25', r["nivel_riesgo"],
                          r.get("descripcion_ataque", r.get("descripcion", "No disponible")),
                          r.get("justificacion", "No disponible")]
                    for col, val in enumerate(dr, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = bd
                        if col == 9:
                            if val == "EXTREMO": c.fill = fe; c.font = Font(bold=True, color="FFFFFF")
                            elif val == "ALTO": c.fill = fa; c.font = Font(bold=True)
                            elif val == "MEDIO": c.fill = fm
                
                for col in range(1, len(headers)+1):
                    ws.column_dimensions[chr(64+col) if col <= 26 else 'Z'].width = 20
                
                ws2 = wb.create_sheet("Controles ISO 27002")
                for col, h in enumerate(["Código", "Nombre del Control", "Sector Recomendado"], 1):
                    c = ws2.cell(row=1, column=col, value=h)
                    c.fill = hf; c.font = hfn; c.border = bd
                for i, ctrl in enumerate(controles_rec, 2):
                    ws2.cell(row=i, column=1, value=ctrl["codigo"]).border = bd
                    ws2.cell(row=i, column=2, value=ctrl["nombre"]).border = bd
                    ws2.cell(row=i, column=3, value=sector_filtro_mostrar).border = bd
                
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                wb.save(tmp.name)
                tmp.close()
                with open(tmp.name, 'rb') as f:
                    st.session_state["excel_datos"] = f.read()
                os.unlink(tmp.name)
                st.rerun()
        
        with col_ex2:
            if "excel_datos" in st.session_state:
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=st.session_state["excel_datos"],
                    file_name=f"Threat_Intelligence_Ransomware_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_ransom_excel"
                )
            else:
                st.info("Pulsa 'Generar Excel'\ny luego descarga aquí")
    
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Autodiagnóstico de Madurez ───────────────────────────────────────
if opcion_menu == "🔬 Autodiagnóstico":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    from autodiagnostico import mostrar_autodiagnostico
    mostrar_autodiagnostico()

    # Manejar generación con IA
    if st.session_state.get("generar_auto_propuesta", False):
        st.session_state["generar_auto_propuesta"] = False
        prompt = st.session_state.get("auto_prompt", "")
        if prompt:
            from langchain_core.messages import SystemMessage, HumanMessage
            with st.spinner("🤖 Generando propuesta personalizada..."):
                try:
                    llm_local_auto = None
                    if proveedor == "DeepSeek":
                        api_key_l = os.getenv("DEEPSEEK_API_KEY")
                        if api_key_l:
                            llm_local_auto = ChatOpenAI(model=modelo, temperature=temperatura, api_key=api_key_l, base_url="https://api.deepseek.com/v1")
                    elif proveedor == "Groq (Gratis)":
                        api_key_l = os.getenv("GROQ_API_KEY")
                        if api_key_l:
                            llm_local_auto = ChatOpenAI(model=modelo, temperature=temperatura, api_key=api_key_l, base_url="https://api.groq.com/openai/v1")
                    elif proveedor == "Google (Gemini)":
                        api_key_l = os.getenv("GOOGLE_API_KEY")
                        if api_key_l:
                            llm_local_auto = ChatGoogleGenerativeAI(model=modelo, temperature=temperatura, google_api_key=api_key_l)
                    else:
                        api_key_l = os.getenv("OPENAI_API_KEY")
                        if api_key_l:
                            llm_local_auto = ChatOpenAI(model=modelo, temperature=temperatura, api_key=api_key_l)

                    if llm_local_auto:
                        respuesta = llm_local_auto.invoke([
                            SystemMessage(content="Eres un consultor senior en ciberseguridad e IA para infraestructura crítica. Especialista en ISO 27001, ISO 42001, NIST AI RMF, ISO 19011 e ISO 23894."),
                            HumanMessage(content=prompt)
                        ])
                        st.session_state.auto_propuesta = respuesta.content
                    else:
                        st.error("No hay conexión con el proveedor de IA. Verifica la configuración.")
                except Exception as e:
                    st.error(f"Error generando propuesta: {e}")
            st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Dashboard ─────────────────────────────────────────────────────────────
if opcion_menu == "Dashboard":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 📊 Dashboard Consolidado 360° (Cumplimiento, Madurez y Riesgos)")
    
    sg = stats_globales()
    sv = stats_vulnerabilidades()
    total_r = sg.get("total_reportes") or 0
    total_v = sv.get("total") or 0
    
    nivel_madurez = "No evaluado"
    pct_madurez = 0.0
    if st.session_state.get("auto_resultados"):
        nivel_madurez = st.session_state["auto_resultados"].get("nivel_global", "No evaluado")
        pct_madurez = st.session_state["auto_resultados"].get("porcentaje_global", 0.0)

    c1, c2, c3, c4 = st.columns(4)
    def metric_html(val, label, color="#a78bfa"):
        return f"<div class='metric-box'><div class='metric-value' style='color:{color}'>{val}</div><div class='metric-label'>{label}</div></div>"
    
    c1.markdown(metric_html(total_r, "Políticas Auditadas", "#a78bfa"), unsafe_allow_html=True)
    c2.markdown(metric_html(f"{(sg.get('avg_cumplimiento') or 0.0):.1f}%", "Cumplimiento Promedio (Gap)", "#34d399"), unsafe_allow_html=True)
    c3.markdown(metric_html(f"{pct_madurez:.1f}%" if pct_madurez > 0 else "N/A", f"Madurez: {nivel_madurez}", "#60a5fa"), unsafe_allow_html=True)
    c4.markdown(metric_html(total_v, "Vulnerabilidades IA Activas", "#fb923c" if total_v > 0 else "#34d399"), unsafe_allow_html=True)

    st.markdown("---")
    col_g1, col_g2 = st.columns(2)
    
    with col_g1:
        if total_r == 0:
            st.info("Aún no hay análisis de confluencia. Sube un documento en 'Auditoría de IA' para ver gráficos.")
        else:
            conformes_t    = sg.get("total_conformes", 0) or 0
            parciales_t    = sg.get("total_parciales", 0) or 0
            no_conformes_t = sg.get("total_no_conformes", 0) or 0
            fig_donut = go.Figure(go.Pie(
                labels=["✅ Conformes", "⚠️ Parciales", "❌ No Conformes"],
                values=[conformes_t, parciales_t, no_conformes_t],
                hole=0.55,
                marker_colors=["#059669", "#d97706", "#dc2626"],
                textfont_size=12,
            ))
            fig_donut.update_layout(
                title="Distribución global de controles (Auditoría)",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font_color="white", showlegend=True,
                legend=dict(orientation="h", y=-0.15),
                height=320,
            )
            st.plotly_chart(fig_donut, use_container_width=True)

    with col_g2:
        if total_v == 0:
            st.info("No hay vulnerabilidades IA en la base de datos. Actualiza la base en 'Matriz Riesgos IA' para ver estadísticas.")
        else:
            extremos = sv.get("extremos") or 0
            altos = sv.get("altos") or 0
            medios = sv.get("medios") or 0
            bajos = sv.get("bajos") or 0
            
            fig_bar_vuln = go.Figure(go.Bar(
                x=["Extremo", "Alto", "Medio", "Bajo"],
                y=[extremos, altos, medios, bajos],
                marker_color=["#ef4444", "#f97316", "#eab308", "#10b981"],
                text=[str(v) for v in [extremos, altos, medios, bajos]],
                textposition="outside",
            ))
            fig_bar_vuln.update_layout(
                title="Vulnerabilidades de IA por Nivel de Riesgo",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font_color="white", yaxis_range=[0, max(extremos, altos, medios, bajos) * 1.2 + 2],
                height=320,
            )
            st.plotly_chart(fig_bar_vuln, use_container_width=True)

    st.markdown("---")
    col_g3, col_g4 = st.columns(2)
    
    with col_g3:
        if total_r > 0:
            reportes = listar_reportes(10)
            if reportes:
                reportes_cron = sorted(reportes, key=lambda x: x.get("created_at", ""))
                fechas = [r["created_at"][:16].replace("T", " ") for r in reportes_cron]
                pcts = [r["cumplimiento_pct"] for r in reportes_cron]
                nombres = [r["nombre_doc"][:20] for r in reportes_cron]
                
                fig_line = go.Figure()
                fig_line.add_trace(go.Scatter(
                    x=fechas,
                    y=pcts,
                    mode='lines+markers',
                    name='Cumplimiento',
                    line=dict(color='#8b5cf6', width=3),
                    marker=dict(size=8, color='#34d399', symbol='circle'),
                    text=[f"{n} ({p}%)" for n, p in zip(nombres, pcts)],
                    hoverinfo='text+x'
                ))
                fig_line.update_layout(
                    title="Tendencia Temporal de Cumplimiento (Últimos 10 Análisis)",
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="white", yaxis_range=[0, 110],
                    xaxis_tickangle=-30, height=320,
                    margin=dict(l=40, r=40, t=40, b=40)
                )
                st.plotly_chart(fig_line, use_container_width=True)
                
    with col_g4:
        if pct_madurez > 0:
            res = st.session_state["auto_resultados"]
            fig_radar = go.Figure()
            fig_radar.add_trace(go.Scatterpolar(
                r=[res["dimensiones"][d]["porcentaje"] for d in res["dimensiones"]],
                theta=list(res["dimensiones"].keys()),
                fill='toself',
                name='Madurez IA-SEC',
                line_color='#7c3aed'
            ))
            fig_radar.update_layout(
                polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                showlegend=False,
                paper_bgcolor="rgba(0,0,0,0)",
                font_color="white",
                title="Madurez Organizacional por Dimensión",
                height=320
            )
            st.plotly_chart(fig_radar, use_container_width=True)
        else:
            st.info("Completa el cuestionario en 'Autodiagnóstico' para habilitar el gráfico radar de madurez organizacional.")
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab Historial ─────────────────────────────────────────────────────────────
if opcion_menu == "Historial":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 🗂️ Historial de Reportes")
    
    todos_reportes = listar_reportes(200)
    if not todos_reportes:
        st.info("No hay reportes guardados aún.")
    else:
        col_f1, col_f2, col_f3 = st.columns([2, 1, 1])
        filtro_nombre = col_f1.text_input("🔍 Filtrar por nombre del documento", key="hist_filtro_nombre")
        min_pct, max_pct = col_f2.slider(
            "📈 Porcentaje de cumplimiento",
            0, 100, (0, 100),
            key="hist_filtro_pct"
        )
        
        reportes_filtrados = []
        for r in todos_reportes:
            nombre_doc = r.get("nombre_doc", "")
            pct = r.get("cumplimiento_pct", 0.0)
            if filtro_nombre and filtro_nombre.lower() not in nombre_doc.lower():
                continue
            if not (min_pct <= pct <= max_pct):
                continue
            reportes_filtrados.append(r)
            
        col_f3.markdown(f"<div style='text-align:right;padding-top:25px;font-size:0.9rem;color:rgba(255,255,255,0.6)'>Encontrados: <b>{len(reportes_filtrados)}</b> / {len(todos_reportes)}</div>", unsafe_allow_html=True)
        
        if not reportes_filtrados:
            st.warning("Ningún reporte coincide con los criterios de búsqueda.")
        else:
            items_por_pagina = 10
            total_paginas = ((len(reportes_filtrados) - 1) // items_por_pagina) + 1
            
            if "pagina_actual" not in st.session_state:
                st.session_state["pagina_actual"] = 1
                
            if st.session_state["pagina_actual"] > total_paginas:
                st.session_state["pagina_actual"] = total_paginas
            if st.session_state["pagina_actual"] < 1:
                st.session_state["pagina_actual"] = 1
                
            inicio = (st.session_state["pagina_actual"] - 1) * items_por_pagina
            fin = inicio + items_por_pagina
            
            for r in reportes_filtrados[inicio:fin]:
                ts = r["created_at"][:16].replace("T", " ")
                pct = r["cumplimiento_pct"]
                with st.expander(f"📄 {r['nombre_doc']} — {ts} — {pct}% cumplimiento"):
                    col_a, col_b, col_c = st.columns(3)
                    col_a.metric("✅ Conformes",    r["conformes"])
                    col_b.metric("⚠️ Parciales",    r["parciales"])
                    col_c.metric("❌ No conformes", r["no_conformes"])
                    st.markdown(f"**Modelo:** {r['proveedor']} / {r['modelo']}  |  **Usuario:** {r['usuario']}")
                    st.markdown("---")
                    st.markdown(r["resultado"])
                    col_dl1, col_dl2, col_dl3 = st.columns(3)
                    with col_dl1:
                        st.download_button(
                            "⬇️ .md",
                            data=r["resultado"],
                            file_name=f"reporte_{r['id']}.md",
                            mime="text/markdown",
                            key=f"md_{r['id']}",
                        )
                    with col_dl2:
                        stats_r = {"conformes": r["conformes"], "parciales": r["parciales"],
                                   "no_conformes": r["no_conformes"], "cumplimiento_pct": r["cumplimiento_pct"]}
                        pdf_b = generar_pdf(r["resultado"], r["nombre_doc"], r["usuario"],
                                            r["proveedor"], r["modelo"], stats_r)
                        st.download_button(
                            "📄 Simple",
                            data=pdf_b,
                            file_name=f"reporte_{r['id']}.pdf",
                            mime="application/pdf",
                            key=f"pdf_{r['id']}",
                        )
                    with col_dl3:
                        from pdf_export import generar_informe_auditoria_profesional
                        prof_b = generar_informe_auditoria_profesional(
                            resultado_md=r["resultado"],
                            nombre_empresa=r["nombre_doc"],
                            sector="",
                            tamano_empresa="",
                            nombre_doc=r["nombre_doc"],
                            usuario=r["usuario"],
                            proveedor=r["proveedor"],
                            modelo=r["modelo"],
                            stats=stats_r,
                        )
                        st.download_button(
                            "📄 Profesional",
                            data=prof_b,
                            file_name=f"informe_{r['id']}.pdf",
                            mime="application/pdf",
                            key=f"prof_{r['id']}",
                        )
            
            if total_paginas > 1:
                st.markdown("---")
                col_p1, col_p2, col_p3 = st.columns([1, 2, 1])
                with col_p1:
                    if st.button("⬅️ Anterior", disabled=(st.session_state["pagina_actual"] == 1), key="btn_pag_prev"):
                        st.session_state["pagina_actual"] -= 1
                        st.rerun()
                with col_p2:
                    st.markdown(f"<div class='pag-info'>Página <b>{st.session_state['pagina_actual']}</b> de {total_paginas}</div>", unsafe_allow_html=True)
                with col_p3:
                    if st.button("Siguiente ➡️", disabled=(st.session_state["pagina_actual"] == total_paginas), key="btn_pag_next"):
                        st.session_state["pagina_actual"] += 1
                        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

# ── Tab 4: Logs de Auditoría ─────────────────────────────────────────────────
if opcion_menu == "Logs":
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("#### 🔍 Registro de Eventos — Control A.8.15 ISO/IEC 27001:2022")
    st.markdown("Trazabilidad completa de accesos y análisis realizados en la plataforma.")

    col_r, col_d = st.columns([1, 1])
    with col_r:
        if st.button("🔄 Refrescar", key="btn_refresh_logs"):
            st.rerun()
    with col_d:
        entries = get_log_entries(200)
        if entries:
            import json as _json
            raw = "\n".join([_json.dumps(e, ensure_ascii=False) for e in entries])
            st.download_button("⬇️ Exportar Logs (.jsonl)", data=raw, file_name="audit_log.jsonl", mime="application/json")

    entries = get_log_entries(100)
    if not entries:
        st.info("No hay eventos registrados aún.")
    else:
        st.markdown(f"**{len(entries)} eventos registrados**")
        tabla_logs = "<table class='result-table'><tr><th>Timestamp (UTC)</th><th>Evento</th><th>Usuario</th><th>Resultado</th><th>Detalle</th></tr>"
        for e in reversed(entries):
            ts = e.get('timestamp','')[:19].replace('T',' ')
            evento = e.get('event', '')
            usr = e.get('user', '')
            res = e.get('result', '')
            det = e.get('detail', '')
            color = '#34d399' if res == 'success' else '#f87171'
            badge = f"<span style='color:{color};font-weight:600'>{res.upper()}</span>"
            tabla_logs += f"<tr><td style='font-size:0.8rem;color:#94a3b8'>{ts}</td><td><b>{evento}</b></td><td>{usr}</td><td>{badge}</td><td style='font-size:0.82rem'>{det}</td></tr>"
        tabla_logs += "</table>"
        st.markdown(tabla_logs, unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

# ── Footer CID INFOSEC (visible en todas las pantallas) ────────────────────
st.markdown(f"""
<div class='cid-footer'>
  <strong>CID INFOSEC</strong> v2.0 — Plataforma de Auditoría de Ciberseguridad e Inteligencia Artificial<br>
  &copy; {datetime.now().year} &middot; Banco Agrario de Colombia &middot; ISO 27001 &middot; ISO 42001 &middot; NIST AI RMF &middot; ISO 19011 &middot; ISO 23894<br>
  <span style='font-size:0.68rem;'>Confidencial &middot; Uso interno autorizado</span>
</div>
""", unsafe_allow_html=True)
