import json
from dataclasses import dataclass, field
from typing import Optional, Dict
import os
from langchain_openai import ChatOpenAI
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage, SystemMessage


# ══════════════════════════════════════════════════════════════════════════════
# DEFINICIONES DE LA MATRIZ DE INTEROPERABILIDAD
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ControlInteroperable:
    """Representa un punto de confluencia entre los tres marcos normativos."""
    id: str
    dimension: str
    iso_27001_ref: str
    iso_27001_desc: str
    iso_42001_ref: str
    iso_42001_desc: str
    nist_ai_rmf_ref: str
    nist_ai_rmf_desc: str
    relevancia_ic: str  # Alta, Muy Alta, Crítica
    requisito_ic: str   # Requisito específico para infraestructura crítica
    riesgo_sin_control: str


# ── Cargar Matriz desde JSON ─────────────────────────────────────────────────

def cargar_matriz() -> list[ControlInteroperable]:
    ruta_json = os.path.join(os.path.dirname(__file__), "mapping_normativo.json")
    try:
        with open(ruta_json, "r", encoding="utf-8") as f:
            datos = json.load(f)
            return [ControlInteroperable(**d) for d in datos]
    except Exception as e:
        print(f"Error cargando matriz: {e}")
        return []

MATRIZ_INTEROPERABILIDAD = cargar_matriz()

# ── Lógica de Cumplimiento Cuantitativo ──────────────────────────────────────

def calcular_cumplimiento(estado: str) -> float:
    """
    Transforma estados cualitativos en valores cuantitativos (0.0 a 1.0).
    """
    equivalencias = {
        "No iniciado": 0.0,
        "En proceso inicial": 0.25,
        "Implementado parcialmente": 0.50,
        "Mayormente implementado": 0.75,
        "Cumplimiento total": 1.0,
        "No evaluado": 0.0
    }
    return equivalencias.get(estado, 0.0)


# ── Sectores de Infraestructura Crítica ──────────────────────────────────────

SECTORES_IC = {
    "energia": {
        "nombre": "Energía y Utilities",
        "descripcion": "Generación, transmisión y distribución de energía eléctrica, gas natural, petróleo",
        "regulaciones_adicionales": ["NERC CIP", "IEC 62351", "IEC 62443"],
        "activos_criticos": [
            "SCADA/ICS", "PLCs", "RTUs", "Subestaciones eléctricas",
            "Redes de transmisión", "Sistemas de medición inteligente (AMI)",
            "Modelos predictivos de carga (IA)", "Turbinas con control autónomo"
        ],
        "riesgos_especificos": [
            "Sabotaje de sistemas SCADA mediante IA adversarial",
            "Manipulación de modelos de predicción de demanda",
            "Ataques a firmware de medidores inteligentes",
            "Falsos positivos en detección de anomalías que causen desconexión",
            "Envenenamiento de datos de sensores IoT industriales"
        ],
    },
    "salud": {
        "nombre": "Salud y Ciencias de la Vida",
        "descripcion": "Hospitales, sistemas de salud pública, dispositivos médicos, investigación farmacéutica",
        "regulaciones_adicionales": ["HIPAA", "FDA 21 CFR Part 11", "MDR (EU)"],
        "activos_criticos": [
            "Sistemas de Historia Clínica Electrónica (HCE)",
            "Dispositivos médicos conectados (IoMT)",
            "Modelos de IA para diagnóstico (radiología, patología)",
            "Robots quirúrgicos", "Sistemas de dosificación automatizada",
            "Plataformas de telemedicina", "Bases de datos genómicas"
        ],
        "riesgos_especificos": [
            "Sesgo en modelos de IA diagnóstica que afecte poblaciones vulnerables",
            "Manipulación de imágenes médicas para alterar diagnósticos",
            "Ataques a sistemas de dosificación automatizada",
            "Fuga de datos genómicos y de salud mental",
            "Fallo de IA en triage que retrase atención crítica"
        ],
    },
    "finanzas": {
        "nombre": "Servicios Financieros",
        "descripcion": "Bancos, seguros, mercados de valores, sistemas de pago, fintech",
        "regulaciones_adicionales": ["PCI DSS v4.0", "SOX", "Basel III/IV", "DORA (EU)"],
        "activos_criticos": [
            "Core bancario", "Sistemas de trading algorítmico",
            "Modelos de scoring crediticio (IA)", "Sistemas de detección de fraude",
            "Pasarelas de pago", "Infraestructura SWIFT/SPEI",
            "Modelos de riesgo de mercado", "KYC/AML automatizado"
        ],
        "riesgos_especificos": [
            "Manipulación de modelos de trading para flash crashes",
            "Sesgo discriminatorio en scoring crediticio",
            "Evasión de modelos AML/detección de fraude",
            "Ataques adversariales a sistemas de reconocimiento facial",
            "Deepfakes para suplantación en procesos KYC"
        ],
    },
    "transporte": {
        "nombre": "Transporte y Logística",
        "descripcion": "Aviación, ferrocarriles, transporte marítimo, logística, vehículos autónomos",
        "regulaciones_adicionales": ["DO-326A (aviación)", "TSA Cybersecurity Directive", "IMO MSC-FAL.1"],
        "activos_criticos": [
            "Sistemas de gestión de tráfico aéreo (ATM)",
            "Señalización ferroviaria (ERTMS/ETCS)",
            "Vehículos autónomos (niveles 3-5)",
            "Sistemas de navegación GPS/GNSS",
            "Plataformas de logística predictiva",
            "Drones de carga", "Sistemas portuarios automatizados"
        ],
        "riesgos_especificos": [
            "Spoofing GPS en vehículos autónomos o drones",
            "Ataques a señalización ferroviaria que provoquen colisiones",
            "Envenenamiento de datos de tráfico para desviar rutas",
            "Interferencia con sensores LiDAR mediante perturbaciones adversariales",
            "Manipulación de modelos de mantenimiento predictivo"
        ],
    },
    "gobierno": {
        "nombre": "Gobierno y Defensa",
        "descripcion": "Servicios gubernamentales, defensa nacional, inteligencia, infraestructura electoral",
        "regulaciones_adicionales": ["FedRAMP", "CMMC", "NIST SP 800-53", "Common Criteria"],
        "activos_criticos": [
            "Sistemas de identificación ciudadana",
            "Infraestructura electoral", "Sistemas de vigilancia (IA)",
            "Redes clasificadas", "Sistemas de mando y control (C2)",
            "Modelos de inteligencia predictiva",
            "Plataformas de servicios digitales ciudadanos"
        ],
        "riesgos_especificos": [
            "Desinformación generada por IA para desestabilización",
            "Manipulación de sistemas electorales automatizados",
            "Sesgo en vigilancia automatizada contra minorías",
            "Ataques a sistemas de identificación biométrica",
            "Exfiltración de modelos de inteligencia clasificados"
        ],
    },
    "telecomunicaciones": {
        "nombre": "Telecomunicaciones",
        "descripcion": "Redes móviles, ISP, infraestructura de internet, satélites, centros de datos",
        "regulaciones_adicionales": ["3GPP Security", "ETSI NFV SEC", "ITU-T X.805"],
        "activos_criticos": [
            "Redes 5G/6G core", "Sistemas de gestión de red (NMS)",
            "CDNs y DNS raíz", "Cables submarinos",
            "Estaciones base", "Centros de datos Tier III/IV",
            "IA para optimización de red", "Sistemas anti-DDoS"
        ],
        "riesgos_especificos": [
            "Ataques a infraestructura 5G Open RAN",
            "Manipulación de IA de optimización de red para degradar servicio",
            "Backdoors en equipos de red con firmware IA",
            "Ataques a sistemas de detección de anomalías de tráfico",
            "Suplantación de estaciones base mediante SDR"
        ],
    },
    "agua": {
        "nombre": "Agua y Saneamiento",
        "descripcion": "Plantas de tratamiento, distribución de agua potable, gestión de aguas residuales",
        "regulaciones_adicionales": ["AWIA (America's Water Infrastructure Act)", "EPA Guidelines"],
        "activos_criticos": [
            "Sistemas SCADA de plantas de tratamiento",
            "Sensores de calidad de agua (IoT)",
            "Sistemas de dosificación química automatizada",
            "Modelos predictivos de demanda hídrica",
            "Infraestructura de bombeo", "EDAR automatizadas"
        ],
        "riesgos_especificos": [
            "Manipulación de dosificación química (ej. caso Oldsmar 2021)",
            "Envenenamiento de sensores IoT de calidad de agua",
            "Ataques a modelos predictivos de demanda para causar escasez",
            "Alteración de datos de monitoreo ambiental",
            "Sabotaje de sistemas de alerta temprana de contaminación"
        ],
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE ANÁLISIS Y GENERACIÓN DE PROPUESTAS
# ══════════════════════════════════════════════════════════════════════════════

def obtener_dimensiones() -> list[str]:
    """Retorna la lista de dimensiones únicas de la matriz."""
    return list(dict.fromkeys(c.dimension for c in MATRIZ_INTEROPERABILIDAD))


def obtener_controles_por_dimension(dimension: str) -> list[ControlInteroperable]:
    """Retorna los controles de una dimensión específica."""
    return [c for c in MATRIZ_INTEROPERABILIDAD if c.dimension == dimension]


def calcular_cobertura(controles_evaluados: dict[str, str]) -> dict:
    """
    Calcula el porcentaje de cobertura por dimensión considerando el peso de cada estado.
    """
    dimensiones = obtener_dimensiones()
    cobertura = {}
    
    for dim in dimensiones:
        controles_dim = obtener_controles_por_dimension(dim)
        total_posible = len(controles_dim)  # Peso máximo (1.0 por control)
        acumulado = 0.0
        
        for c in controles_dim:
            estado = controles_evaluados.get(c.id, "No evaluado")
            acumulado += calcular_cumplimiento(estado)
            
        cobertura[dim] = {
            "total": total_posible,
            "implementado_valor": acumulado,
            "porcentaje": round((acumulado / total_posible * 100), 1) if total_posible > 0 else 0,
        }
    
    return cobertura


def obtener_sector(sector_id: str) -> Optional[dict]:
    """Retorna la información de un sector de IC."""
    return SECTORES_IC.get(sector_id)


def generar_resumen_interoperabilidad() -> dict:
    """Genera estadísticas resumen de la matriz completa."""
    total = len(MATRIZ_INTEROPERABILIDAD)
    por_relevancia = {"Crítica": 0, "Muy Alta": 0, "Alta": 0}
    por_dimension = {}
    
    for c in MATRIZ_INTEROPERABILIDAD:
        por_relevancia[c.relevancia_ic] = por_relevancia.get(c.relevancia_ic, 0) + 1
        por_dimension[c.dimension] = por_dimension.get(c.dimension, 0) + 1
    
    return {
        "total_controles": total,
        "por_relevancia": por_relevancia,
        "por_dimension": por_dimension,
        "dimensiones": obtener_dimensiones(),
    }


def generar_df_interoperabilidad() -> 'pd.DataFrame':
    """Retorna la matriz completa como un DataFrame de Pandas."""
    import pandas as pd
    data = []
    for c in MATRIZ_INTEROPERABILIDAD:
        data.append({
            "ID": c.id,
            "Dimensión": c.dimension,
            "ISO 27001": c.iso_27001_ref,
            "ISO 42001": c.iso_42001_ref,
            "NIST AI RMF": c.nist_ai_rmf_ref,
            "Relevancia IC": c.relevancia_ic,
            "Estado Actual": "No evaluado",
            "Plan de Acción": ""
        })
    return pd.DataFrame(data)


# ══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE PROPUESTAS CON LLM
# ══════════════════════════════════════════════════════════════════════════════

def generar_propuesta_ic(
    sector_id: str,
    controles_evaluados: dict[str, str],
    contexto_organizacion: str,
    proveedor: str,
    modelo_llm: str,
    temperatura: float,
) -> str:
    """
    Genera una propuesta de implementación personalizada para una organización
    de infraestructura crítica usando el LLM seleccionado.
    """
    sector = SECTORES_IC.get(sector_id)
    if not sector:
        return "ERROR: Sector no válido"
    
    # Construir el contexto de la evaluación basado en el cumplimiento cuantitativo
    brechas = []
    parciales = []
    implementados = []
    
    for control_id, estado in controles_evaluados.items():
        control = next((c for c in MATRIZ_INTEROPERABILIDAD if c.id == control_id), None)
        if not control:
            continue
        
        valor = calcular_cumplimiento(estado)
        if valor <= 0.25:
            brechas.append((control, estado))
        elif valor < 1.0:
            parciales.append((control, estado))
        else:
            implementados.append((control, estado))
    
    # Formatear brechas para el prompt
    brechas_texto = ""
    for c, est in brechas:
        brechas_texto += f"""
- **{c.id} ({c.dimension})**: {c.requisito_ic}
  - Estado actual: {est}
  - Riesgo: {c.riesgo_sin_control}
"""
    
    parciales_texto = ""
    for c, est in parciales:
        parciales_texto += f"""
- **{c.id} ({c.dimension})**: {c.requisito_ic}
  - Estado actual: {est}
"""

    # ── RAG: Buscar contexto normativo relevante ──
    contexto_rag_texto = ""
    try:
        from core.rag_engine import get_rag
        engine = get_rag()
        if engine.is_active:
            consulta = f"{sector['nombre']} {contexto_organizacion[:300]} controles de seguridad implementación"
            ctx = engine.get_context(consulta, top_k=3)
            if ctx:
                contexto_rag_texto = f"\n\nContexto normativo relevante:\n{ctx}\n"
    except Exception:
        pass

    # Construir el LLM
    if proveedor == "Google (Gemini)":
        api_key = os.getenv("GOOGLE_API_KEY")
        if not api_key:
            return "ERROR: No se encontró GOOGLE_API_KEY en el archivo .env"
        llm = ChatGoogleGenerativeAI(model=modelo_llm, temperature=temperatura, google_api_key=api_key)
    elif proveedor == "DeepSeek":
        api_key = os.getenv("DEEPSEEK_API_KEY")
        if not api_key:
            return "ERROR: No se encontró DEEPSEEK_API_KEY en el archivo .env"
        llm = ChatOpenAI(model=modelo_llm, temperature=temperatura, api_key=api_key, base_url="https://api.deepseek.com/v1")
    else:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return "ERROR: No se encontró OPENAI_API_KEY en el archivo .env"
        llm = ChatOpenAI(model=modelo_llm, temperature=temperatura, api_key=api_key)

    system_prompt = f"""Eres un consultor senior experto en ciberseguridad e inteligencia artificial para infraestructuras críticas.
Tu dominio incluye ISO/IEC 27001:2022, NIST AI RMF 1.0, ISO/IEC 42001:2023, 
ISO 19011:2018 (directrices para auditoría de sistemas de gestión) e ISO/IEC 23894:2023 (gestión de riesgos de IA),
así como regulaciones sectoriales específicas.

Tu tarea es generar una PROPUESTA DE IMPLEMENTACIÓN detallada y accionable para una organización del sector de infraestructura crítica: **{sector['nombre']}**.

Regulaciones sectoriales adicionales relevantes: {', '.join(sector['regulaciones_adicionales'])}

Activos críticos del sector: {', '.join(sector['activos_criticos'][:5])}

Riesgos específicos del sector:
{chr(10).join('- ' + r for r in sector['riesgos_especificos'][:4])}

INSTRUCCIONES — La propuesta DEBE estructurarse siguiendo la metodología de auditoría ISO 19011:2018:

**Fase 1 — Iniciación (ISO 19011 Cláusula 6.2):**
- Determinar la viabilidad de la auditoría basada en las brechas identificadas
- Definir el alcance del programa de implementación

**Fase 2 — Preparación (ISO 19011 Cláusula 6.3):**
- Revisar la documentación existente y las brechas detectadas
- Elaborar el plan de trabajo con recursos y cronograma

**Fase 3 — Ejecución (ISO 19011 Cláusula 6.4):**
- Implementar los controles priorizados
- Recolectar evidencia de conformidad
- Incluir el catálogo de riesgos de IA según ISO 23894:2023

**Fase 4 — Informe (ISO 19011 Cláusula 6.5):**
- Documentar los resultados de la implementación
- Emitir el informe de conformidad y el estado de riesgos residuales

**Fase 5 — Seguimiento (ISO 19011 Cláusula 6.6):**
- Planificar verificaciones periódicas
- Establecer indicadores de monitoreo continuo

Estructura requerida de la propuesta:
1. **Resumen Ejecutivo**: Estado actual, nivel de riesgo y urgencia, basado en las brechas evaluadas.
2. **Metodología**: ISO 19011:2018 como marco de auditoría con las 5 fases.
3. **Catálogo de Riesgos de IA (ISO 23894:2023)**: Identificar los riesgos específicos del sector (opacidad algorítmica, deriva del modelo, vulnerabilidad adversarial, sesgo, privacidad, impacto social, seguridad física).
4. **Hoja de Ruta de Implementación por Fases** (Fase 1: 0-3m, Fase 2: 3-6m, Fase 3: 6-12m).
5. **Controles Prioritarios**: Justificación de urgencia alineada con la criticidad y el nivel de riesgo.
6. **Arquitectura de Seguridad de IA recomendada**.
7. **Métricas y KPIs** para monitoreo continuo (ISO 23894 Cláusula 6.2).
8. **Quick Wins** (2 semanas).
9. **Plan de Seguimiento**: Próximas auditorías, frecuencia de revisión, indicadores de mejora.

La propuesta debe ser específica para {sector['nombre']} y alineada con los cinco marcos normativos. Formato: Markdown estructurado con emojis."""

    user_prompt = f"""## Contexto de la Organización
{contexto_organizacion}

## Evaluación de Interoperabilidad — Resultados

### Brechas Críticas (No iniciado / Proceso inicial): {len(brechas)}
{brechas_texto if brechas_texto else "Ninguna crítica."}

### Avances Parciales: {len(parciales)}
{parciales_texto if parciales_texto else "Ninguno."}

### Implementados totalmente: {len(implementados)}

### Catálogo de Riesgos de IA (ISO 23894:2023) a considerar en la propuesta:
- **Opacidad algorítmica**: Imposibilidad de auditar la lógica interna del modelo
- **Deriva del modelo (Model Drift)**: Degradación silenciosa del rendimiento
- **Vulnerabilidad adversarial**: Manipulación de entradas para inducir errores
- **Sesgo algorítmico**: Resultados desfavorables para ciertos grupos
- **Privacidad**: Fuga de datos sensibles desde los modelos
- **Impacto social**: Efectos en confianza pública y equidad
- **Seguridad física**: Daños por fallos en IA en entornos críticos
{contexto_rag_texto}
Genera la propuesta de implementación completa siguiendo la metodología ISO 19011:2018. Fundamenta tus recomendaciones en el contexto normativo disponible."""

    mensajes = [
        SystemMessage(content=system_prompt),
        HumanMessage(content=user_prompt),
    ]
    
    respuesta = llm.invoke(mensajes)
    return respuesta.content


def generar_markdown_matriz(sector_id: Optional[str] = None) -> str:
    """Genera la matriz completa en formato Markdown para exportación."""
    sector = SECTORES_IC.get(sector_id) if sector_id else None
    
    md = "# Matriz de Interoperabilidad Normativa\n"
    md += "## ISO/IEC 27001:2022 + ISO/IEC 42001:2023 + NIST AI RMF 1.0\n"
    md += "### Enfoque: Infraestructuras Críticas\n\n"
    
    if sector:
        md += f"**Sector:** {sector['nombre']}\n"
    
    md += "---\n\n"
    
    for dimension in obtener_dimensiones():
        controles = obtener_controles_por_dimension(dimension)
        md += f"## {dimension}\n\n"
        md += "| ID | ISO 27001 | ISO 42001 | NIST AI RMF | Relevancia IC | Requisito IC |\n"
        md += "|:---|:----------|:----------|:------------|:--------------|:-------------|\n"
        
        for c in controles:
            md += f"| {c.id} | {c.iso_27001_ref} | {c.iso_42001_ref} | {c.nist_ai_rmf_ref} | {c.relevancia_ic} | {c.requisito_ic[:80]}... |\n"
        
        md += "\n"
    
    return md


def exportar_matriz_excel(df_evaluaciones: 'pd.DataFrame') -> bytes:
    """
    Genera un archivo Excel profesional con 5 pestañas:
      1. Confluencia IA-SEC   — 3 normas lado a lado (23 controles)
      2. ISO 27001            — Gap Analysis por control
      3. ISO 42001            — Gap Analysis por control
      4. NIST AI RMF          — Gap Analysis por control
      5. Evaluación Riesgos   — Matriz PxI (ISO 31000/27005/23894)
    Formato profesional estilo GAP.xlsx con semáforos de cumplimiento y
    escala de madurez MSPI (Inexistente → Optimizado).
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # IDs de TODOS los 23 controles del mapping
    confluencia_controles = [c.id for c in MATRIZ_INTEROPERABILIDAD]

    escala_valores = {
        "No evaluado": {"pct": 0.0, "text": "Inexistente", "bg": "FFC7CE", "fg": "9C0006"},
        "No iniciado": {"pct": 0.0, "text": "Inexistente", "bg": "FFC7CE", "fg": "9C0006"},
        "En proceso inicial": {"pct": 0.20, "text": "Inicial", "bg": "FFD966", "fg": "7F6000"},
        "Implementado parcialmente": {"pct": 0.50, "text": "Repetible", "bg": "F4B084", "fg": "C65911"},
        "Mayormente implementado": {"pct": 0.80, "text": "Gestionado", "bg": "E2EFDA", "fg": "375623"},
        "Cumplimiento total": {"pct": 1.0, "text": "Optimizado", "bg": "C6EFCE", "fg": "006100"}
    }

    normas = {
        "ISO 27001": {
            "title": "Análisis de Brechas (GAP Analysis) — ISO/IEC 27001:2022",
            "ref_attr": "iso_27001_ref",
            "desc_attr": "iso_27001_desc",
            "color_tema": "1F4E78"
        },
        "ISO 42001": {
            "title": "Análisis de Brechas (GAP Analysis) — ISO/IEC 42001:2023",
            "ref_attr": "iso_42001_ref",
            "desc_attr": "iso_42001_desc",
            "color_tema": "7C3AED"
        },
        "NIST AI RMF": {
            "title": "Análisis de Brechas (GAP Analysis) — NIST AI RMF 1.0",
            "ref_attr": "nist_ai_rmf_ref",
            "desc_attr": "nist_ai_rmf_desc",
            "color_tema": "2E7D32"
        }
    }

    font_family = "Arial"
    medium_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for tab_name, config in normas.items():
        ws = wb.create_sheet(title=tab_name)
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells("B1:H1")
        ws["B1"] = config["title"]
        ws["B1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
        ws["B1"].fill = PatternFill(start_color=config["color_tema"], end_color=config["color_tema"], fill_type="solid")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.merge_cells("B2:H2")
        ws["B2"] = "Framework IA-SEC · Matriz de Cumplimiento Integrada (23 controles)"
        ws["B2"].font = Font(name=font_family, size=11, italic=True, color="555555")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 15

        headers = [
            ("B", "Control"),
            ("C", "Descripción del control"),
            ("D", "Estado inicial"),
            ("E", "Nivel de cumplimiento %"),
            ("F", "Plan de acción"),
            ("G", "% de avance alcanzado"),
            ("H", "Qué voy a hacer para lograrlo ?")
        ]

        for col, text in headers:
            cell = ws[col + "4"]
            cell.value = text
            cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=config["color_tema"], end_color=config["color_tema"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = medium_border
        ws.row_dimensions[4].height = 32

        current_row = 5
        for cid in confluencia_controles:
            control_obj = next((c for c in MATRIZ_INTEROPERABILIDAD if c.id == cid), None)
            if not control_obj:
                continue

            eval_row = df_evaluaciones[df_evaluaciones['ID'] == cid]
            estado_actual = "No evaluado"
            plan_accion = ""
            if not eval_row.empty:
                estado_actual = eval_row['Estado Actual'].values[0]
                plan_accion = eval_row['Plan de Acción'].values[0] or ""

            val_data = escala_valores.get(estado_actual, escala_valores["No evaluado"])

            ws[f"B{current_row}"] = getattr(control_obj, config["ref_attr"])
            ws[f"C{current_row}"] = getattr(control_obj, config["desc_attr"])
            ws[f"D{current_row}"] = control_obj.riesgo_sin_control
            ws[f"E{current_row}"] = val_data["pct"]
            ws[f"E{current_row}"].number_format = '0%'
            ws[f"F{current_row}"] = plan_accion if plan_accion else "Documentar e implementar el control según los requisitos del marco normativo."
            ws[f"G{current_row}"] = f"{int(val_data['pct']*100)}% - {val_data['text']}"
            ws[f"H{current_row}"] = control_obj.requisito_ic

            for col, _ in headers:
                cell = ws[f"{col}{current_row}"]
                cell.font = Font(name=font_family, size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = thin_border

            sem_cell = ws[f"G{current_row}"]
            sem_cell.alignment = Alignment(horizontal="center", vertical="center")
            sem_cell.font = Font(name=font_family, size=10, bold=True, color=val_data["fg"])
            sem_cell.fill = PatternFill(start_color=val_data["bg"], end_color=val_data["bg"], fill_type="solid")

            ws[f"E{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
            current_row += 1

        end_row = current_row - 1
        for r in range(5, end_row + 1):
            ws.row_dimensions[r].height = 70

        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 48

        eq_row = end_row + 2
        ws.merge_cells(f"B{eq_row}:D{eq_row}")
        ws[f"B{eq_row}"] = "Tabla de Escala de Valoración IA-SEC (Basada en MSPI MinTIC Colombia)"
        ws[f"B{eq_row}"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        ws[f"B{eq_row}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        ws[f"B{eq_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[eq_row].height = 25

        ws[f"B{eq_row+1}"] = "Calificación"
        ws[f"C{eq_row+1}"] = "Descripción"
        ws[f"D{eq_row+1}"] = "Criterio de Evaluación"
        ws.row_dimensions[eq_row+1].height = 20

        for col in ["B", "C", "D"]:
            cell = ws[f"{col}{eq_row+1}"]
            cell.font = Font(name=font_family, size=10, bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center" if col != "D" else "left", vertical="center")
            cell.border = thin_border

        escala_filas = [
            ("0%", "Inexistente", "Total falta de cualquier proceso reconocible. La organización ni siquiera ha reconocido que hay un problema.", "FFC7CE", "9C0006"),
            ("20%", "Inicial", "Existe evidencia de reconocimiento del problema, sin procesos estandarizados. Depende de cada individuo y es reactivo.", "FFD966", "7F6000"),
            ("40%", "Repetible", "Los procesos siguen un patrón regular, sin comunicación ni formación formal. Alta dependencia de personas clave.", "F4B084", "C65911"),
            ("60%", "Definido", "Los procesos se documentan y comunican. Los controles son eficientes y se aplican casi siempre de forma coherente.", "FFF2CC", "8C6B00"),
            ("80%", "Gestionado", "Los controles se monitorean, miden y evalúan de forma continua. Se pueden tomar medidas de corrección ágiles.", "E2EFDA", "375623"),
            ("100%", "Optimizado", "Las buenas prácticas se automatizan por completo. Los procesos se redefinen basándose en mejora continua.", "C6EFCE", "006100")
        ]

        curr_eq_row = eq_row + 2
        for cal, desc, crit, bg, fg in escala_filas:
            ws[f"B{curr_eq_row}"] = cal
            ws[f"C{curr_eq_row}"] = desc
            ws[f"D{curr_eq_row}"] = crit
            ws[f"B{curr_eq_row}"].font = Font(name=font_family, size=9, bold=True, color=fg)
            ws[f"B{curr_eq_row}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            ws[f"B{curr_eq_row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"C{curr_eq_row}"].font = Font(name=font_family, size=9, bold=True, color=fg)
            ws[f"C{curr_eq_row}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            ws[f"C{curr_eq_row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"D{curr_eq_row}"].font = Font(name=font_family, size=9)
            ws[f"D{curr_eq_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for col in ["B", "C", "D"]:
                ws[f"{col}{curr_eq_row}"].border = thin_border
            ws.row_dimensions[curr_eq_row].height = 35
            curr_eq_row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑA CONFLUENCIA IA-SEC (3 normas integradas, 23 controles)
    # ══════════════════════════════════════════════════════════════════════════
    ws_conf = wb.create_sheet(title="Confluencia IA-SEC", index=0)

    ws_conf.merge_cells("A1:L1")
    ws_conf["A1"] = "Matriz de Confluencia Normativa IA-SEC — ISO 27001 · ISO 42001 · NIST AI RMF"
    ws_conf["A1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    ws_conf["A1"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    ws_conf["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_conf.row_dimensions[1].height = 40

    ws_conf.merge_cells("A2:L2")
    ws_conf["A2"] = "Framework IA-SEC V6 · Basado en NIST AI RMF 1.0 + ISO/IEC 42001:2023 + ISO/IEC 27001:2022"
    ws_conf["A2"].font = Font(name=font_family, size=11, italic=True, color="BBBBBB")
    ws_conf["A2"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    ws_conf["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_conf.row_dimensions[2].height = 25
    ws_conf.merge_cells("A3:L3")
    ws_conf["A3"] = "Herramienta de apoyo al auditor · 23 controles correlacionados · 9 dimensiones operativas"
    ws_conf["A3"].font = Font(name=font_family, size=10, italic=True, color="888888")
    ws_conf["A3"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    ws_conf["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws_conf.row_dimensions[3].height = 22

    conf_headers = [
        ("A", "ID"),
        ("B", "Dimensión"),
        ("C", "ISO 27001\nReferencia"),
        ("D", "ISO 27001\nDescripción"),
        ("E", "ISO 42001\nReferencia"),
        ("F", "ISO 42001\nDescripción"),
        ("G", "NIST AI RMF\nReferencia"),
        ("H", "NIST AI RMF\nDescripción"),
        ("I", "Relevancia\nIC"),
        ("J", "Requisito para Infraestructura Crítica"),
        ("K", "Riesgo sin control"),
        ("L", "Estado")
    ]

    for col, text in conf_headers:
        cell = ws_conf[col + "4"]
        cell.value = text
        cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = medium_border
    ws_conf.row_dimensions[4].height = 45

    col_widths = {"A": 10, "B": 20, "C": 18, "D": 35, "E": 18, "F": 35,
                  "G": 18, "H": 35, "I": 12, "J": 45, "K": 35, "L": 14}
    for col, w in col_widths.items():
        ws_conf.column_dimensions[col].width = w

    color_rel = {"Crítica": "FFC7CE", "Muy Alta": "FFD966", "Alta": "E2EFDA"}
    color_rel_fg = {"Crítica": "9C0006", "Muy Alta": "7F6000", "Alta": "375623"}

    conf_row = 5
    for cid in confluencia_controles:
        ctrl = next((c for c in MATRIZ_INTEROPERABILIDAD if c.id == cid), None)
        if not ctrl:
            continue

        eval_row = df_evaluaciones[df_evaluaciones['ID'] == cid]
        estado_actual = "No evaluado"
        if not eval_row.empty:
            estado_actual = eval_row['Estado Actual'].values[0]
        val_data = escala_valores.get(estado_actual, escala_valores["No evaluado"])

        ws_conf[f"A{conf_row}"] = ctrl.id
        ws_conf[f"B{conf_row}"] = ctrl.dimension
        ws_conf[f"C{conf_row}"] = ctrl.iso_27001_ref
        ws_conf[f"D{conf_row}"] = ctrl.iso_27001_desc
        ws_conf[f"E{conf_row}"] = ctrl.iso_42001_ref
        ws_conf[f"F{conf_row}"] = ctrl.iso_42001_desc
        ws_conf[f"G{conf_row}"] = ctrl.nist_ai_rmf_ref
        ws_conf[f"H{conf_row}"] = ctrl.nist_ai_rmf_desc
        ws_conf[f"I{conf_row}"] = ctrl.relevancia_ic
        ws_conf[f"J{conf_row}"] = ctrl.requisito_ic
        ws_conf[f"K{conf_row}"] = ctrl.riesgo_sin_control
        ws_conf[f"L{conf_row}"] = val_data["text"]

        for col_letter in [c for c, _ in conf_headers]:
            cell = ws_conf[f"{col_letter}{conf_row}"]
            cell.font = Font(name=font_family, size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

        rel = ctrl.relevancia_ic
        bg = color_rel.get(rel, "FFFFFF")
        fg_col = color_rel_fg.get(rel, "000000")
        ws_conf[f"I{conf_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_conf[f"I{conf_row}"].font = Font(name=font_family, size=9, bold=True, color=fg_col)
        ws_conf[f"I{conf_row}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws_conf[f"L{conf_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_conf[f"L{conf_row}"].font = Font(name=font_family, size=9, bold=True, color=val_data["fg"])
        ws_conf[f"L{conf_row}"].fill = PatternFill(start_color=val_data["bg"], end_color=val_data["bg"], fill_type="solid")
        ws_conf.row_dimensions[conf_row].height = 85
        conf_row += 1

    eq_start = conf_row + 1
    ws_conf.merge_cells(f"A{eq_start}:D{eq_start}")
    ws_conf[f"A{eq_start}"] = "Escala de Valoración IA-SEC (MSPI - MinTIC Colombia)"
    ws_conf[f"A{eq_start}"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    ws_conf[f"A{eq_start}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_conf[f"A{eq_start}"].alignment = Alignment(horizontal="center", vertical="center")
    ws_conf.row_dimensions[eq_start].height = 25

    for i, (col, label) in enumerate([("A", "Calificación"), ("B", "Nivel"), ("C", "Descripción")], 0):
        cell = ws_conf[f"{chr(65+i)}{eq_start+1}"]
        cell.value = label
        cell.font = Font(name=font_family, size=10, bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    e_filas = [
        ("0%", "No implementado", "El control no existe ni ha sido considerado", "FFC7CE", "9C0006"),
        ("20%", "Inicial", "Identificado pero no formalizado ni documentado", "FFD966", "7F6000"),
        ("40%", "Repetible", "En proceso de implementación pero aún no efectivo", "F4B084", "C65911"),
        ("60%", "Definido", "Documentado, implementado y conocido por responsables", "FFF2CC", "8C6B00"),
        ("80%", "Gestionado", "Monitoreado y medido periódicamente con indicadores", "E2EFDA", "375623"),
        ("100%", "Optimizado", "Mejora continua basada en evidencia y lecciones aprendidas", "C6EFCE", "006100")
    ]
    erow = eq_start + 2
    for cal, nivel, desc, bg, fg in e_filas:
        ws_conf[f"A{erow}"] = cal
        ws_conf[f"B{erow}"] = nivel
        ws_conf[f"C{erow}"] = desc
        for c in ["A", "B", "C"]:
            ws_conf[f"{c}{erow}"].font = Font(name=font_family, size=9, bold=True, color=fg)
            ws_conf[f"{c}{erow}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            ws_conf[f"{c}{erow}"].alignment = Alignment(horizontal="center" if c != "C" else "left", vertical="center")
            ws_conf[f"{c}{erow}"].border = thin_border
        ws_conf.row_dimensions[erow].height = 25
        erow += 1

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑA EVALUACIÓN DE RIESGOS (ISO 31000 / ISO 27005 / ISO 23894)
    # ══════════════════════════════════════════════════════════════════════════
    ws_risk = wb.create_sheet(title="Evaluación de Riesgos IA-SEC")

    riesgos_fundamentales = [
        ("OPA-01", "Opacidad algorítmica", "Imposibilidad de auditar la lógica interna del modelo de IA ('caja negra')."),
        ("DER-01", "Deriva del modelo (Model Drift)", "Degradación silenciosa del rendimiento del modelo cuando los datos de entrada mutan."),
        ("ADV-01", "Vulnerabilidad adversarial", "Atacantes manipulan entradas del modelo para inducir errores indetectables."),
        ("SES-01", "Sesgo algorítmico", "Resultados sistemáticamente desfavorables para ciertos grupos poblacionales."),
        ("PRI-01", "Fuga de datos sensibles", "El modelo memoriza y reproduce datos sensibles de entrenamiento en sus outputs."),
        ("SOC-01", "Impacto social y reputacional", "El despliegue de IA afecta la confianza pública y la equidad social."),
        ("FIS-01", "Seguridad física y funcional", "Fallos en sistemas de IA pueden causar daños físicos a personas o infraestructura.")
    ]

    ids_risk_ctrl = [c.id for c in MATRIZ_INTEROPERABILIDAD]
    evaluaciones = []
    for rid, rname, rdesc in riesgos_fundamentales:
        for cid in ids_risk_ctrl[:5]:  # top 5 controles
            prob = 3
            imp = 4
            just = f"Riesgo '{rname}' evaluado contra control {cid}. Sin controles adecuados, el impacto potencial en infraestructura crítica es significativo."
            evaluaciones.append((rid, rname, rdesc, cid, prob, imp, just))

    ws_risk.merge_cells("A1:J1")
    ws_risk["A1"] = "Evaluación de Riesgos IA-SEC — Metodología ISO 31000 / ISO 27005 / ISO 23894"
    ws_risk["A1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    ws_risk["A1"].fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
    ws_risk["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_risk.row_dimensions[1].height = 40

    ws_risk.merge_cells("A2:J2")
    ws_risk["A2"] = "Matriz Cualitativa 4×4 (Probabilidad × Impacto) · Riesgo = P × I → Bajo · Medio · Alto · Extremo"
    ws_risk["A2"].font = Font(name=font_family, size=10, italic=True, color="777777")
    ws_risk["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_risk.row_dimensions[2].height = 22

    risk_headers = [
        ("A", "ID\nRiesgo"), ("B", "Riesgo IA"), ("C", "Descripción"),
        ("D", "Control\nIA-SEC"), ("E", "Control\nISO 27001"),
        ("F", "Prob.\n(1-4)"), ("G", "Impacto\n(1-4)"),
        ("H", "Nivel de\nRiesgo"), ("I", "Color"), ("J", "Justificación")
    ]
    for col, text in risk_headers:
        cell = ws_risk[col + "3"]
        cell.value = text
        cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = medium_border
    ws_risk.row_dimensions[3].height = 40

    risk_cw = {"A": 10, "B": 18, "C": 35, "D": 12, "E": 22, "F": 10, "G": 10, "H": 12, "I": 10, "J": 40}
    for col, w in risk_cw.items():
        ws_risk.column_dimensions[col].width = w

    def nivel_riesgo(p, i):
        v = p * i
        if v >= 12:
            return "Extremo", "000000", "FF0000"
        elif v >= 8:
            return "Alto", "000000", "FF6600"
        elif v >= 5:
            return "Medio", "000000", "FFD700"
        else:
            return "Bajo", "000000", "92D050"

    risk_row = 4
    for (rid, rname, rdesc, cid, prob, imp, just) in evaluaciones:
        ctrl = next((c for c in MATRIZ_INTEROPERABILIDAD if c.id == cid), None)
        if not ctrl:
            continue
        nivel, fg_nivel, bg_nivel = nivel_riesgo(prob, imp)
        ws_risk[f"A{risk_row}"] = rid
        ws_risk[f"B{risk_row}"] = rname
        ws_risk[f"C{risk_row}"] = rdesc
        ws_risk[f"D{risk_row}"] = cid
        ws_risk[f"E{risk_row}"] = ctrl.iso_27001_ref
        ws_risk[f"F{risk_row}"] = prob
        ws_risk[f"G{risk_row}"] = imp
        ws_risk[f"H{risk_row}"] = nivel
        ws_risk[f"I{risk_row}"] = f"{prob*imp} pts"
        ws_risk[f"J{risk_row}"] = just

        for cl in ["A","B","C","D","E","F","G","H","I","J"]:
            cell = ws_risk[f"{cl}{risk_row}"]
            cell.font = Font(name=font_family, size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

        for cn in ["H", "I"]:
            cell = ws_risk[f"{cn}{risk_row}"]
            cell.fill = PatternFill(start_color=bg_nivel, end_color=bg_nivel, fill_type="solid")
            cell.font = Font(name=font_family, size=9, bold=True, color=fg_nivel)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for cn in ["F", "G"]:
            ws_risk[f"{cn}{risk_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_risk.row_dimensions[risk_row].height = 50
        risk_row += 1

    # Leyenda matriz de riesgo
    lr = risk_row + 1
    ws_risk.merge_cells(f"A{lr}:E{lr}")
    ws_risk[f"A{lr}"] = "Matriz de Calor — Nivel de Riesgo (Probabilidad × Impacto) · ISO 31000 / ISO 27005"
    ws_risk[f"A{lr}"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    ws_risk[f"A{lr}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_risk[f"A{lr}"].alignment = Alignment(horizontal="center", vertical="center")
    ws_risk.row_dimensions[lr].height = 25

    ws_risk[f"A{lr+1}"] = "Prob \\ Impacto"
    ws_risk[f"A{lr+1}"].font = Font(name=font_family, size=9, bold=True)
    ws_risk[f"A{lr+1}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_risk[f"A{lr+1}"].border = medium_border

    for i, imp_label in enumerate(["Menor (1)", "Moderado (2)", "Grave (3)", "Crítico (4)"]):
        cell = ws_risk[f"{chr(66+i)}{lr+1}"]
        cell.value = imp_label
        cell.font = Font(name=font_family, size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = medium_border

    probs = [("Alta (4)", [4, 8, 12, 16]), ("Media (3)", [3, 6, 9, 12]),
             ("Baja (2)", [2, 4, 6, 8]), ("Muy Baja (1)", [1, 2, 3, 4])]
    for ro, (pl, vals) in enumerate(probs):
        r = lr + 2 + ro
        ws_risk[f"A{r}"] = pl
        ws_risk[f"A{r}"].font = Font(name=font_family, size=9, bold=True)
        ws_risk[f"A{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_risk[f"A{r}"].border = medium_border
        for co, val in enumerate(vals):
            cell = ws_risk[f"{chr(66+co)}{r}"]
            cell.value = val
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if val >= 12:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
            elif val >= 8:
                cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
            elif val >= 5:
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                cell.font = Font(name=font_family, size=9, bold=True)
            else:
                cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                cell.font = Font(name=font_family, size=9, bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
