"""
Genera la Matriz de Cumplimiento Integrada IA-SEC para el Sector Salud
en el mismo formato Excel profesional que arroja el software Auditor ISO.
"""
import json
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RUTA_MAPPING = os.path.join(os.path.dirname(__file__), "mapping_normativo.json")

with open(RUTA_MAPPING, "r", encoding="utf-8") as f:
    MATRIZ = json.load(f)

# 5 controles de confluencia para el sector salud (mapeo directo del Framework IA-SEC V6)
confluencia_controles_salud = [
    {
        "id": "GOB-01",
        "estado_inicial": "La alta dirección no ha establecido una política de IA para los sistemas clínicos. No existe un comité de ética de IA con competencia técnica para evaluar modelos de diagnóstico asistido, triage automatizado o dosificación de medicamentos. No se han definido niveles de autonomía permitidos para IA en decisiones críticas de salud.",
        "plan_accion_base": "Constituir un Comité de Ética de IA Clínica con participación de dirección médica, seguridad de la información y cumplimiento regulatorio. Definir política de IA que establezca niveles de autonomía, umbrales de supervisión humana obligatoria y protocolos de fallback."
    },
    {
        "id": "RIE-01",
        "estado_inicial": "Las evaluaciones de riesgo de seguridad de TI tradicionales no contemplan los riesgos específicos de IA en el entorno clínico: deriva del modelo diagnóstico, sesgo algorítmico en poblaciones vulnerables, ataques adversariales a imágenes médicas o impacto en cascada por fallo de IA en sistemas de soporte vital.",
        "plan_accion_base": "Incorporar la evaluación de riesgos de IA al proceso de gestión de riesgos institucional (ISO 27001 Cláusula 6.1.2), incluyendo: análisis de sesgo demográfico, pruebas de robustez adversarial, y mapeo de impacto en cascada por fallo de IA en servicios críticos de salud."
    },
    {
        "id": "TEC-02",
        "estado_inicial": "Los registros de eventos clínicos automatizados no capturan las decisiones detalladas de los modelos de IA (versión del modelo, parámetros, confianza, entrada exacta). No existe un registro inmutable a prueba de manipulaciones que permita realizar análisis forense post-incidente o auditoría regulatoria.",
        "plan_accion_base": "Implementar logging inmutable (WORM/blockchain) para todas las decisiones de IA clínica: entrada del modelo, salida, versión, confianza y timestamp. Retención mínima de 7 años con capacidad de auditoría forense ante incidentes."
    },
    {
        "id": "HUM-01",
        "estado_inicial": "Los sistemas de IA para diagnóstico y triage operan sin un mecanismo formal de supervisión humana significativa (Human-in-the-Loop). No se ha definido qué decisiones clínicas automatizadas requieren revisión obligatoria por un profesional de la salud antes de ejecutarse.",
        "plan_accion_base": "Implementar Human-in-the-Loop obligatorio para todas las decisiones clínicas irreversibles (cambios de dosificación, diagnósticos de alto riesgo, priorización de urgencias). Definir tiempos de reacción humana validados y mecanismos de override con registro de auditoría."
    },
    {
        "id": "RES-01",
        "estado_inicial": "No existe un plan de continuidad operativa específico para fallos catastróficos de los sistemas de IA clínica. No se ha definido un RTO para la conmutación a modo manual sin IA, ni se realizan simulacros periódicos de operación degradada.",
        "plan_accion_base": "Desarrollar plan de continuidad para sistemas de IA clínica con: RTO < 15 minutos, modo degradado sin IA validado semestralmente mediante simulacros, hot standby de modelos críticos y procedimientos de conmutación manual documentados."
    }
]

escala_valores = {
    "No evaluado": {"pct": 0.0, "text": "Inexistente", "bg": "FFC7CE", "fg": "9C0006"},
    "No iniciado": {"pct": 0.0, "text": "Inexistente", "bg": "FFC7CE", "fg": "9C0006"},
    "En proceso inicial": {"pct": 0.20, "text": "Inicial", "bg": "FFD966", "fg": "7F6000"},
    "Implementado parcialmente": {"pct": 0.50, "text": "Repetible", "bg": "F4B084", "fg": "C65911"},
    "Mayormente implementado": {"pct": 0.80, "text": "Gestionado", "bg": "E2EFDA", "fg": "375623"},
    "Cumplimiento total": {"pct": 1.0, "text": "Optimizado", "bg": "C6EFCE", "fg": "006100"}
}

normas = {
    "ISO 27001": {
        "title": "Análisis de Brechas (GAP Analysis) — ISO/IEC 27001:2022",
        "ref_attr": "iso_27001_ref",
        "desc_attr": "iso_27001_desc",
        "color_tema": "1F4E78"
    },
    "ISO 42001": {
        "title": "Análisis de Brechas (GAP Analysis) — ISO/IEC 42001:2023",
        "ref_attr": "iso_42001_ref",
        "desc_attr": "iso_42001_desc",
        "color_tema": "7C3AED"
    },
    "NIST AI RMF": {
        "title": "Análisis de Brechas (GAP Analysis) — NIST AI RMF 1.0",
        "ref_attr": "nist_ai_rmf_ref",
        "desc_attr": "nist_ai_rmf_desc",
        "color_tema": "2E7D32"
    }
}

font_family = "Arial"
medium_border = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='medium', color='000000')
)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

estados_iniciales = {c["id"]: c["estado_inicial"] for c in confluencia_controles_salud}
planes_base = {c["id"]: c["plan_accion_base"] for c in confluencia_controles_salud}
ids_control = [c["id"] for c in confluencia_controles_salud]

wb = openpyxl.Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

for tab_name, config in normas.items():
    ws = wb.create_sheet(title=tab_name)
    ws.views.sheetView[0].showGridLines = True

    # Título
    ws.merge_cells("B1:H1")
    ws["B1"] = config["title"]
    ws["B1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    ws["B1"].fill = PatternFill(start_color=config["color_tema"], end_color=config["color_tema"], fill_type="solid")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtítulo
    ws.merge_cells("B2:H2")
    ws["B2"] = "Sector: Salud e Infraestructura Crítica de Salud · Framework IA-SEC V6"
    ws["B2"].font = Font(name=font_family, size=11, italic=True, color="555555")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 15

    headers = [
        ("B", "Control"),
        ("C", "Descripción del control"),
        ("D", "Estado inicial"),
        ("E", "Nivel de cumplimiento %"),
        ("F", "Plan de acción"),
        ("G", "% de avance alcanzado"),
        ("H", "Qué voy a hacer para lograrlo ?")
    ]

    for col, text in headers:
        cell = ws[col + "4"]
        cell.value = text
        cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=config["color_tema"], end_color=config["color_tema"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = medium_border
    ws.row_dimensions[4].height = 32

    current_row = 5
    for cid in ids_control:
        control_obj = next((c for c in MATRIZ if c["id"] == cid), None)
        if not control_obj:
            continue

        estado_actual = "No evaluado"
        plan_accion = ""
        val_data = escala_valores.get(estado_actual, escala_valores["No evaluado"])

        ws[f"B{current_row}"] = control_obj[config["ref_attr"]]
        ws[f"C{current_row}"] = control_obj[config["desc_attr"]]
        ws[f"D{current_row}"] = estados_iniciales[cid]
        ws[f"E{current_row}"] = val_data["pct"]
        ws[f"E{current_row}"].number_format = '0%'
        ws[f"F{current_row}"] = plan_accion if plan_accion else planes_base[cid]
        ws[f"G{current_row}"] = f"{int(val_data['pct']*100)}% - {val_data['text']}"
        ws[f"H{current_row}"] = control_obj["requisito_ic"]

        for col, _ in headers:
            cell = ws[f"{col}{current_row}"]
            cell.font = Font(name=font_family, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

        sem_cell = ws[f"G{current_row}"]
        sem_cell.alignment = Alignment(horizontal="center", vertical="center")
        sem_cell.font = Font(name=font_family, size=10, bold=True, color=val_data["fg"])
        sem_cell.fill = PatternFill(start_color=val_data["bg"], end_color=val_data["bg"], fill_type="solid")

        ws[f"E{current_row}"].alignment = Alignment(horizontal="right", vertical="center")

        current_row += 1

    for r in range(5, 10):
        ws.row_dimensions[r].height = 80

    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 42
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 48

    # Tabla de escala al final
    eq_row = 11
    ws.merge_cells(f"B{eq_row}:E{eq_row}")
    ws[f"B{eq_row}"] = "Tabla de Escala de Valoración IA-SEC (Basada en MSPI MinTIC Colombia)"
    ws[f"B{eq_row}"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    ws[f"B{eq_row}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws[f"B{eq_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[eq_row].height = 25

    ws[f"B{eq_row+1}"] = "Calificación"
    ws[f"C{eq_row+1}"] = "Nivel IA-SEC"
    ws[f"D{eq_row+1}"] = "Descripción"
    ws[f"E{eq_row+1}"] = "Valor"
    ws.row_dimensions[eq_row+1].height = 20

    for col in ["B", "C", "D", "E"]:
        cell = ws[f"{col}{eq_row+1}"]
        cell.font = Font(name=font_family, size=10, bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    escala_filas = [
        ("0%", "No implementado", "El control no existe ni ha sido considerado por la organización", "0.0", "FFC7CE", "9C0006"),
        ("20%", "Inicial", "El control está identificado pero no ha sido formalizado ni documentado", "0.25", "FFD966", "7F6000"),
        ("40%", "Repetible", "El control está en proceso de implementación pero aún no es efectivo", "0.50", "F4B084", "C65911"),
        ("60%", "Definido", "El control está documentado, implementado y conocido por los responsables", "0.75", "FFF2CC", "8C6B00"),
        ("80%", "Gestionado", "El control se monitorea y mide periódicamente con indicadores definidos", "0.80", "E2EFDA", "375623"),
        ("100%", "Optimizado", "El control mejora continuamente con base en evidencia y lecciones aprendidas", "1.0", "C6EFCE", "006100")
    ]

    curr_eq_row = eq_row + 2
    for cal, nivel, desc, valor, bg, fg in escala_filas:
        ws[f"B{curr_eq_row}"] = cal
        ws[f"C{curr_eq_row}"] = nivel
        ws[f"D{curr_eq_row}"] = desc
        ws[f"E{curr_eq_row}"] = valor

        for col in ["B", "C"]:
            cell = ws[f"{col}{curr_eq_row}"]
            cell.font = Font(name=font_family, size=9, bold=True, color=fg)
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws[f"D{curr_eq_row}"].font = Font(name=font_family, size=9)
        ws[f"D{curr_eq_row}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws[f"D{curr_eq_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{curr_eq_row}"].border = thin_border

        ws[f"E{curr_eq_row}"].font = Font(name=font_family, size=9, bold=True, color=fg)
        ws[f"E{curr_eq_row}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws[f"E{curr_eq_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"E{curr_eq_row}"].border = thin_border

        ws.row_dimensions[curr_eq_row].height = 35
        curr_eq_row += 1

    # Información adicional: contexto salud
    info_row = curr_eq_row + 1
    ws.merge_cells(f"B{info_row}:E{info_row}")
    ws[f"B{info_row}"] = "Contexto: Sector Salud - Infraestructura Crítica (CONPES 3995)"
    ws[f"B{info_row}"].font = Font(name=font_family, size=10, bold=True, color="1F4E78")
    ws[f"B{info_row}"].alignment = Alignment(horizontal="left", vertical="center")

    riesgos_row = info_row + 1
    riesgos_texto = (
        "Riesgos específicos de IA en salud: (1) Sesgo en modelos diagnósticos que afecte poblaciones vulnerables, "
        "(2) Manipulación adversarial de imágenes médicas para alterar diagnósticos, "
        "(3) Deriva silenciosa de modelos de triage que degrade la calidad de atención, "
        "(4) Fuga de datos genómicos y de salud mental desde modelos de IA, "
        "(5) Fallo de IA en dosificación automatizada con consecuencias para pacientes."
    )
    ws.merge_cells(f"B{riesgos_row}:H{riesgos_row}")
    ws[f"B{riesgos_row}"] = riesgos_texto
    ws[f"B{riesgos_row}"].font = Font(name=font_family, size=9, color="555555")
    ws[f"B{riesgos_row}"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[riesgos_row].height = 45

# ══════════════════════════════════════════════════════════════════════════════
# PESTAÑA ADICIONAL: CONFLUENCIA IA-SEC (3 normas integradas)
# ══════════════════════════════════════════════════════════════════════════════
ws_conf = wb.create_sheet(title="Confluencia IA-SEC", index=0)

# Título principal
ws_conf.merge_cells("A1:L1")
ws_conf["A1"] = "Matriz de Confluencia Normativa IA-SEC — Sector Salud"
ws_conf["A1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
ws_conf["A1"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
ws_conf["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_conf.row_dimensions[1].height = 40

ws_conf.merge_cells("A2:L2")
ws_conf["A2"] = "ISO/IEC 27001:2022  ·  ISO/IEC 42001:2023  ·  NIST AI RMF 1.0  ·  Framework IA-SEC V6"
ws_conf["A2"].font = Font(name=font_family, size=11, italic=True, color="BBBBBB")
ws_conf["A2"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
ws_conf["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_conf.row_dimensions[2].height = 25

ws_conf.merge_cells("A3:L3")
ws_conf["A3"] = "Sector: Salud e Infraestructura Crítica de Salud (CONPES 3995) — Basado en los 5 puntos de correlación del Framework IA-SEC"
ws_conf["A3"].font = Font(name=font_family, size=10, italic=True, color="888888")
ws_conf["A3"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
ws_conf["A3"].alignment = Alignment(horizontal="center", vertical="center")
ws_conf.row_dimensions[3].height = 22

# Encabezados de la tabla de confluencia
conf_headers = [
    ("A", "ID Control"),
    ("B", "Dimensión IA-SEC"),
    ("C", "ISO 27001:2022\nReferencia"),
    ("D", "ISO 27001:2022\nDescripción"),
    ("E", "ISO 42001:2023\nReferencia"),
    ("F", "ISO 42001:2023\nDescripción"),
    ("G", "NIST AI RMF 1.0\nReferencia"),
    ("H", "NIST AI RMF 1.0\nDescripción"),
    ("I", "Relevancia\nIC"),
    ("J", "Requisito para Infraestructura Crítica de Salud"),
    ("K", "Riesgo sin control"),
    ("L", "Estado\nActual")
]

for col, text in conf_headers:
    cell = ws_conf[col + "4"]
    cell.value = text
    cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = medium_border
ws_conf.row_dimensions[4].height = 45

# Anchos de columna
col_widths = {"A": 12, "B": 20, "C": 18, "D": 35, "E": 18, "F": 35,
              "G": 18, "H": 35, "I": 12, "J": 45, "K": 35, "L": 14}
for col, w in col_widths.items():
    ws_conf.column_dimensions[col].width = w

# Llenar los 5 controles
color_relevancia = {"Crítica": "FFC7CE", "Muy Alta": "FFD966", "Alta": "E2EFDA"}
color_relevancia_fg = {"Crítica": "9C0006", "Muy Alta": "7F6000", "Alta": "375623"}

conf_row = 5
for cid in ids_control:
    ctrl = next((c for c in MATRIZ if c["id"] == cid), None)
    if not ctrl:
        continue

    val_data = escala_valores["No evaluado"]

    ws_conf[f"A{conf_row}"] = ctrl["id"]
    ws_conf[f"B{conf_row}"] = ctrl["dimension"]
    ws_conf[f"C{conf_row}"] = ctrl["iso_27001_ref"]
    ws_conf[f"D{conf_row}"] = ctrl["iso_27001_desc"]
    ws_conf[f"E{conf_row}"] = ctrl["iso_42001_ref"]
    ws_conf[f"F{conf_row}"] = ctrl["iso_42001_desc"]
    ws_conf[f"G{conf_row}"] = ctrl["nist_ai_rmf_ref"]
    ws_conf[f"H{conf_row}"] = ctrl["nist_ai_rmf_desc"]
    ws_conf[f"I{conf_row}"] = ctrl["relevancia_ic"]
    ws_conf[f"J{conf_row}"] = estados_iniciales[cid]
    ws_conf[f"K{conf_row}"] = ctrl["riesgo_sin_control"]
    ws_conf[f"L{conf_row}"] = val_data["text"]

    for col_letter in [c for c, _ in conf_headers]:
        cell = ws_conf[f"{col_letter}{conf_row}"]
        cell.font = Font(name=font_family, size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border

    # Color por relevancia
    rel = ctrl["relevancia_ic"]
    bg = color_relevancia.get(rel, "FFFFFF")
    fg_col = color_relevancia_fg.get(rel, "000000")
    ws_conf[f"I{conf_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws_conf[f"I{conf_row}"].font = Font(name=font_family, size=9, bold=True, color=fg_col)
    ws_conf[f"I{conf_row}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    ws_conf[f"L{conf_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws_conf[f"L{conf_row}"].font = Font(name=font_family, size=9, bold=True, color=val_data["fg"])
    ws_conf[f"L{conf_row}"].fill = PatternFill(start_color=val_data["bg"], end_color=val_data["bg"], fill_type="solid")

    ws_conf.row_dimensions[conf_row].height = 85
    conf_row += 1

# Escala de valoración en la pestaña de confluencia
eq_start = conf_row + 1
ws_conf.merge_cells(f"A{eq_start}:D{eq_start}")
ws_conf[f"A{eq_start}"] = "Escala de Valoración IA-SEC (Basada en MSPI MinTIC Colombia)"
ws_conf[f"A{eq_start}"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
ws_conf[f"A{eq_start}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
ws_conf[f"A{eq_start}"].alignment = Alignment(horizontal="center", vertical="center")
ws_conf.row_dimensions[eq_start].height = 25

for i, (col, label) in enumerate([("A", "Calificación"), ("B", "Nivel"), ("C", "Descripción")], 0):
    cell = ws_conf[f"{chr(65+i)}{eq_start+1}"]
    cell.value = label
    cell.font = Font(name=font_family, size=10, bold=True)
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

escala_filas_conf = [
    ("0%", "No implementado", "El control no existe ni ha sido considerado", "FFC7CE", "9C0006"),
    ("20%", "Inicial", "Identificado pero no formalizado ni documentado", "FFD966", "7F6000"),
    ("40%", "Repetible", "En proceso de implementación pero aún no efectivo", "F4B084", "C65911"),
    ("60%", "Definido", "Documentado, implementado y conocido por responsables", "FFF2CC", "8C6B00"),
    ("80%", "Gestionado", "Monitoreado y medido periódicamente con indicadores", "E2EFDA", "375623"),
    ("100%", "Optimizado", "Mejora continua basada en evidencia y lecciones aprendidas", "C6EFCE", "006100")
]

erow = eq_start + 2
for cal, nivel, desc, bg, fg in escala_filas_conf:
    ws_conf[f"A{erow}"] = cal
    ws_conf[f"B{erow}"] = nivel
    ws_conf[f"C{erow}"] = desc
    for c in ["A", "B", "C"]:
        ws_conf[f"{c}{erow}"].font = Font(name=font_family, size=9, bold=True, color=fg)
        ws_conf[f"{c}{erow}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws_conf[f"{c}{erow}"].alignment = Alignment(horizontal="center" if c != "C" else "left", vertical="center")
        ws_conf[f"{c}{erow}"].border = thin_border
    ws_conf.row_dimensions[erow].height = 25
    erow += 1

# ══════════════════════════════════════════════════════════════════════════════
# PESTAÑA: EVALUACIÓN DE RIESGOS IA-SEC (ISO 31000 / ISO 27005)
# ══════════════════════════════════════════════════════════════════════════════
ws_risk = wb.create_sheet(title="Evaluación de Riesgos IA-SEC")

# Metodología: ISO 31000 / ISO 27005 — Matriz Cualitativa 4×4
# Probabilidad: Muy Baja(1), Baja(2), Media(3), Alta(4)
# Impacto: Menor(1), Moderado(2), Grave(3), Crítico(4)
# Riesgo = Probabilidad × Impacto → Bajo(1-4), Medio(5-7), Alto(8-11), Extremo(12-16)

riesgos_fundamentales_ia = [
    {
        "id": "OPA-01",
        "riesgo": "Opacidad algorítmica",
        "descripcion": "Imposibilidad de auditar la lógica interna del modelo de IA ('caja negra'). Pérdida de trazabilidad sobre el proceso que lleva a una decisión clínica.",
        "origen": "ISO/IEC 42001 A.6.2.4 — Explicabilidad"
    },
    {
        "id": "DER-01",
        "riesgo": "Deriva del modelo (Model Drift)",
        "descripcion": "Degradación silenciosa del rendimiento del modelo cuando el entorno o los datos de entrada mutan. No hay fallos binarios sino deterioro gradual.",
        "origen": "ISO/IEC 42001 A.6 — Monitoreo continuo"
    },
    {
        "id": "ADV-01",
        "riesgo": "Vulnerabilidad adversarial",
        "descripcion": "Atacantes manipulan el modelo inyectando entradas alteradas para inducir errores. Los controles de seguridad tradicionales no detectan la anomalía.",
        "origen": "NIST AI RMF MAP 2.3 — Robustez adversarial"
    },
    {
        "id": "SES-01",
        "riesgo": "Sesgo algorítmico",
        "descripcion": "El modelo produce resultados sistemáticamente desfavorables para ciertos grupos poblacionales debido a datos de entrenamiento no representativos.",
        "origen": "NIST AI RMF MEASURE 2.6-2.11 — Equidad"
    },
    {
        "id": "DAT-01",
        "riesgo": "Fuga de datos sensibles",
        "descripcion": "El modelo memoriza y reproduce datos sensibles de entrenamiento (historias clínicas, datos genómicos) en sus outputs.",
        "origen": "ISO/IEC 27001 A.8.12 — DLP"
    }
]

# Matriz de evaluación: cada riesgo × cada control de confluencia
# Valores: Probabilidad (1-4), Impacto (1-4), Riesgo calculado
# Datos precargados con ejemplo de evaluación para salud

evaluacion_riesgos = [
    # (id_riesgo, id_control, probabilidad, impacto, justificacion)
    ("OPA-01", "GOB-01", 3, 4, "Sin política de IA que exija explicabilidad, la opacidad es alta y su impacto en diagnósticos erróneos es crítico para el paciente."),
    ("OPA-01", "RIE-01", 3, 4, "Las evaluaciones de riesgo tradicionales no incluyen opacidad algorítmica, dejando modelos de 'caja negra' sin supervisión."),
    ("OPA-01", "TEC-02", 2, 4, "Sin registros detallados de decisiones del modelo, es imposible auditar la lógica de un diagnóstico automatizado."),
    ("OPA-01", "HUM-01", 3, 3, "Sin supervisión HITL, decisiones inexplicables del modelo se ejecutan sin revisión humana."),
    ("OPA-01", "RES-01", 2, 3, "En modo degradado la opacidad se reduce (operación manual), pero el riesgo persiste si no hay plan de contingencia."),

    ("DER-01", "GOB-01", 3, 3, "Sin política de monitoreo continuo, la deriva del modelo no se detecta hasta que causa daño."),
    ("DER-01", "RIE-01", 4, 4, "La deriva es el riesgo más desatendido en salud: un modelo preciso al despliegue puede volverse peligroso sin que nadie lo note."),
    ("DER-01", "TEC-02", 4, 4, "Sin logging de inferencia es imposible detectar cuándo el modelo empezó a derivar, impidiendo el análisis forense."),
    ("DER-01", "HUM-01", 3, 4, "Operadores no capacitados para detectar deriva confían ciegamente en outputs degradados del modelo."),
    ("DER-01", "RES-01", 4, 4, "Sin plan de continuidad, la deriva del modelo puede causar parálisis operativa total en el servicio clínico."),

    ("ADV-01", "GOB-01", 3, 4, "La política de IA no exige pruebas de robustez adversarial, dejando modelos clínicos expuestos."),
    ("ADV-01", "RIE-01", 3, 4, "Evaluaciones de riesgo que ignoran ataques adversariales subestiman gravemente la exposición del sistema."),
    ("ADV-01", "TEC-02", 4, 4, "Sin logs inmutables, un ataque adversarial exitoso no deja rastro para investigación forense."),
    ("ADV-01", "HUM-01", 3, 3, "El personal clínico no está entrenado para identificar entradas adversariales manipuladas."),
    ("ADV-01", "RES-01", 3, 4, "La dependencia de IA sin plan B hace que un ataque adversarial exitoso paralice el servicio crítico."),

    ("SES-01", "GOB-01", 4, 4, "Sin política de equidad en IA, el sesgo algorítmico puede discriminar sistemáticamente a poblaciones vulnerables en diagnósticos y triage."),
    ("SES-01", "RIE-01", 3, 4, "Los procesos de evaluación de riesgo estándar no miden sesgo algorítmico ni disparate impact."),
    ("SES-01", "TEC-02", 2, 4, "Sin registros de decisiones segmentados por grupo poblacional, el sesgo es indetectable."),
    ("SES-01", "HUM-01", 3, 4, "Personal no capacitado en detección de sesgo no puede identificar outputs discriminatorios del modelo."),
    ("SES-01", "RES-01", 2, 3, "En operación manual se elimina el sesgo del modelo, pero la transición no está planificada."),

    ("DAT-01", "GOB-01", 3, 4, "Sin política de protección de datos en IA, los modelos pueden memorizar y exponer historias clínicas."),
    ("DAT-01", "RIE-01", 3, 4, "Los análisis de riesgo de datos no cubren la memorización de datos de entrenamiento por parte del modelo."),
    ("DAT-01", "TEC-02", 4, 4, "Sin controles DLP en el pipeline ML, los datos sensibles de pacientes pueden filtrarse a través de outputs del modelo."),
    ("DAT-01", "HUM-01", 2, 4, "El personal clínico puede sin saberlo extraer datos sensibles al interactuar con el modelo (prompt injection)."),
    ("DAT-01", "RES-01", 2, 3, "En contingencia, la prioridad operativa puede relajar controles de privacidad aumentando el riesgo de fuga.")
]

# Título
ws_risk.merge_cells("A1:J1")
ws_risk["A1"] = "Evaluación de Riesgos IA-SEC — Sector Salud"
ws_risk["A1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
ws_risk["A1"].fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
ws_risk["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_risk.row_dimensions[1].height = 40

ws_risk.merge_cells("A2:J2")
ws_risk["A2"] = "Metodología: ISO 31000 / ISO 27005 · Matriz Cualitativa 4×4 (Probabilidad × Impacto) · Riesgo = P × I"
ws_risk["A2"].font = Font(name=font_family, size=10, italic=True, color="777777")
ws_risk["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_risk.row_dimensions[2].height = 22

# Tabla de evaluación
risk_headers = [
    ("A", "ID\nRiesgo"),
    ("B", "Riesgo IA"),
    ("C", "Descripción del riesgo"),
    ("D", "Control\nIA-SEC"),
    ("E", "Control\nISO 27001"),
    ("F", "Probabilidad\n(1-4)"),
    ("G", "Impacto\n(1-4)"),
    ("H", "Nivel de\nRiesgo"),
    ("I", "Color"),
    ("J", "Justificación de la evaluación")
]

for col, text in risk_headers:
    cell = ws_risk[col + "3"]
    cell.value = text
    cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = medium_border
ws_risk.row_dimensions[3].height = 40

risk_col_widths = {"A": 10, "B": 18, "C": 35, "D": 12, "E": 25, "F": 12, "G": 12, "H": 12, "I": 10, "J": 40}
for col, w in risk_col_widths.items():
    ws_risk.column_dimensions[col].width = w

def nivel_riesgo(p, i):
    valor = p * i
    if valor >= 12:
        return "Extremo", "4. Extremo", "000000", "FF0000"
    elif valor >= 8:
        return "Alto", "3. Alto", "000000", "FF6600"
    elif valor >= 5:
        return "Medio", "2. Medio", "000000", "FFD700"
    else:
        return "Bajo", "1. Bajo", "000000", "92D050"

risk_row = 4
for (rid, cid, prob, imp, just) in evaluacion_riesgos:
    riesgo_info = next((r for r in riesgos_fundamentales_ia if r["id"] == rid), None)
    ctrl_info = next((c for c in MATRIZ if c["id"] == cid), None)
    if not riesgo_info or not ctrl_info:
        continue

    nivel, label_nivel, fg_nivel, bg_nivel = nivel_riesgo(prob, imp)

    ws_risk[f"A{risk_row}"] = rid
    ws_risk[f"B{risk_row}"] = riesgo_info["riesgo"]
    ws_risk[f"C{risk_row}"] = riesgo_info["descripcion"]
    ws_risk[f"D{risk_row}"] = cid
    ws_risk[f"E{risk_row}"] = ctrl_info.get("iso_27001_ref", "")
    ws_risk[f"F{risk_row}"] = prob
    ws_risk[f"G{risk_row}"] = imp
    ws_risk[f"H{risk_row}"] = label_nivel
    ws_risk[f"I{risk_row}"] = nivel
    ws_risk[f"J{risk_row}"] = just

    for col_letter in ["A","B","C","D","E","F","G","H","I","J"]:
        cell = ws_risk[f"{col_letter}{risk_row}"]
        cell.font = Font(name=font_family, size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border

    # Color de nivel de riesgo en columna H e I
    for col_nivel in ["H", "I"]:
        cell = ws_risk[f"{col_nivel}{risk_row}"]
        cell.fill = PatternFill(start_color=bg_nivel, end_color=bg_nivel, fill_type="solid")
        cell.font = Font(name=font_family, size=9, bold=True, color=fg_nivel)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Centrar probabilidad e impacto
    for col_num in ["F", "G"]:
        ws_risk[f"{col_num}{risk_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws_risk.row_dimensions[risk_row].height = 55
    risk_row += 1

# Leyenda de la matriz de riesgo
leyenda_row = risk_row + 1
ws_risk.merge_cells(f"A{leyenda_row}:E{leyenda_row}")
ws_risk[f"A{leyenda_row}"] = "Matriz de Calor — Nivel de Riesgo (Probabilidad × Impacto)"
ws_risk[f"A{leyenda_row}"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
ws_risk[f"A{leyenda_row}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
ws_risk[f"A{leyenda_row}"].alignment = Alignment(horizontal="center", vertical="center")
ws_risk.row_dimensions[leyenda_row].height = 25

# Cabecera de la leyenda: fila de encabezados
lr = leyenda_row + 1
ws_risk[f"A{lr}"] = "Probabilidad \\ Impacto"
ws_risk[f"A{lr}"].font = Font(name=font_family, size=9, bold=True)
ws_risk[f"A{lr}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_risk[f"A{lr}"].border = medium_border
ws_risk_pivot_lr = lr

impactos = ["Menor (1)", "Moderado (2)", "Grave (3)", "Crítico (4)"]
for i, imp_label in enumerate(impactos):
    cell = ws_risk[f"{chr(66+i)}{lr}"]
    cell.value = imp_label
    cell.font = Font(name=font_family, size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = medium_border

probabilidades = [
    ("Alta (4)", [4, 8, 12, 16]),
    ("Media (3)", [3, 6, 9, 12]),
    ("Baja (2)", [2, 4, 6, 8]),
    ("Muy Baja (1)", [1, 2, 3, 4])
]

for row_offset, (prob_label, valores) in enumerate(probabilidades):
    r = ws_risk_pivot_lr + 1 + row_offset
    ws_risk[f"A{r}"] = prob_label
    ws_risk[f"A{r}"].font = Font(name=font_family, size=9, bold=True)
    ws_risk[f"A{r}"].alignment = Alignment(horizontal="center", vertical="center")
    ws_risk[f"A{r}"].border = medium_border
    for col_offset, val in enumerate(valores):
        cell = ws_risk[f"{chr(66+col_offset)}{r}"]
        cell.value = val
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if val >= 12:
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
        elif val >= 8:
            cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
        elif val >= 5:
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.font = Font(name=font_family, size=9, bold=True)
        else:
            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            cell.font = Font(name=font_family, size=9, bold=True)

for r in range(ws_risk_pivot_lr, ws_risk_pivot_lr + 5):
    ws_risk.row_dimensions[r].height = 28

# Leyenda de colores
lr2 = lr + 6
ws_risk.merge_cells(f"A{lr2}:E{lr2}")
ws_risk[f"A{lr2}"] = "Escala de Nivel de Riesgo"
ws_risk[f"A{lr2}"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
ws_risk[f"A{lr2}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
ws_risk[f"A{lr2}"].alignment = Alignment(horizontal="center", vertical="center")
ws_risk.row_dimensions[lr2].height = 25

leyenda_riesgos = [
    ("1 - 4", "Bajo", "Riesgo aceptable. Se mantienen controles existentes. No requiere acción inmediata.", "92D050"),
    ("5 - 7", "Medio", "Riesgo tolerable. Requiere monitoreo periódico y acciones de mejora planificadas.", "FFD700"),
    ("8 - 11", "Alto", "Riesgo inaceptable. Requiere acciones correctivas inmediatas con plan de tratamiento documentado.", "FF6600"),
    ("12 - 16", "Extremo", "Riesgo crítico. Requiere acción inmediata, escalamiento a dirección y posible detención del sistema.", "FF0000")
]

for i, (rango, nivel, desc, color) in enumerate(leyenda_riesgos):
    r = lr2 + 1 + i
    ws_risk[f"A{r}"] = rango
    ws_risk[f"B{r}"] = nivel
    ws_risk[f"C{r}"] = desc
    ws_risk.merge_cells(f"C{r}:E{r}")
    for col_let in ["A", "B", "C"]:
        cell = ws_risk[f"{col_let}{r}"]
        cell.font = Font(name=font_family, size=9, bold=True if col_let != "C" else False, color="000000")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_let != "C" else "left", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_risk.row_dimensions[r].height = 30

# Leyenda de riesgos fundamentales
lr3 = lr2 + 6
ws_risk.merge_cells(f"A{lr3}:E{lr3}")
ws_risk[f"A{lr3}"] = "Riesgos Fundamentales de IA (Framework IA-SEC V6)"
ws_risk[f"A{lr3}"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
ws_risk[f"A{lr3}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
ws_risk[f"A{lr3}"].alignment = Alignment(horizontal="center", vertical="center")
ws_risk.row_dimensions[lr3].height = 25

for i, rinfo in enumerate(riesgos_fundamentales_ia):
    r = lr3 + 1 + i
    ws_risk[f"A{r}"] = rinfo["id"]
    ws_risk[f"B{r}"] = rinfo["riesgo"]
    ws_risk[f"C{r}"] = rinfo["descripcion"]
    ws_risk[f"D{r}"] = rinfo["origen"]
    ws_risk.merge_cells(f"C{r}:D{r}")
    for col_let in ["A", "B", "C", "D"]:
        cell = ws_risk[f"{col_let}{r}"]
        cell.font = Font(name=font_family, size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_risk[f"A{r}"].font = Font(name=font_family, size=9, bold=True)
    ws_risk[f"B{r}"].font = Font(name=font_family, size=9, bold=True)
    ws_risk.row_dimensions[r].height = 35

buffer = io.BytesIO()
wb.save(buffer)

output_path = os.path.join(os.path.dirname(__file__), "Matriz_Cumplimiento_IA_SEC_Salud.xlsx")
with open(output_path, "wb") as f:
    f.write(buffer.getvalue())

print(f"Excel generado: {output_path}")
print(f"Tamaño: {len(buffer.getvalue()) / 1024:.1f} KB")
sheets = wb.sheetnames
for s in sheets:
    print(f"  • {s}")
