import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta

BOGOTA_OFFSET = timedelta(hours=-5)

def _ahora_bogota():
    return (datetime.now(timezone.utc) + BOGOTA_OFFSET).replace(tzinfo=None)
from typing import Optional

from .risk_matrix import calcular_riesgo, criticidad_compuesta, COLOR_MAPA, resumen_estadistico


NIVEL_ES = {
    "EXTREMO": "EXTREMO",
    "ALTO": "ALTO",
    "MEDIO": "MEDIO",
    "BAJO": "BAJO",
}

NIVEL_COLOR = {
    "EXTREMO": ("FF0000", "FFFFFF"),
    "ALTO": ("FF6600", "000000"),
    "MEDIO": ("FFD700", "000000"),
    "BAJO": ("92D050", "000000"),
}

PRIORIDAD_ES = {
    "EXTREMO": "Crítica",
    "ALTO": "Alta",
    "MEDIO": "Media",
    "BAJO": "Baja",
}


def exportar_matriz_riesgos_ia(
    vulnerabilidades: list[dict],
    contexto_organizacion: Optional[str] = None,
    respuestas_cuestionario: Optional[dict] = None,
) -> bytes:
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    font_family = "Arial"
    medium_border = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="medium", color="000000"),
        bottom=Side(style="medium", color="000000"),
    )
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    stats = resumen_estadistico(vulnerabilidades)

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑA 1: TABLERO DE CONTROL (Dashboard)
    # ══════════════════════════════════════════════════════════════════════════
    ws_dash = wb.create_sheet(title="Tablero de Control", index=0)

    ws_dash.merge_cells("A1:H1")
    ws_dash["A1"] = "MATRIZ DE RIESGOS DE INTELIGENCIA ARTIFICIAL"
    ws_dash["A1"].font = Font(name=font_family, size=18, bold=True, color="FFFFFF")
    ws_dash["A1"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 45

    ws_dash.merge_cells("A2:H2")
    ws_dash["A2"] = (
        f"Generado: {_ahora_bogota().strftime('%d/%m/%Y %I:%M %p')} Hora Bogotá | "
        f"Total vulnerabilidades activas: {stats['total']} | "
        f"Riesgo promedio: {stats['riesgo_promedio']}/25 | "
        f"Criticidad promedio: {stats['criticidad_promedio']}%"
    )
    ws_dash["A2"].font = Font(name=font_family, size=11, italic=True, color="BBBBBB")
    ws_dash["A2"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    ws_dash["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[2].height = 25

    if contexto_organizacion:
        ws_dash.merge_cells("A3:H3")
        ws_dash["A3"] = f"Contexto: {contexto_organizacion[:200]}"
        ws_dash["A3"].font = Font(name=font_family, size=9, italic=True, color="888888")
        ws_dash["A3"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        ws_dash["A3"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash.row_dimensions[3].height = 20

    row = 5
    ws_dash[f"A{row}"] = "RESUMEN DE RIESGOS DE IA"
    ws_dash[f"A{row}"].font = Font(name=font_family, size=14, bold=True, color="1A1A2E")
    row += 1

    metrics = [
        ("Total Vulnerabilidades", stats["total"], "1A1A2E"),
        ("Riesgo Promedio (1-25)", stats["riesgo_promedio"], "FF6600" if (stats["riesgo_promedio"] or 0) > 10 else "FFD700"),
        ("Criticidad Promedio", f"{stats['criticidad_promedio']}%", "FF0000" if (stats["criticidad_promedio"] or 0) > 60 else "FF6600"),
        ("EXTREMO", stats["por_nivel"].get("EXTREMO", 0), "FF0000"),
        ("ALTO", stats["por_nivel"].get("ALTO", 0), "FF6600"),
        ("MEDIO", stats["por_nivel"].get("MEDIO", 0), "FFD700"),
        ("BAJO", stats["por_nivel"].get("BAJO", 0), "92D050"),
    ]

    for col_offset, (label, value, color) in enumerate(metrics):
        c = col_offset + 1
        if c > 7:
            continue
        cell_label = ws_dash.cell(row=row, column=c)
        cell_label.value = label
        cell_label.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
        cell_label.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_label.border = medium_border

        cell_val = ws_dash.cell(row=row + 1, column=c)
        cell_val.value = value
        cell_val.font = Font(name=font_family, size=16, bold=True, color=color)
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        cell_val.border = medium_border

    row += 3

    ws_dash[f"A{row}"] = "DISTRIBUCIÓN POR FUENTE DE ORIGEN"
    ws_dash[f"A{row}"].font = Font(name=font_family, size=12, bold=True, color="1A1A2E")
    row += 1

    ws_dash[f"A{row}"] = "Fuente"
    ws_dash[f"B{row}"] = "Cantidad"
    ws_dash[f"A{row}"].font = Font(name=font_family, size=10, bold=True)
    ws_dash[f"B{row}"].font = Font(name=font_family, size=10, bold=True)
    ws_dash[f"A{row}"].border = medium_border
    ws_dash[f"B{row}"].border = medium_border
    row += 1

    for fuente, cantidad in sorted(stats["por_fuente"].items(), key=lambda x: -x[1]):
        ws_dash[f"A{row}"] = fuente
        ws_dash[f"B{row}"] = cantidad
        ws_dash[f"A{row}"].font = Font(name=font_family, size=10)
        ws_dash[f"B{row}"].font = Font(name=font_family, size=10)
        ws_dash[f"A{row}"].border = thin_border
        ws_dash[f"B{row}"].border = thin_border
        ws_dash[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    ws_dash[f"A{row}"] = "DISTRIBUCIÓN POR TIPO DE VULNERABILIDAD"
    ws_dash[f"A{row}"].font = Font(name=font_family, size=12, bold=True, color="1A1A2E")
    row += 1

    ws_dash[f"A{row}"] = "Tipo"
    ws_dash[f"B{row}"] = "Cantidad"
    ws_dash[f"A{row}"].font = Font(name=font_family, size=10, bold=True)
    ws_dash[f"B{row}"].font = Font(name=font_family, size=10, bold=True)
    ws_dash[f"A{row}"].border = medium_border
    ws_dash[f"B{row}"].border = medium_border
    row += 1

    for clasif, cantidad in sorted(stats["por_clasificacion"].items(), key=lambda x: -x[1]):
        ws_dash[f"A{row}"] = clasif
        ws_dash[f"B{row}"] = cantidad
        ws_dash[f"A{row}"].font = Font(name=font_family, size=10)
        ws_dash[f"B{row}"].font = Font(name=font_family, size=10)
        ws_dash[f"A{row}"].border = thin_border
        ws_dash[f"B{row}"].border = thin_border
        ws_dash[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

    ws_dash.column_dimensions["A"].width = 35
    ws_dash.column_dimensions["B"].width = 15
    for c in "CDEFGH":
        ws_dash.column_dimensions[c].width = 18

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑA 2: CATÁLOGO DE VULNERABILIDADES
    # ══════════════════════════════════════════════════════════════════════════
    ws_vuln = wb.create_sheet(title="Catálogo de Vulnerabilidades")

    ws_vuln.merge_cells("A1:O1")
    ws_vuln["A1"] = "CATÁLOGO DE VULNERABILIDADES DE INTELIGENCIA ARTIFICIAL"
    ws_vuln["A1"].font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    ws_vuln["A1"].fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
    ws_vuln["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_vuln.row_dimensions[1].height = 40

    vuln_headers = [
        ("A", "ID", 18),
        ("B", "Fuente", 14),
        ("C", "Descripción del Riesgo", 40),
        ("D", "Causa Raíz", 35),
        ("E", "Cómo se Ejecuta el Ataque", 40),
        ("F", "Cómo Prevenirlo", 40),
        ("G", "Solución Recomendada", 40),
        ("H", "Clasificación", 16),
        ("I", "Táctica MITRE", 22),
        ("J", "Probabilidad\n(1-5)", 12),
        ("K", "Severidad\n(1-5)", 12),
        ("L", "Riesgo\n(P×S)", 10),
        ("M", "Nivel", 10),
        ("N", "Normas Aplicables", 40),
        ("O", "Normativa Colombiana", 40),
    ]

    for col_letter, text, width in vuln_headers:
        cell = ws_vuln[f"{col_letter}2"]
        cell.value = text
        cell.font = Font(name=font_family, size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = medium_border
        ws_vuln.column_dimensions[col_letter].width = width
    ws_vuln.row_dimensions[2].height = 35

    row = 3
    for v in vulnerabilidades:
        p = v.get("probabilidad_base", 3)
        s = v.get("severidad_tecnica", 3)
        riesgo = calcular_riesgo(p, s)
        nivel = riesgo["nivel_riesgo"]
        color_ahex = riesgo["color_ahex"]
        nivel_es = NIVEL_ES.get(nivel, nivel)

        ws_vuln[f"A{row}"] = v.get("id_vulnerabilidad", "PENDIENTE")
        ws_vuln[f"B{row}"] = v.get("fuente_origen", "?")
        ws_vuln[f"C{row}"] = v.get("descripcion_riesgo", v.get("resumen_es", ""))
        ws_vuln[f"D{row}"] = v.get("causa_raiz", "No especificada")
        ws_vuln[f"E{row}"] = v.get("como_se_ejecuta", "Consultar la fuente original para detalles técnicos.")
        ws_vuln[f"F{row}"] = v.get("como_prevenirlo", "Aplicar los controles recomendados por las normas aplicables.")
        ws_vuln[f"G{row}"] = v.get("accion_remediacion", "No especificada")
        ws_vuln[f"H{row}"] = v.get("clasificacion_tipo", "Técnico")
        ws_vuln[f"I{row}"] = v.get("tactica_mitre", "")
        ws_vuln[f"J{row}"] = p
        ws_vuln[f"K{row}"] = s
        ws_vuln[f"L{row}"] = riesgo["riesgo_valor"]
        ws_vuln[f"M{row}"] = nivel_es
        normas = v.get("normas_aplicables", [])
        ws_vuln[f"N{row}"] = "; ".join(normas) if isinstance(normas, list) else str(normas)
        norm_col = v.get("normativa_colombia", [])
        ws_vuln[f"O{row}"] = "\n".join(norm_col) if isinstance(norm_col, list) else str(norm_col)

        for col_letter, _, _ in vuln_headers:
            cell = ws_vuln[f"{col_letter}{row}"]
            cell.font = Font(name=font_family, size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

        for cl in ["J", "K", "L"]:
            ws_vuln[cl + str(row)].alignment = Alignment(horizontal="center", vertical="center")

        ws_vuln[f"M{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_vuln[f"M{row}"].font = Font(name=font_family, size=9, bold=True,
                                         color=NIVEL_COLOR.get(nivel, ("000000", "FFFFFF"))[1])
        ws_vuln[f"M{row}"].fill = PatternFill(start_color=color_ahex, end_color=color_ahex, fill_type="solid")

        ws_vuln.row_dimensions[row].height = 65
        row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑA 3: CLASIFICACIÓN POR TÁCTICA
    # ══════════════════════════════════════════════════════════════════════════
    ws_mitre = wb.create_sheet(title="Clasificación por Táctica")

    ws_mitre.merge_cells("A1:E1")
    ws_mitre["A1"] = "CLASIFICACIÓN DE VULNERABILIDADES POR TÁCTICA DE ATAQUE"
    ws_mitre["A1"].font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    ws_mitre["A1"].fill = PatternFill(start_color="2D1B69", end_color="2D1B69", fill_type="solid")
    ws_mitre["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_mitre.row_dimensions[1].height = 35

    mitre_headers = [
        ("A", "Código de Táctica", 22),
        ("B", "Descripción en Español", 45),
        ("C", "Vulnerabilidades\nAsociadas", 16),
        ("D", "Riesgo\nPromedio", 14),
        ("E", "% del\nTotal", 12),
    ]

    for cl, text, w in mitre_headers:
        cell = ws_mitre[f"{cl}2"]
        cell.value = text
        cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2D1B69", end_color="2D1B69", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = medium_border
        ws_mitre.column_dimensions[cl].width = w
    ws_mitre.row_dimensions[2].height = 30

    from .risk_matrix import TACTICAS_MITRE_ATLAS, clasificar_tactica

    tacticas_agrupadas = {}
    for v in vulnerabilidades:
        tact = v.get("tactica_mitre", "AML.PENDIENTE")
        if tact not in tacticas_agrupadas:
            tacticas_agrupadas[tact] = {"count": 0, "sum_riesgo": 0}
        tacticas_agrupadas[tact]["count"] += 1
        p = v.get("probabilidad_base", 3)
        s = v.get("severidad_tecnica", 3)
        if isinstance(p, (int, float)) and isinstance(s, (int, float)):
            tacticas_agrupadas[tact]["sum_riesgo"] += p * s

    total_vuln = len(vulnerabilidades) or 1
    row = 3
    for tactica, data in sorted(tacticas_agrupadas.items(), key=lambda x: -x[1]["count"]):
        desc = clasificar_tactica(tactica)
        prom = round(data["sum_riesgo"] / data["count"], 1) if data["count"] > 0 else 0
        pct = round(data["count"] / total_vuln * 100, 1)

        ws_mitre[f"A{row}"] = tactica
        ws_mitre[f"B{row}"] = desc
        ws_mitre[f"C{row}"] = data["count"]
        ws_mitre[f"D{row}"] = prom
        ws_mitre[f"E{row}"] = f"{pct}%"

        for cl, _, _ in mitre_headers:
            ws_mitre[f"{cl}{row}"].font = Font(name=font_family, size=10)
            ws_mitre[f"{cl}{row}"].alignment = Alignment(vertical="center", wrap_text=True)
            ws_mitre[f"{cl}{row}"].border = thin_border

        for cl in ["C", "D", "E"]:
            ws_mitre[f"{cl}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        ws_mitre.row_dimensions[row].height = 25
        row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑA 4: MAPA DE CALOR (Probabilidad × Severidad)
    # ══════════════════════════════════════════════════════════════════════════
    ws_heat = wb.create_sheet(title="Mapa de Calor P×S")

    ws_heat.merge_cells("A1:G1")
    ws_heat["A1"] = "MAPA DE CALOR — Probabilidad × Severidad (Metodología ISO 31000)"
    ws_heat["A1"].font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    ws_heat["A1"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_heat["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_heat.row_dimensions[1].height = 35

    ws_heat["A3"] = "Probabilidad \\ Severidad"
    ws_heat["A3"].font = Font(name=font_family, size=10, bold=True)
    ws_heat["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_heat["A3"].border = medium_border

    severidad_labels = [
        "Muy Bajo\n(1)",
        "Bajo\n(2)",
        "Medio\n(3)",
        "Alto\n(4)",
        "Muy Alto\n(5)",
    ]
    for i, label in enumerate(severidad_labels):
        cell = ws_heat.cell(row=3, column=i + 2)
        cell.value = label
        cell.font = Font(name=font_family, size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = medium_border

    prob_labels = ["Muy Alta (5)", "Alta (4)", "Media (3)", "Baja (2)", "Muy Baja (1)"]
    prob_values = [5, 4, 3, 2, 1]

    for row_offset, prob_val in enumerate(prob_values):
        r = 4 + row_offset
        ws_heat.cell(row=r, column=1, value=prob_labels[row_offset])
        ws_heat.cell(row=r, column=1).font = Font(name=font_family, size=9, bold=True)
        ws_heat.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_heat.cell(row=r, column=1).border = medium_border

        for sev in range(1, 6):
            riesgo = calcular_riesgo(prob_val, sev)
            cell = ws_heat.cell(row=r, column=sev + 1)
            cell.value = riesgo["riesgo_valor"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.font = Font(name=font_family, size=10, bold=True,
                              color="FFFFFF" if riesgo["nivel_riesgo"] == "EXTREMO" else "000000")
            cell.fill = PatternFill(start_color=riesgo["color_ahex"],
                                    end_color=riesgo["color_ahex"], fill_type="solid")

    leyenda_row = 11
    ws_heat.cell(row=leyenda_row, column=1, value="NIVEL DE RIESGO")
    ws_heat.cell(row=leyenda_row, column=1).font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    ws_heat.cell(row=leyenda_row, column=1).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws_heat.merge_cells(f"A{leyenda_row}:E{leyenda_row}")

    leyenda = [
        ("1 - 5", "BAJO", "Riesgo aceptable. Los controles existentes son suficientes. No requiere acción inmediata.", "92D050"),
        ("6 - 11", "MEDIO", "Riesgo tolerable. Requiere monitoreo periódico y acciones de mejora planificadas.", "FFD700"),
        ("12 - 19", "ALTO", "Riesgo inaceptable. Requiere acciones correctivas inmediatas con plan de tratamiento.", "FF6600"),
        ("20 - 25", "EXTREMO", "Riesgo crítico. Requiere acción inmediata, escalamiento a dirección y posible detener el sistema.", "FF0000"),
    ]
    for i, (rango, nivel, desc, color) in enumerate(leyenda):
        r = leyenda_row + 1 + i
        ws_heat.cell(row=r, column=1, value=rango)
        ws_heat.cell(row=r, column=1).font = Font(name=font_family, size=9, bold=True)
        ws_heat.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_heat.cell(row=r, column=1).border = thin_border

        ws_heat.cell(row=r, column=2, value=nivel)
        ws_heat.cell(row=r, column=2).font = Font(name=font_family, size=9, bold=True)
        ws_heat.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws_heat.cell(row=r, column=2).border = thin_border

        ws_heat.cell(row=r, column=3, value=desc)
        ws_heat.merge_cells(f"C{r}:E{r}")
        ws_heat.cell(row=r, column=3).font = Font(name=font_family, size=9)
        ws_heat.cell(row=r, column=3).alignment = Alignment(vertical="center", wrap_text=True)
        ws_heat.cell(row=r, column=3).border = thin_border

        for ci in range(1, 4):
            ws_heat.cell(row=r, column=ci).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws_heat.row_dimensions[r].height = 30

    ws_heat.column_dimensions["A"].width = 18
    for i in range(2, 7):
        ws_heat.column_dimensions[get_column_letter(i)].width = 16

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑA 5: CONTROLES Y NORMAS APLICABLES
    # ══════════════════════════════════════════════════════════════════════════
    ws_nist = wb.create_sheet(title="Controles y Normas Aplicables")

    ws_nist.merge_cells("A1:G1")
    ws_nist["A1"] = "CONTROLES DE SEGURIDAD Y NORMAS APLICABLES POR VULNERABILIDAD"
    ws_nist["A1"].font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    ws_nist["A1"].fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    ws_nist["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_nist.row_dimensions[1].height = 35

    nist_headers = [
        ("A", "Vulnerabilidad", 22),
        ("B", "Control de Seguridad Recomendado", 40),
        ("C", "Normas Aplicables", 45),
        ("D", "Tipo de Vulnerabilidad", 20),
        ("E", "Prioridad", 12),
        ("F", "Nivel de Riesgo", 12),
        ("G", "Normativa Colombiana", 45),
    ]

    for cl, text, w in nist_headers:
        cell = ws_nist[f"{cl}2"]
        cell.value = text
        cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = medium_border
        ws_nist.column_dimensions[cl].width = w
    ws_nist.row_dimensions[2].height = 30

    row = 3
    for v in vulnerabilidades:
        p = v.get("probabilidad_base", 3)
        s = v.get("severidad_tecnica", 3)
        riesgo = calcular_riesgo(p, s)
        nivel = riesgo["nivel_riesgo"]
        prioridad = PRIORIDAD_ES.get(nivel, "Media")

        ws_nist[f"A{row}"] = v.get("id_vulnerabilidad", "PENDIENTE")
        ws_nist[f"B{row}"] = v.get("controles_nist", v.get("accion_remediacion", "No especificado"))
        normas = v.get("normas_aplicables", [])
        ws_nist[f"C{row}"] = "; ".join(normas) if isinstance(normas, list) else str(normas)
        ws_nist[f"D{row}"] = v.get("clasificacion_tipo", "Técnico")
        ws_nist[f"E{row}"] = prioridad
        ws_nist[f"F{row}"] = NIVEL_ES.get(nivel, nivel)
        norm_col = v.get("normativa_colombia", [])
        ws_nist[f"G{row}"] = "\n".join(norm_col) if isinstance(norm_col, list) else str(norm_col)

        for cl, _, _ in nist_headers:
            cell = ws_nist[f"{cl}{row}"]
            cell.font = Font(name=font_family, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

        for cl in ["E", "F"]:
            ws_nist[f"{cl}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        prior_colors = {"Crítica": ("FF0000", "FFFFFF"), "Alta": ("FF6600", "000000"),
                        "Media": ("FFD700", "000000"), "Baja": ("92D050", "000000")}
        bg, fg = prior_colors.get(prioridad, ("FFFFFF", "000000"))
        ws_nist[f"E{row}"].font = Font(name=font_family, size=10, bold=True, color=fg)
        ws_nist[f"E{row}"].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        nivel_bg, nivel_fg = NIVEL_COLOR.get(nivel, ("FFFFFF", "000000"))
        ws_nist[f"F{row}"].font = Font(name=font_family, size=10, bold=True, color=nivel_fg)
        ws_nist[f"F{row}"].fill = PatternFill(start_color=nivel_bg, end_color=nivel_bg, fill_type="solid")

        ws_nist.row_dimensions[row].height = 45
        row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑA 6: METODOLOGÍA (opcional, informativa)
    # ══════════════════════════════════════════════════════════════════════════
    ws_meta = wb.create_sheet(title="Metodología", index=1)

    ws_meta.merge_cells("A1:F1")
    ws_meta["A1"] = "METODOLOGÍA DE LA MATRIZ DE RIESGOS DE IA"
    ws_meta["A1"].font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    ws_meta["A1"].fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    ws_meta["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_meta.row_dimensions[1].height = 35

    info = [
        ("MARCO NORMATIVO BASE", [
            "ISO/IEC 27001:2022 — Seguridad de la información (controles base: gestión de vulnerabilidades, control de acceso, protección de datos).",
            "NIST AI RMF 1.0 — Gestión de riesgos de IA (funciones: GOVERN, MAP, MEASURE, MANAGE).",
            "ISO/IEC 42001:2023 — Sistema de Gestión de IA (controles específicos para sistemas de IA).",
        ]),
        ("NORMATIVA COLOMBIANA APLICABLE", [
            "Ley 1581 de 2012 — Protección de Datos Personales. Aplica a vulnerabilidades de privacidad en sistemas de IA que procesen datos de colombianos.",
            "Decreto 1377 de 2013 — Reglamentación de la Ley 1581. Autorización del titular y aviso de privacidad en sistemas IA.",
            "Ley 1266 de 2008 — Habeas Data Financiero. Aplica a IA en scoring crediticio y detección de fraude financiero.",
            "CONPES 3995 de 2020 — Política Nacional de Ciberseguridad. Protección de infraestructuras críticas con IA.",
            "CONPES 4069 de 2022 — Política Nacional de Inteligencia Artificial. Eje 2: marco ético. Eje 4: seguridad en IA.",
            "Proyecto Ley 059 de 2023 — Regulación de IA en Colombia (en trámite). Evaluaciones de riesgo obligatorias para IA de alto impacto.",
            "Circular Externa 007 de 2018 — Superintendencia Financiera. Gestión de riesgos ciberseguridad en entidades vigiladas.",
            "Resolución MinTIC 20223040026965 de 2022 — Requisitos mínimos de ciberseguridad para entidades públicas.",
            "Ley 2195 de 2022 — Transparencia y acceso a la información. IA del Estado debe ser explicable y auditable.",
            "Decreto 620 de 2020 — Política de confianza digital y seguridad digital.",
            "Norma Técnica Colombiana NTC-ISO 27001 — Adaptación nacional del estándar ISO 27001.",
        ]),
        ("MAPEO: VULNERABILIDAD ↔ NORMATIVA COLOMBIANA", [
            "Adversarial → Ley 1581/2012 (principio de seguridad), Circular 007/2018 SFC, CONPES 3995/2020.",
            "Privacidad → Ley 1581/2012 (Art. 4, 10, 17), Decreto 1377/2013, CONPES 4069/2022 Eje 2.",
            "Sesgo → Ley 1581/2012 Art. 4 (no discriminación), CONPES 4069/2022 Eje 2, Proyecto Ley 059/2023.",
            "Seguridad Física → CONPES 3995/2020 (infraestructura crítica), Resolución MinTIC 2022.",
            "Datos → Ley 1581/2012, Decreto 1377/2013 Art. 11, CONPES 4069/2022.",
            "Societal → CONPES 4069/2022, Proyecto Ley 059/2023, Ley 1712/2014 transparencia.",
            "Técnico → Resolución MinTIC 2022, CONPES 3995/2020, Decreto 620/2020, NTC-ISO 27001.",
        ]),
        ("FUENTES DE INFORMACIÓN", [
            "MITRE ATLAS — Framework de tácticas y técnicas adversariales contra sistemas de IA (14 tácticas, 100+ técnicas).",
            "AVID (AI Vulnerability Database) — Base de datos de vulnerabilidades de IA documentadas con casos reales.",
            "MIT AI Risk Repository — Taxonomía del MIT con 7 dominios y 24 subdominios de riesgos de IA.",
        ]),
        ("CLASIFICACIÓN DE VULNERABILIDADES", [
            "Adversarial — Ataques que manipulan el sistema de IA para obtener resultados incorrectos.",
            "Sesgo — Discriminación algorítmica contra grupos poblacionales.",
            "Privacidad — Exposición de información sensible a través del modelo de IA.",
            "Seguridad Física — Riesgos que pueden causar daños a personas o infraestructura.",
            "Datos — Contaminación o robo de datos de entrenamiento.",
            "Societal — Impactos negativos en la sociedad (desinformación, vigilancia, desigualdad).",
            "Técnico — Debilidades en la infraestructura del sistema de IA.",
        ]),
        ("CÁLCULO DEL RIESGO (Matriz 5×5)", [
            "Riesgo = Probabilidad (1-5) × Severidad (1-5). Rango: 1 a 25.",
            "BAJO (1-5): Riesgo aceptable. Controles existentes suficientes.",
            "MEDIO (6-11): Riesgo tolerable. Monitoreo y mejora planificada.",
            "ALTO (12-19): Riesgo inaceptable. Acciones correctivas inmediatas.",
            "EXTREMO (20-25): Riesgo crítico. Escalamiento a dirección.",
            "Criticidad = ((Probabilidad + Severidad) / 10) × 100. Escala 0-100%.",
        ]),
        ("NORMAS APLICABLES POR TIPO", [
            "Adversarial → NIST AI RMF 3.0 Robustness Testing, ISO 42001 A.7.3, OWASP LLM Top 10.",
            "Sesgo → NIST AI RMF 2.6 Fairness, ISO 42001 A.6.2, ISO 23894 5.3.",
            "Privacidad → NIST AI RMF 2.0 Privacy, ISO 27001 A.8.12, GDPR Art. 22.",
            "Seguridad Física → NIST AI RMF 3.0 Safety, ISO 23894 5.5, IEC 62443.",
            "Datos → NIST AI RMF 2.0 Data Integrity, ISO 27001 A.8.10, ISO 42001 A.7.1.",
            "Societal → NIST AI RMF 2.0 Transparency, ISO 42001 A.6.1, C2PA Standards.",
        ]),
    ]

    current_row = 3
    for title, items in info:
        ws_meta.merge_cells(f"A{current_row}:F{current_row}")
        ws_meta[f"A{current_row}"] = title
        ws_meta[f"A{current_row}"].font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
        ws_meta[f"A{current_row}"].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        ws_meta[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws_meta.row_dimensions[current_row].height = 28
        current_row += 1

        for item in items:
            ws_meta.merge_cells(f"A{current_row}:F{current_row}")
            ws_meta[f"A{current_row}"] = f"• {item}"
            ws_meta[f"A{current_row}"].font = Font(name=font_family, size=10)
            ws_meta[f"A{current_row}"].alignment = Alignment(vertical="center", wrap_text=True)
            ws_meta.row_dimensions[current_row].height = 22
            current_row += 1

        current_row += 1

    for c in "ABCDEF":
        ws_meta.column_dimensions[c].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
